"""
Sistema TPV para Heladería - Aplicación Principal
Autor: Jesus
Versión: 3.0.0 - ELITE EDITION
Descripción: Aplicación con interfaz ultra moderna, animaciones fluidas y diseño premium
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import sys
import threading
import json
import tempfile
import subprocess
import urllib.request
from datetime import datetime, timedelta
from typing import Optional
import logging

APP_VERSION = "3.0.0"
UPDATE_URL  = "https://tpvelite.surge.sh/version.json"

# Configurar e importar matplotlib
try:
    import matplotlib
    matplotlib.use('TkAgg')  # Usar backend Tkinter
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.warning(f"Matplotlib no disponible: {e}")
    MATPLOTLIB_AVAILABLE = False
    FigureCanvasTkAgg = None
    Figure = None

# Importar módulos propios
from database import DatabaseManager
from services import AuthService, ProductoService, VentaService
from models import Producto, Venta, DetalleVenta, Usuario
from utils import (setup_logging, load_config, save_config, resource_path, format_currency,
                   ensure_directory, get_app_data_dir, Validator, ColorHelper)

# Configurar logging
ensure_directory('logs')
setup_logging()
logger = logging.getLogger(__name__)


def _parse_version(v: str):
    """Convierte '3.1.0' en (3, 1, 0) para comparar."""
    try:
        return tuple(int(x) for x in v.strip().split('.'))
    except Exception:
        return (0,)


def check_for_updates(root: tk.Tk):
    """
    Verifica en segundo plano si hay una versión más nueva disponible.
    Si la hay, muestra un diálogo y descarga+ejecuta el instalador.
    """
    def _worker():
        try:
            with urllib.request.urlopen(UPDATE_URL, timeout=5) as resp:
                data = json.loads(resp.read().decode())

            remote_version  = data.get('version', '0')
            download_url    = data.get('download_url', '')
            changelog       = data.get('changelog', '')

            if _parse_version(remote_version) <= _parse_version(APP_VERSION):
                return  # sin actualizaciones

            # Mostrar diálogo en el hilo principal
            root.after(0, lambda: _prompt_update(root, remote_version, download_url, changelog))

        except Exception as e:
            logger.debug(f"Check de actualización omitido: {e}")

    threading.Thread(target=_worker, daemon=True).start()


def _prompt_update(root: tk.Tk, version: str, url: str, changelog: str):
    """Muestra el diálogo de actualización y, si el usuario acepta, descarga e instala."""
    msg = (
        f"¡Nueva versión disponible! ({version})\n\n"
        f"{changelog}\n\n"
        "¿Querés actualizar ahora?"
    )
    if not messagebox.askyesno("Actualización disponible", msg, parent=root):
        return

    # Ventana de progreso
    prog_win = tk.Toplevel(root)
    prog_win.title("Descargando actualización...")
    prog_win.resizable(False, False)
    prog_win.grab_set()
    w, h = 360, 100
    prog_win.geometry(f"{w}x{h}+{root.winfo_x()+(root.winfo_width()-w)//2}+{root.winfo_y()+(root.winfo_height()-h)//2}")
    tk.Label(prog_win, text="Descargando, por favor esperá...", pady=16).pack()
    bar = ttk.Progressbar(prog_win, mode='indeterminate', length=300)
    bar.pack(pady=4)
    bar.start(10)
    prog_win.update()

    def _download():
        try:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.exe')
            tmp.close()
            urllib.request.urlretrieve(url, tmp.name)
            root.after(0, lambda: _launch_installer(root, prog_win, tmp.name))
        except Exception as e:
            root.after(0, lambda: _download_error(prog_win, str(e)))

    threading.Thread(target=_download, daemon=True).start()


def _launch_installer(root: tk.Tk, prog_win: tk.Toplevel, exe_path: str):
    prog_win.destroy()
    subprocess.Popen([exe_path])
    root.destroy()


def _download_error(prog_win: tk.Toplevel, error: str):
    prog_win.destroy()
    messagebox.showerror("Error de descarga",
        f"No se pudo descargar la actualización:\n{error}\n\nIntentalo más tarde.")


# 🎨 PALETA DE COLORES ELITE - Dark Theme Premium
ELITE_COLORS = {
    # Colores principales - Dark elegant
    'primary_dark': '#0a0e27',          # Negro azulado profundo
    'secondary_dark': '#16213e',        # Azul oscuro elegante
    'accent_purple': '#6c5ce7',         # Púrpura vibrante
    'accent_blue': '#00d4ff',           # Azul cyan brillante
    'accent_pink': '#fd79a8',           # Rosa suave
    'accent_green': '#00b894',          # Verde esmeralda
    'accent_orange': '#ff9f43',         # Naranja cálido
    'accent_red': '#ff6b81',            # Rojo coral
    
    # Fondos y superficies
    'bg_primary': '#0f1419',            # Fondo principal oscuro
    'bg_secondary': '#1a1f2e',          # Fondo secundario
    'bg_card': '#1e2531',               # Fondo de tarjetas
    'bg_hover': '#252d3d',              # Hover state
    'bg_active': '#2d3548',             # Active state
    
    # Textos
    'text_primary': '#e8eaed',          # Texto principal
    'text_secondary': '#9ba3af',        # Texto secundario
    'text_muted': '#6b7280',            # Texto atenuado
    
    # Bordes y divisores
    'border': '#2d3748',                # Bordes sutiles
    'border_light': '#3d4758',          # Bordes claros
    
    # Estados
    'success': '#00e676',               # Verde éxito brillante
    'warning': '#ffd600',               # Amarillo advertencia
    'danger': '#ff1744',                # Rojo peligro
    'info': '#00e5ff',                  # Cyan información
    
    # Gradientes
    'gradient_start': '#667eea',        # Inicio gradiente púrpura-azul
    'gradient_end': '#764ba2',          # Fin gradiente púrpura
    'gradient_green_start': '#56ccf2',  # Inicio gradiente verde-azul
    'gradient_green_end': '#2f80ed',    # Fin gradiente azul
    
    # Sidebar elite
    'sidebar_bg': '#0d1117',            # Sidebar casi negro
    'sidebar_hover': '#161b22',         # Sidebar hover
    'sidebar_active': '#21262d',        # Sidebar activo
    'sidebar_border': '#30363d',        # Sidebar borde
}

# 🌞 PALETA DE COLORES LIGHT - Modern Light Theme
LIGHT_COLORS = {
    # Colores principales - Light elegant
    'primary_dark': '#1e293b',          # Azul grisáceo oscuro
    'secondary_dark': '#334155',        # Gris azulado
    'accent_purple': '#7c3aed',         # Púrpura vibrante
    'accent_blue': '#0ea5e9',           # Azul sky
    'accent_pink': '#ec4899',           # Rosa brillante
    'accent_green': '#10b981',          # Verde esmeralda
    'accent_orange': '#ff9f43',         # Naranja cálido
    'accent_red': '#ef4444',            # Rojo brillante
    
    # Fondos y superficies
    'bg_primary': '#f8fafc',            # Fondo principal claro
    'bg_secondary': '#f1f5f9',          # Fondo secundario
    'bg_card': '#ffffff',               # Fondo de tarjetas blanco
    'bg_hover': '#e2e8f0',              # Hover state
    'bg_active': '#cbd5e1',             # Active state
    
    # Textos
    'text_primary': '#0f172a',          # Texto principal oscuro
    'text_secondary': '#475569',        # Texto secundario
    'text_muted': '#94a3b8',            # Texto atenuado
    
    # Bordes y divisores
    'border': '#e2e8f0',                # Bordes sutiles
    'border_light': '#cbd5e1',          # Bordes claros
    
    # Estados
    'success': '#22c55e',               # Verde éxito
    'warning': '#eab308',               # Amarillo advertencia
    'danger': '#ef4444',                # Rojo peligro
    'info': '#06b6d4',                  # Cyan información
    
    # Gradientes
    'gradient_start': '#8b5cf6',        # Inicio gradiente púrpura
    'gradient_end': '#6366f1',          # Fin gradiente índigo
    'gradient_green_start': '#3b82f6',  # Inicio gradiente azul
    'gradient_green_end': '#06b6d4',    # Fin gradiente cyan
    
    # Sidebar light
    'sidebar_bg': '#ffffff',            # Sidebar blanco
    'sidebar_hover': '#f1f5f9',         # Sidebar hover
    'sidebar_active': '#e2e8f0',        # Sidebar activo
    'sidebar_border': '#e2e8f0',        # Sidebar borde
}

class AnimationHelper:
    """Helper class para animaciones fluidas"""
    
    @staticmethod
    def fade_in(widget, duration=300, steps=20):
        """Fade in animation"""
        try:
            alpha = 0.0
            increment = 1.0 / steps
            delay = duration // steps
            
            def animate():
                nonlocal alpha
                if alpha < 1.0:
                    alpha += increment
                    # En tkinter no hay opacity directo, simulamos con estados
                    widget.update()
                    widget.after(delay, animate)
            
            animate()
        except:
            pass
    
    @staticmethod
    def fade_out(widget, duration=300, steps=20, callback=None):
        """Fade out animation"""
        try:
            alpha = 1.0
            increment = 1.0 / steps
            delay = duration // steps
            
            def animate():
                nonlocal alpha
                if alpha > 0:
                    alpha -= increment
                    widget.update()
                    widget.after(delay, animate)
                elif callback:
                    callback()
            
            animate()
        except:
            if callback:
                callback()
    
    @staticmethod
    def slide_in(widget, direction='left', duration=400):
        """Slide in animation from direction"""
        try:
            widget.update_idletasks()
            
            if direction == 'left':
                start_x = -widget.winfo_width()
                end_x = 0
            elif direction == 'right':
                start_x = widget.winfo_width()
                end_x = 0
            elif direction == 'top':
                start_y = -widget.winfo_height()
                end_y = 0
            else:  # bottom
                start_y = widget.winfo_height()
                end_y = 0
            
            # Animación simple con place
            steps = 20
            delay = duration // steps
            
            def animate(step=0):
                if step < steps:
                    progress = step / steps
                    if direction in ['left', 'right']:
                        current_x = start_x + (end_x - start_x) * progress
                        widget.place(x=current_x, rely=0)
                    else:
                        current_y = start_y + (end_y - start_y) * progress
                        widget.place(y=current_y, relx=0)
                    widget.after(delay, lambda: animate(step + 1))
            
            animate()
        except:
            pass
    
    @staticmethod
    def smooth_scroll(canvas, target_y, duration=500):
        """Smooth scroll animation"""
        try:
            current_y = canvas.yview()[0]
            diff = target_y - current_y
            steps = 30
            delay = duration // steps
            
            def animate(step=0):
                if step < steps:
                    progress = AnimationHelper._ease_in_out_cubic(step / steps)
                    new_y = current_y + diff * progress
                    canvas.yview_moveto(new_y)
                    canvas.after(delay, lambda: animate(step + 1))
            
            animate()
        except:
            pass
    
    @staticmethod
    def _ease_in_out_cubic(t):
        """Easing function for smooth animations"""
        if t < 0.5:
            return 4 * t * t * t
        else:
            return 1 - pow(-2 * t + 2, 3) / 2
    
    @staticmethod
    def pulse(widget, duration=600, scale=1.05):
        """Pulse animation effect"""
        try:
            original_width = widget.winfo_width()
            original_height = widget.winfo_height()
            
            steps = 20
            delay = duration // steps
            
            def animate(step=0):
                if step < steps:
                    if step < steps // 2:
                        progress = step / (steps // 2)
                        current_scale = 1 + (scale - 1) * progress
                    else:
                        progress = (step - steps // 2) / (steps // 2)
                        current_scale = scale - (scale - 1) * progress
                    
                    widget.update()
                    widget.after(delay, lambda: animate(step + 1))
            
            animate()
        except:
            pass
    
    @staticmethod
    def shake(widget, duration=500):
        """Shake animation for errors"""
        try:
            original_x = widget.winfo_x()
            amplitude = 10
            steps = 20
            delay = duration // steps
            
            def animate(step=0):
                if step < steps:
                    offset = amplitude * (1 - step / steps) * (1 if step % 2 == 0 else -1)
                    widget.place(x=original_x + offset)
                    widget.after(delay, lambda: animate(step + 1))
                else:
                    widget.place(x=original_x)
            
            animate()
        except:
            pass
    
    @staticmethod
    def bounce_in(widget, duration=600):
        """Bounce in animation"""
        try:
            steps = 30
            delay = duration // steps
            
            def animate(step=0):
                if step < steps:
                    t = step / steps
                    if t < 0.5:
                        scale = 4 * t * t * t
                    else:
                        scale = 1 - pow(-2 * t + 2, 3) / 2
                    widget.update()
                    widget.after(delay, lambda: animate(step + 1))
            
            animate()
        except:
            pass


class ToastNotification:
    """Sistema de notificaciones Toast modernas con animaciones"""
    
    _active_toasts = []  # Lista de toasts activos
    _toast_spacing = 10  # Espaciado entre toasts
    
    def __init__(self, parent, message, toast_type='info', duration=3000):
        """
        Crea una notificación toast
        
        Args:
            parent: Widget padre
            message: Mensaje a mostrar
            toast_type: 'success', 'error', 'warning', 'info'
            duration: Duración en milisegundos
        """
        self.parent = parent
        self.message = message
        self.toast_type = toast_type
        self.duration = duration
        self.toast_frame = None
        
        # Colores según tipo (soporta ambos temas)
        colors = {
            'success': {'bg': '#10b981', 'icon': '✓', 'fg': 'white'},
            'error': {'bg': '#ef4444', 'icon': '✕', 'fg': 'white'},
            'warning': {'bg': '#f59e0b', 'icon': '⚠', 'fg': 'white'},
            'info': {'bg': '#3b82f6', 'icon': 'ℹ', 'fg': 'white'}
        }
        
        self.config = colors.get(toast_type, colors['info'])
        
    def show(self):
        """Muestra la notificación toast"""
        try:
            # Crear frame del toast
            self.toast_frame = tk.Frame(self.parent,
                                       bg=self.config['bg'],
                                       highlightthickness=0)
            
            # Contenedor interno con padding
            content = tk.Frame(self.toast_frame,
                              bg=self.config['bg'])
            content.pack(padx=20, pady=12)
            
            # Icono
            icon_label = tk.Label(content,
                                 text=self.config['icon'],
                                 font=('Segoe UI', 16, 'bold'),
                                 bg=self.config['bg'],
                                 fg=self.config['fg'])
            icon_label.pack(side='left', padx=(0, 12))
            
            # Mensaje
            msg_label = tk.Label(content,
                                text=self.message,
                                font=('Segoe UI', 11),
                                bg=self.config['bg'],
                                fg=self.config['fg'],
                                wraplength=350)
            msg_label.pack(side='left')
            
            # Posicionar en la parte superior derecha
            self.toast_frame.update_idletasks()
            width = self.toast_frame.winfo_reqwidth()
            height = self.toast_frame.winfo_reqheight()
            
            # Calcular posición Y considerando otros toasts
            y_offset = 20
            for toast in ToastNotification._active_toasts:
                if toast.toast_frame and toast.toast_frame.winfo_exists():
                    y_offset += toast.toast_frame.winfo_height() + ToastNotification._toast_spacing
            
            x = self.parent.winfo_width() - width - 20
            
            self.toast_frame.place(x=x, y=-height)  # Empieza fuera de pantalla
            
            # Agregar a la lista de activos
            ToastNotification._active_toasts.append(self)
            
            # Animación de entrada (slide down)
            self._slide_in(y_offset)
            
            # Programar cierre automático
            self.parent.after(self.duration, self.hide)
            
        except Exception as e:
            logger.error(f"Error mostrando toast: {e}")
    
    def _slide_in(self, target_y):
        """Animación de entrada deslizante"""
        try:
            steps = 20
            duration = 300
            delay = duration // steps
            
            start_y = -self.toast_frame.winfo_height()
            diff = target_y - start_y
            
            def animate(step=0):
                if step < steps and self.toast_frame.winfo_exists():
                    progress = AnimationHelper._ease_in_out_cubic(step / steps)
                    current_y = start_y + diff * progress
                    self.toast_frame.place(y=current_y)
                    self.parent.after(delay, lambda: animate(step + 1))
            
            animate()
        except:
            pass
    
    def _slide_out(self, callback):
        """Animación de salida deslizante"""
        try:
            steps = 20
            duration = 300
            delay = duration // steps
            
            start_y = self.toast_frame.winfo_y()
            target_y = -self.toast_frame.winfo_height()
            diff = target_y - start_y
            
            def animate(step=0):
                if step < steps and self.toast_frame.winfo_exists():
                    progress = AnimationHelper._ease_in_out_cubic(step / steps)
                    current_y = start_y + diff * progress
                    self.toast_frame.place(y=current_y)
                    self.parent.after(delay, lambda: animate(step + 1))
                elif callback:
                    callback()
            
            animate()
        except:
            if callback:
                callback()
    
    def hide(self):
        """Oculta y destruye la notificación"""
        def destroy():
            try:
                if self.toast_frame and self.toast_frame.winfo_exists():
                    self.toast_frame.destroy()
                if self in ToastNotification._active_toasts:
                    ToastNotification._active_toasts.remove(self)
                # Reposicionar toasts restantes
                self._reposition_toasts()
            except:
                pass
        
        self._slide_out(destroy)
    
    def _reposition_toasts(self):
        """Reposiciona los toasts activos después de eliminar uno"""
        try:
            y_offset = 20
            for toast in ToastNotification._active_toasts:
                if toast.toast_frame and toast.toast_frame.winfo_exists():
                    toast.toast_frame.place(y=y_offset)
                    y_offset += toast.toast_frame.winfo_height() + ToastNotification._toast_spacing
        except:
            pass


class ModernButton(tk.Button):
    """Botón moderno con efectos hover y estilos personalizados"""
    
    def __init__(self, parent, text="", style='primary', command=None, **kwargs):
        # Definir estilos
        styles = {
            'primary': {
                'bg': ELITE_COLORS['accent_purple'],
                'hover': ELITE_COLORS['accent_blue'],
                'active': ELITE_COLORS['gradient_end']
            },
            'success': {
                'bg': ELITE_COLORS['success'],
                'hover': ELITE_COLORS['accent_green'],
                'active': '#00a67c'
            },
            'danger': {
                'bg': ELITE_COLORS['danger'],
                'hover': ELITE_COLORS['accent_red'],
                'active': '#ff4757'
            },
            'warning': {
                'bg': ELITE_COLORS['warning'],
                'hover': ELITE_COLORS['accent_orange'],
                'active': '#ffa502'
            },
            'info': {
                'bg': ELITE_COLORS['info'],
                'hover': ELITE_COLORS['accent_blue'],
                'active': '#00b8d4'
            },
            'secondary': {
                'bg': ELITE_COLORS['bg_secondary'],
                'hover': ELITE_COLORS['bg_hover'],
                'active': ELITE_COLORS['bg_active']
            }
        }
        
        style_config = styles.get(style, styles['primary'])
        
        # Configuración por defecto
        default_config = {
            'font': ('Segoe UI', 11, 'bold'),
            'bg': style_config['bg'],
            'fg': 'white' if style != 'warning' else ELITE_COLORS['primary_dark'],
            'bd': 0,
            'relief': 'flat',
            'cursor': 'hand2',
            'activebackground': style_config['active'],
            'activeforeground': 'white',
            'padx': 20,
            'pady': 10
        }
        
        # Combinar con kwargs personalizados
        default_config.update(kwargs)
        
        super().__init__(parent, text=text, command=command, **default_config)
        
        self.style_config = style_config
        self.default_bg = style_config['bg']
        self.hover_bg = style_config['hover']
        
        # Eventos hover
        self.bind('<Enter>', self._on_enter)
        self.bind('<Leave>', self._on_leave)
    
    def _on_enter(self, event):
        """Efecto hover"""
        self.configure(bg=self.hover_bg)
        AnimationHelper.pulse(self, duration=300, scale=1.02)
    
    def _on_leave(self, event):
        """Restaurar estado normal"""
        self.configure(bg=self.default_bg)


class ModernCard(tk.Frame):
    """Tarjeta moderna con sombra y efectos"""
    
    def __init__(self, parent, title=None, **kwargs):
        # Frame exterior (sombra)
        super().__init__(parent, bg=ELITE_COLORS['bg_primary'])
        
        # Card principal
        self.card = tk.Frame(self,
                            bg=kwargs.get('bg', ELITE_COLORS['bg_card']),
                            highlightbackground=ELITE_COLORS['border'],
                            highlightthickness=1)
        self.card.pack(fill='both', expand=True, padx=2, pady=2)
        
        # Título si existe
        if title:
            title_frame = tk.Frame(self.card, bg=ELITE_COLORS['bg_secondary'])
            title_frame.pack(fill='x')
            
            tk.Label(title_frame,
                    text=title,
                    font=('Segoe UI', 12, 'bold'),
                    bg=ELITE_COLORS['bg_secondary'],
                    fg=ELITE_COLORS['text_primary']).pack(side='left', padx=20, pady=15)
        
        # Contenedor interno
        self.content = tk.Frame(self.card, bg=kwargs.get('bg', ELITE_COLORS['bg_card']))
        self.content.pack(fill='both', expand=True, padx=20, pady=20)
    
    def get_content(self):
        """Retorna el frame de contenido para agregar widgets"""
        return self.content


class ModernSearchBar(tk.Frame):
    """Barra de búsqueda moderna con icono y filtrado en tiempo real"""
    
    def __init__(self, parent, placeholder="🔍 Buscar...", on_search=None, **kwargs):
        """
        Crea una barra de búsqueda moderna
        
        Args:
            parent: Widget padre
            placeholder: Texto de placeholder
            on_search: Callback que se ejecuta cuando cambia el texto (recibe el texto como parámetro)
        """
        super().__init__(parent, bg=kwargs.get('bg', ELITE_COLORS['bg_card']))
        
        self.on_search = on_search
        self.placeholder = placeholder
        self.showing_placeholder = True
        
        # Contenedor con borde redondeado simulado
        search_container = tk.Frame(self,
                                   bg=ELITE_COLORS['bg_secondary'],
                                   highlightbackground=ELITE_COLORS['border'],
                                   highlightthickness=1)
        search_container.pack(fill='x', padx=5, pady=5)
        
        # Icono de búsqueda
        icon_label = tk.Label(search_container,
                             text="🔍",
                             font=('Segoe UI', 12),
                             bg=ELITE_COLORS['bg_secondary'],
                             fg=ELITE_COLORS['text_muted'])
        icon_label.pack(side='left', padx=(12, 8))
        
        # Entry de búsqueda
        self.search_entry = tk.Entry(search_container,
                                     font=('Segoe UI', 11),
                                     bg=ELITE_COLORS['bg_secondary'],
                                     fg=ELITE_COLORS['text_muted'],
                                     bd=0,
                                     relief='flat',
                                     insertbackground=ELITE_COLORS['text_primary'])
        self.search_entry.pack(side='left', fill='x', expand=True, pady=8, padx=(0, 5))
        
        # Mostrar placeholder inicial
        self.search_entry.insert(0, placeholder)
        
        # Botón de limpiar (inicialmente oculto)
        self.clear_btn = tk.Label(search_container,
                                 text="✕",
                                 font=('Segoe UI', 10, 'bold'),
                                 bg=ELITE_COLORS['bg_secondary'],
                                 fg=ELITE_COLORS['text_muted'],
                                 cursor='hand2')
        self.clear_btn.pack(side='right', padx=(0, 12))
        self.clear_btn.pack_forget()  # Ocultar inicialmente
        
        # Eventos
        self.search_entry.bind('<FocusIn>', self._on_focus_in)
        self.search_entry.bind('<FocusOut>', self._on_focus_out)
        self.search_entry.bind('<KeyRelease>', self._on_key_release)
        self.clear_btn.bind('<Button-1>', self._on_clear)
        
    def _on_focus_in(self, event):
        """Elimina el placeholder cuando se enfoca"""
        if self.showing_placeholder:
            self.search_entry.delete(0, tk.END)
            self.search_entry.configure(fg=ELITE_COLORS['text_primary'])
            self.showing_placeholder = False
    
    def _on_focus_out(self, event):
        """Restaura el placeholder si está vacío"""
        if not self.search_entry.get():
            self.search_entry.insert(0, self.placeholder)
            self.search_entry.configure(fg=ELITE_COLORS['text_muted'])
            self.showing_placeholder = True
            self.clear_btn.pack_forget()
    
    def _on_key_release(self, event):
        """Ejecuta la búsqueda al escribir"""
        text = self.search_entry.get()
        
        # Mostrar/ocultar botón de limpiar
        if text and not self.showing_placeholder:
            self.clear_btn.pack(side='right', padx=(0, 12))
        else:
            self.clear_btn.pack_forget()
        
        # Ejecutar callback de búsqueda
        if self.on_search and not self.showing_placeholder:
            self.on_search(text)
    
    def _on_clear(self, event):
        """Limpia el texto de búsqueda"""
        self.search_entry.delete(0, tk.END)
        self.clear_btn.pack_forget()
        if self.on_search:
            self.on_search("")
        self.search_entry.focus()
    
    def get_text(self):
        """Obtiene el texto actual (sin placeholder)"""
        if self.showing_placeholder:
            return ""
        return self.search_entry.get()
    
    def clear(self):
        """Limpia la búsqueda programáticamente"""
        self.search_entry.delete(0, tk.END)
        self.clear_btn.pack_forget()
        if not self.search_entry.focus_get() == self.search_entry:
            self.search_entry.insert(0, self.placeholder)
            self.search_entry.configure(fg=ELITE_COLORS['text_muted'])
            self.showing_placeholder = True


class ModernDialog:
    """Diálogo modal moderno"""
    
    def __init__(self, parent, title="", width=500, height=400):
        self.result = None
        
        # Crear ventana toplevel
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry(f"{width}x{height}")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.resizable(False, False)
        
        # Centrar ventana
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (height // 2)
        self.dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        # Configurar fondo
        self.dialog.configure(bg=ELITE_COLORS['bg_primary'])
        
        # Header
        self.header = tk.Frame(self.dialog,
                              bg=ELITE_COLORS['secondary_dark'],
                              highlightbackground=ELITE_COLORS['border'],
                              highlightthickness=1)
        self.header.pack(fill='x')
        
        # Título del header
        tk.Label(self.header,
                text=title,
                font=('Segoe UI', 14, 'bold'),
                bg=ELITE_COLORS['secondary_dark'],
                fg=ELITE_COLORS['text_primary']).pack(side='left', padx=20, pady=18)
        
        # Botón cerrar
        close_btn = tk.Button(self.header,
                             text="✕",
                             command=self.cancel,
                             font=('Segoe UI', 16),
                             bg=ELITE_COLORS['secondary_dark'],
                             fg=ELITE_COLORS['text_muted'],
                             bd=0,
                             padx=15,
                             pady=5,
                             cursor='hand2',
                             activebackground=ELITE_COLORS['danger'],
                             activeforeground='white')
        close_btn.pack(side='right', padx=10)
        
        # Contenido
        self.content = tk.Frame(self.dialog, bg=ELITE_COLORS['bg_primary'])
        self.content.pack(fill='both', expand=True, padx=0, pady=0)
        
        # Footer para botones
        self.footer = tk.Frame(self.dialog,
                              bg=ELITE_COLORS['bg_secondary'],
                              highlightbackground=ELITE_COLORS['border'],
                              highlightthickness=1)
        self.footer.pack(fill='x', side='bottom')
        
        # Animación de entrada
        AnimationHelper.fade_in(self.dialog, duration=300)
    
    def get_content(self):
        """Retorna el frame de contenido"""
        return self.content
    
    def get_footer(self):
        """Retorna el frame del footer"""
        return self.footer
    
    def cancel(self):
        """Cancela el diálogo"""
        self.result = None
        AnimationHelper.fade_out(self.dialog, duration=200, callback=self.dialog.destroy)
    
    def accept(self, value=True):
        """Acepta el diálogo con un valor"""
        self.result = value
        AnimationHelper.fade_out(self.dialog, duration=200, callback=self.dialog.destroy)
    
    def wait_result(self):
        """Espera y retorna el resultado"""
        self.dialog.wait_window()
        return self.result


class LoadingSpinner:
    """Spinner de carga moderno"""
    
    def __init__(self, parent, text="Cargando..."):
        self.parent = parent
        self.text = text
        self.running = False
        
        # Overlay
        self.overlay = tk.Frame(parent,
                               bg=ELITE_COLORS['bg_primary'])
        
        # Contenedor del spinner
        spinner_container = tk.Frame(self.overlay,
                                    bg=ELITE_COLORS['bg_card'],
                                    highlightbackground=ELITE_COLORS['accent_purple'],
                                    highlightthickness=2)
        spinner_container.place(relx=0.5, rely=0.5, anchor='center')
        
        # Contenido
        content = tk.Frame(spinner_container, bg=ELITE_COLORS['bg_card'])
        content.pack(padx=40, pady=30)
        
        # Spinner (usando caracteres unicode)
        self.spinner_label = tk.Label(content,
                                      text="⟳",
                                      font=('Segoe UI', 32),
                                      bg=ELITE_COLORS['bg_card'],
                                      fg=ELITE_COLORS['accent_purple'])
        self.spinner_label.pack()
        
        # Texto
        tk.Label(content,
                text=text,
                font=('Segoe UI', 11),
                bg=ELITE_COLORS['bg_card'],
                fg=ELITE_COLORS['text_secondary']).pack(pady=(10, 0))
        
        self.angle = 0
    
    def show(self):
        """Muestra el spinner"""
        self.overlay.place(x=0, y=0, relwidth=1, relheight=1)
        self.running = True
        self._animate()
        AnimationHelper.fade_in(self.overlay, duration=200)
    
    def hide(self):
        """Oculta el spinner"""
        self.running = False
        AnimationHelper.fade_out(self.overlay, duration=200,
                                callback=lambda: self.overlay.place_forget())
    
    def _animate(self):
        """Anima el spinner"""
        if self.running:
            # Rotar el icono
            spinners = ['⟳', '⟲', '⟳', '⟲']
            self.angle = (self.angle + 1) % len(spinners)
            self.spinner_label.config(text=spinners[self.angle])
            self.parent.after(150, self._animate)


class ModernTooltip:
    """Tooltip moderno con estilo elite"""
    
    def __init__(self, widget, text, delay=500):
        self.widget = widget
        self.text = text
        self.delay = delay
        self.tooltip_window = None
        self.timer_id = None
        
        # Bind eventos
        widget.bind('<Enter>', self._on_enter)
        widget.bind('<Leave>', self._on_leave)
        widget.bind('<Button>', self._on_leave)
    
    def _on_enter(self, event=None):
        """Cuando el mouse entra"""
        self._cancel_timer()
        self.timer_id = self.widget.after(self.delay, self._show_tooltip)
    
    def _on_leave(self, event=None):
        """Cuando el mouse sale"""
        self._cancel_timer()
        self._hide_tooltip()
    
    def _cancel_timer(self):
        """Cancela el timer"""
        if self.timer_id:
            self.widget.after_cancel(self.timer_id)
            self.timer_id = None
    
    def _show_tooltip(self):
        """Muestra el tooltip"""
        if self.tooltip_window:
            return
        
        # Posición del widget
        x = self.widget.winfo_rootx()
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        
        # Crear ventana del tooltip
        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_attributes('-topmost', True)
        
        # Frame del tooltip
        tooltip_frame = tk.Frame(self.tooltip_window,
                                bg=ELITE_COLORS['bg_secondary'],
                                highlightbackground=ELITE_COLORS['accent_purple'],
                                highlightthickness=1)
        tooltip_frame.pack()
        
        # Texto
        label = tk.Label(tooltip_frame,
                        text=self.text,
                        font=('Segoe UI', 9),
                        bg=ELITE_COLORS['bg_secondary'],
                        fg=ELITE_COLORS['text_primary'],
                        padx=12,
                        pady=8,
                        justify='left')
        label.pack()
        
        # Posicionar
        self.tooltip_window.wm_geometry(f"+{x}+{y}")
        
        # Animación de entrada
        AnimationHelper.fade_in(tooltip_frame, duration=200)
    
    def _hide_tooltip(self):
        """Oculta el tooltip"""
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None


class ProgressBar:
    """Barra de progreso moderna"""
    
    def __init__(self, parent, width=300, height=6):
        self.parent = parent
        self.width = width
        self.height = height
        self.progress = 0
        
        # Container
        self.container = tk.Frame(parent,
                                 bg=ELITE_COLORS['bg_secondary'],
                                 highlightbackground=ELITE_COLORS['border'],
                                 highlightthickness=1)
        
        # Canvas para la barra
        self.canvas = tk.Canvas(self.container,
                               width=width,
                               height=height,
                               bg=ELITE_COLORS['bg_secondary'],
                               highlightthickness=0)
        self.canvas.pack(padx=2, pady=2)
        
        # Barra de fondo
        self.bg_bar = self.canvas.create_rectangle(
            0, 0, width, height,
            fill=ELITE_COLORS['bg_secondary'],
            outline=''
        )
        
        # Barra de progreso
        self.progress_bar = self.canvas.create_rectangle(
            0, 0, 0, height,
            fill=ELITE_COLORS['accent_purple'],
            outline=''
        )
    
    def set_progress(self, percent):
        """Establece el progreso (0-100)"""
        self.progress = max(0, min(100, percent))
        new_width = (self.width * self.progress) / 100
        
        # Animación suave
        self._animate_to(new_width)
        
        # Cambiar color según progreso
        if self.progress < 30:
            color = ELITE_COLORS['danger']
        elif self.progress < 70:
            color = ELITE_COLORS['warning']
        else:
            color = ELITE_COLORS['success']
        
        self.canvas.itemconfig(self.progress_bar, fill=color)
    
    def _animate_to(self, target_width):
        """Anima el cambio de tamaño"""
        current_coords = self.canvas.coords(self.progress_bar)
        current_width = current_coords[2] if len(current_coords) >= 3 else 0
        
        steps = 15
        diff = target_width - current_width
        increment = diff / steps
        
        def animate(step=0):
            if step < steps:
                new_width = current_width + (increment * step)
                self.canvas.coords(self.progress_bar, 0, 0, new_width, self.height)
                self.parent.after(20, lambda: animate(step + 1))
            else:
                self.canvas.coords(self.progress_bar, 0, 0, target_width, self.height)
        
        animate()
    
    def pack(self, **kwargs):
        """Pack el container"""
        self.container.pack(**kwargs)
    
    def grid(self, **kwargs):
        """Grid el container"""
        self.container.grid(**kwargs)
    
    @staticmethod
    def rotate_icon(label, duration=1000):
        """Rotate an emoji icon (loading effect)"""
        icons = ['◐', '◓', '◑', '◒']
        current = [0]
        
        def animate():
            if hasattr(label, 'winfo_exists') and label.winfo_exists():
                label.configure(text=icons[current[0] % len(icons)])
                current[0] += 1
                label.after(duration // len(icons), animate)
        
        animate()


class LoadingSpinner:
    """Loading spinner moderno"""
    
    def __init__(self, parent, text="Cargando..."):
        self.overlay = tk.Toplevel(parent)
        self.overlay.withdraw()
        self.overlay.overrideredirect(True)
        self.overlay.attributes('-topmost', True)
        self.overlay.configure(bg=ELITE_COLORS['bg_primary'])
        
        # Frame semi-transparente simulado
        frame = tk.Frame(self.overlay,
                        bg=ELITE_COLORS['bg_secondary'],
                        highlightbackground=ELITE_COLORS['border'],
                        highlightthickness=2)
        frame.pack(padx=3, pady=3)
        
        content = tk.Frame(frame, bg=ELITE_COLORS['bg_secondary'])
        content.pack(padx=50, pady=40)
        
        # Spinner icon
        self.spinner_label = tk.Label(content,
                                      text='◐',
                                      font=('Segoe UI', 32),
                                      bg=ELITE_COLORS['bg_secondary'],
                                      fg=ELITE_COLORS['accent_blue'])
        self.spinner_label.pack(pady=(0, 15))
        
        # Texto
        tk.Label(content,
                text=text,
                font=('Segoe UI', 12),
                bg=ELITE_COLORS['bg_secondary'],
                fg=ELITE_COLORS['text_primary']).pack()
        
        # Centrar
        self.overlay.update_idletasks()
        x = (parent.winfo_screenwidth() - self.overlay.winfo_width()) // 2
        y = (parent.winfo_screenheight() - self.overlay.winfo_height()) // 2
        self.overlay.geometry(f'+{x}+{y}')
        
        # Iniciar animación
        AnimationHelper.rotate_icon(self.spinner_label, duration=800)
    
    def show(self):
        """Muestra el spinner"""
        self.overlay.deiconify()
        self.overlay.grab_set()
    
    def hide(self):
        """Oculta el spinner"""
        try:
            self.overlay.grab_release()
            self.overlay.destroy()
        except:
            pass


class ModernButton(tk.Button):
    """Botón moderno con efectos avanzados"""
    
    def __init__(self, parent, text, command=None, style='primary', **kwargs):
        # Estilos predefinidos
        styles = {
            'primary': {
                'bg': ELITE_COLORS['accent_purple'],
                'hover': ELITE_COLORS['accent_blue'],
                'active': ELITE_COLORS['gradient_end']
            },
            'success': {
                'bg': ELITE_COLORS['success'],
                'hover': ELITE_COLORS['accent_green'],
                'active': '#00a67c'
            },
            'danger': {
                'bg': ELITE_COLORS['danger'],
                'hover': ELITE_COLORS['accent_red'],
                'active': '#ff4757'
            },
            'warning': {
                'bg': ELITE_COLORS['warning'],
                'hover': ELITE_COLORS['accent_orange'],
                'active': '#ffa502'
            },
            'info': {
                'bg': ELITE_COLORS['info'],
                'hover': ELITE_COLORS['accent_blue'],
                'active': '#00b8d4'
            }
        }
        
        style_config = styles.get(style, styles['primary'])
        
        super().__init__(
            parent,
            text=text,
            command=command,
            font=('Segoe UI', 11, 'bold'),
            bg=style_config['bg'],
            fg='white',
            bd=0,
            padx=kwargs.get('padx', 25),
            pady=kwargs.get('pady', 12),
            cursor='hand2',
            relief='flat',
            activebackground=style_config['active'],
            activeforeground='white',
            **{k: v for k, v in kwargs.items() if k not in ['padx', 'pady']}
        )
        
        self.default_bg = style_config['bg']
        self.hover_bg = style_config['hover']
        
        # Bind hover effects
        self.bind('<Enter>', self._on_enter)
        self.bind('<Leave>', self._on_leave)
    
    def _on_enter(self, e):
        self.configure(bg=self.hover_bg)
        AnimationHelper.pulse(self, duration=300, scale=1.03)
    
    def _on_leave(self, e):
        self.configure(bg=self.default_bg)


class ModernTPV:
    """Aplicación TPV Elite - Diseño premium con animaciones"""
    
    def __init__(self, root: tk.Tk):
        self.root = root

        # Cargar config primero — todo lo demás depende de ella
        self.config = load_config('config.json')
        _b = self.config.get('business', {})
        self.root.title(f"{_b.get('tipo_emoji','🏪')} {_b.get('nombre','Sistema TPV Elite')} — {_b.get('tipo_label','')}")
        self.current_theme = self.config.get('theme', 'dark')  # 'dark' o 'light'
        
        # Usar colores según tema
        self.colors = ELITE_COLORS if self.current_theme == 'dark' else LIGHT_COLORS
        
        # Configurar ventana principal
        self._setup_window()
        
        # Inicializar servicios
        self._init_services()
        
        # Variables de la aplicación
        self.current_cart = []
        self.current_screen = None
        self.animation_helper = AnimationHelper()
        
        # Configurar estilos elite
        self._configure_elite_styles()
        
        # Crear interfaz principal
        self._create_main_layout()
        
        # Configurar atajos de teclado
        self._setup_keyboard_shortcuts()
        
        # Mostrar pantalla de login con animación o wizard si es primera vez
        if not self.config.get('business', {}).get('configured', False):
            self.show_setup_wizard()
        else:
            self._apply_business_branding()
            self.show_login_screen()

        # Verificar actualizaciones en segundo plano (3 seg después del inicio)
        self.root.after(3000, lambda: check_for_updates(self.root))

        logger.info("✨ Aplicación ELITE iniciada correctamente")
    
    def _setup_keyboard_shortcuts(self):
        """Configura los atajos de teclado de la aplicación"""
        # F1 - Ayuda
        self.root.bind('<F1>', lambda e: self._safe_navigate(self.show_help))
        
        # F5 - Actualizar pantalla actual
        self.root.bind('<F5>', lambda e: self._refresh_current_screen())
        
        # Escape - Cerrar diálogo actual (si existe)
        # Este se maneja en cada diálogo individual
        
        # Navegación con Ctrl+Número
        self.root.bind('<Control-Key-1>', lambda e: self._safe_navigate(self.show_dashboard))
        self.root.bind('<Control-Key-2>', lambda e: self._safe_navigate(self.show_pos_screen))
        self.root.bind('<Control-Key-3>', lambda e: self._safe_navigate(self.show_sales_history))
        self.root.bind('<Control-Key-4>', lambda e: self._safe_navigate(self.show_products_manager))
        self.root.bind('<Control-Key-5>', lambda e: self._safe_navigate(self.show_reports))
        self.root.bind('<Control-Key-6>', lambda e: self._safe_navigate(self.show_settings))
        self.root.bind('<Control-Key-7>', lambda e: self._safe_navigate(self.show_users_manager))
        
        logger.info("Atajos de teclado configurados")
    
    def _safe_navigate(self, navigation_func):
        """Navega de forma segura solo si el usuario está autenticado"""
        if self.auth_service.current_user:
            try:
                navigation_func()
            except Exception as e:
                logger.error(f"Error en navegación: {e}")
                messagebox.showerror("Error", f"Error al navegar: {str(e)}")
        else:
            logger.warning("Intento de navegación sin autenticación")
    
    def _refresh_current_screen(self):
        """Actualiza la pantalla actual"""
        if not self.auth_service.current_user:
            return
        
        try:
            if self.current_screen == "dashboard":
                self.show_dashboard()
            elif self.current_screen == "pos":
                self.show_pos_screen()
            elif self.current_screen == "sales":
                self.show_sales_history()
            elif self.current_screen == "products":
                self.show_products_manager()
            elif self.current_screen == "orders":
                self.show_orders_manager()
            elif self.current_screen == "reports":
                self.show_reports()
            elif self.current_screen == "settings":
                self.show_settings()
            elif self.current_screen == "help":
                self.show_help()
            else:
                self.show_dashboard()
                
            logger.info(f"Pantalla '{self.current_screen}' actualizada")
        except Exception as e:
            logger.error(f"Error al actualizar pantalla: {e}")
            messagebox.showerror("Error", f"Error al actualizar: {str(e)}")
    
    def _setup_window(self):
        """Configura la ventana principal"""
        # Obtener dimensiones de pantalla
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Configurar tamaño mínimo
        min_width = self.config.get('app', {}).get('min_width', 800)
        min_height = self.config.get('app', {}).get('min_height', 600)
        self.root.minsize(min_width, min_height)
        
        # Maximizar ventana
        self.root.state('zoomed')
        
        # Configurar grid principal
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        # Aplicar tema oscuro elite
        self.root.configure(bg=self.colors['bg_primary'])
        
        # Intentar cargar ícono
        try:
            icon_path = resource_path('icon.ico')
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except:
            pass
    
    def _init_services(self):
        """Inicializa los servicios de la aplicación"""
        # Usar la base de datos local en el directorio de la aplicación
        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'heladeria.db')

        # Asegurar directorios para backups y reportes
        self.app_data_dir = get_app_data_dir()
        ensure_directory(self.app_data_dir)
        ensure_directory(os.path.join(self.app_data_dir, 'backups'))
        ensure_directory(os.path.join(self.app_data_dir, 'reports'))
        ensure_directory(os.path.join(self.app_data_dir, 'invoices'))

        # Inicializar base de datos
        self.db = DatabaseManager(db_path)

        # Inicializar cloud sync (opcional — si no hay credenciales, modo offline)
        self._init_cloud_sync()

        # Inicializar servicios
        self.auth_service = AuthService(self.db)
        self.producto_service = ProductoService(self.db)
        self.venta_service = VentaService(self.db, self.producto_service)

        logger.info(f"Servicios inicializados correctamente - DB: {db_path}")

    def _init_cloud_sync(self):
        """Inicializa el servicio de sincronización con Supabase."""
        try:
            import cloud_sync

            config_cloud = self.config.get('cloud', {})
            config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')

            # Generar tenant_id si aún no existe
            if not config_cloud.get('tenant_id'):
                nuevo_id = cloud_sync.generar_tenant_id()
                cloud_sync.guardar_tenant_id(config_path, nuevo_id)
                config_cloud['tenant_id'] = nuevo_id
                # Recargar config en memoria
                self.config = load_config('config.json')
                logger.info(f"Tenant ID generado: {nuevo_id}")

            sync = cloud_sync.inicializar(config_cloud)

            if sync.activo:
                # Pull inicial en background para sincronizar con otros dispositivos
                threading.Thread(
                    target=sync.pull_all,
                    args=(self.db,),
                    daemon=True,
                    name='cloud-pull-init'
                ).start()
                logger.info("Cloud sync activo — pull inicial iniciado en background.")
            else:
                logger.info("Cloud sync en modo offline (sin credenciales o deshabilitado).")
        except Exception as exc:
            logger.warning(f"Cloud sync no pudo inicializarse: {exc}")
    
    def _configure_elite_styles(self):
        """Configura estilos ELITE con diseño premium"""
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Obtener configuración de fuentes
        fonts = self.config.get('fonts', {})
        font_family = 'Segoe UI'
        
        # ============= FRAMES ELITE =============
        self.style.configure('Main.TFrame',
                           background=self.colors['bg_primary'])
        
        self.style.configure('Sidebar.TFrame',
                           background=self.colors['sidebar_bg'],
                           relief='flat')
        
        self.style.configure('Header.TFrame',
                           background=self.colors['secondary_dark'],
                           relief='flat')
        
        self.style.configure('Card.TFrame',
                           background=self.colors['bg_card'],
                           relief='flat',
                           borderwidth=0)
        
        # ============= LABELS ELITE =============
        self.style.configure('Title.TLabel',
                           font=(font_family, 22, 'bold'),
                           background=self.colors['secondary_dark'],
                           foreground=self.colors['text_primary'])
        
        self.style.configure('Subtitle.TLabel',
                           font=(font_family, 16, 'bold'),
                           background=self.colors['bg_card'],
                           foreground=self.colors['text_primary'])
        
        self.style.configure('Info.TLabel',
                           font=(font_family, 11),
                           background=self.colors['bg_card'],
                           foreground=self.colors['text_secondary'])
        
        self.style.configure('Sidebar.TLabel',
                           font=(font_family, 11),
                           background=self.colors['sidebar_bg'],
                           foreground=self.colors['text_secondary'])
        
        self.style.configure('StatValue.TLabel',
                           font=(font_family, 28, 'bold'),
                           background=self.colors['bg_card'],
                           foreground=self.colors['accent_blue'])
        
        self.style.configure('StatLabel.TLabel',
                           font=(font_family, 10),
                           background=self.colors['bg_card'],
                           foreground=self.colors['text_muted'])
        
        # ============= BUTTONS ELITE con efectos =============
        self.style.configure('Primary.TButton',
                           font=(font_family, 11, 'bold'),
                           background=self.colors['accent_purple'],
                           foreground='white',
                           borderwidth=0,
                           focuscolor='none',
                           relief='flat',
                           padding=(25, 12))
        
        self.style.map('Primary.TButton',
                      background=[('active', self.colors['accent_blue']),
                                ('pressed', self.colors['gradient_end'])])
        
        self.style.configure('Success.TButton',
                           font=(font_family, 11, 'bold'),
                           background=self.colors['success'],
                           foreground=self.colors['primary_dark'],
                           borderwidth=0,
                           focuscolor='none',
                           relief='flat',
                           padding=(20, 10))
        
        self.style.map('Success.TButton',
                      background=[('active', self.colors['accent_green']),
                                ('pressed', '#00a67c')])
        
        self.style.configure('Danger.TButton',
                           font=(font_family, 11, 'bold'),
                           background=self.colors['danger'],
                           foreground='white',
                           borderwidth=0,
                           focuscolor='none',
                           relief='flat',
                           padding=(20, 10))
        
        self.style.map('Danger.TButton',
                      background=[('active', self.colors['accent_red']),
                                ('pressed', '#ff4757')])
        
        self.style.configure('Warning.TButton',
                           font=(font_family, 11, 'bold'),
                           background=self.colors['warning'],
                           foreground=self.colors['primary_dark'],
                           borderwidth=0,
                           focuscolor='none',
                           relief='flat',
                           padding=(20, 10))
        
        self.style.map('Warning.TButton',
                      background=[('active', self.colors['accent_orange']),
                                ('pressed', '#ffa502')])
        
        self.style.configure('Info.TButton',
                           font=(font_family, 11, 'bold'),
                           background=self.colors['info'],
                           foreground=self.colors['primary_dark'],
                           borderwidth=0,
                           focuscolor='none',
                           relief='flat',
                           padding=(20, 10))
        
        self.style.map('Info.TButton',
                      background=[('active', self.colors['accent_blue']),
                                ('pressed', '#00b8d4')])
        
        # Botones sidebar con efecto glassmorphism
        self.style.configure('Sidebar.TButton',
                           font=(font_family, 11),
                           background=self.colors['sidebar_bg'],
                           foreground=self.colors['text_secondary'],
                           borderwidth=0,
                           focuscolor='none',
                           anchor='w',
                           relief='flat',
                           padding=(20, 16))
        
        self.style.map('Sidebar.TButton',
                      background=[('active', self.colors['sidebar_hover']),
                                ('pressed', self.colors['sidebar_active'])],
                      foreground=[('active', self.colors['text_primary']),
                                ('pressed', self.colors['accent_blue'])])
        
        self.style.configure('SidebarActive.TButton',
                           font=(font_family, 11, 'bold'),
                           background=self.colors['sidebar_active'],
                           foreground=self.colors['accent_blue'],
                           borderwidth=0,
                           focuscolor='none',
                           anchor='w',
                           relief='flat',
                           padding=(20, 16))
        
        # ============= TREEVIEW ELITE =============
        self.style.configure('Modern.Treeview',
                           font=(font_family, 10),
                           background=self.colors['bg_card'],
                           foreground=self.colors['text_primary'],
                           fieldbackground=self.colors['bg_card'],
                           borderwidth=0,
                           relief='flat',
                           rowheight=38)

        self.style.configure('Modern.Treeview.Heading',
                           font=(font_family, 10, 'bold'),
                           background=self.colors['secondary_dark'],
                           foreground=self.colors['accent_blue'],
                           borderwidth=0,
                           relief='flat',
                           padding=(8, 10))

        self.style.map('Modern.Treeview',
                      background=[('selected', self.colors['accent_purple'])],
                      foreground=[('selected', 'white')])

        self.style.map('Modern.Treeview.Heading',
                      background=[('active', self.colors['bg_hover'])])
        
        # ============= LABELFRAME ELITE =============
        self.style.configure('Card.TLabelframe',
                           font=(font_family, 12, 'bold'),
                           background=self.colors['bg_card'],
                           foreground=self.colors['accent_blue'],
                           borderwidth=1,
                           bordercolor=self.colors['border'],
                           relief='flat')
        
        self.style.configure('Card.TLabelframe.Label',
                           font=(font_family, 12, 'bold'),
                           background=self.colors['bg_card'],
                           foreground=self.colors['accent_blue'])
        
        # ============= ENTRY ELITE =============
        self.style.configure('Elite.TEntry',
                           font=(font_family, 11),
                           fieldbackground=self.colors['bg_secondary'],
                           foreground=self.colors['text_primary'],
                           borderwidth=1,
                           relief='flat',
                           insertcolor=self.colors['accent_blue'])
        
        # ============= COMBOBOX ELITE =============
        self.style.configure('Elite.TCombobox',
                           font=(font_family, 11),
                           fieldbackground=self.colors['bg_secondary'],
                           foreground=self.colors['text_primary'],
                           background=self.colors['bg_secondary'],
                           borderwidth=1,
                           relief='flat',
                           arrowcolor=self.colors['accent_blue'])
        
        # ============= SCROLLBAR ELITE =============
        self.style.configure('Elite.Vertical.TScrollbar',
                           background=self.colors['bg_secondary'],
                           troughcolor=self.colors['bg_primary'],
                           borderwidth=0,
                           relief='flat',
                           arrowsize=0)
        
        self.style.map('Elite.Vertical.TScrollbar',
                      background=[('active', self.colors['accent_purple'])])
        
        # ============= SEPARATOR =============
        self.style.configure('Elite.TSeparator',
                           background=self.colors['border'])
        
        logger.info("✨ Estilos ELITE configurados correctamente")
    
    def toggle_theme(self):
        """Cambia entre tema claro y oscuro con animación"""
        try:
            # Cambiar tema
            self.current_theme = 'light' if self.current_theme == 'dark' else 'dark'
            self.colors = LIGHT_COLORS if self.current_theme == 'light' else ELITE_COLORS
            
            # Guardar en configuración
            self.config['theme'] = self.current_theme
            import json
            config_path = os.path.join(os.path.dirname(__file__), 'config.json')
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4)
            
            # Mostrar notificación
            theme_name = "Claro ☀️" if self.current_theme == 'light' else "Oscuro 🌙"
            toast = ToastNotification(self.root, f"Tema cambiado a {theme_name}", 'info', 2000)
            toast.show()
            
            # Reconfigurar estilos
            self._configure_elite_styles()
            
            # Refrescar pantalla actual
            if self.auth_service.current_user and self.current_screen:
                self._refresh_current_screen()
            
            logger.info(f"Tema cambiado a: {self.current_theme}")
            
        except Exception as e:
            logger.error(f"Error cambiando tema: {e}")
    
    def _create_theme_toggle(self, parent):
        """Crea el botón toggle para cambiar de tema"""
        toggle_frame = tk.Frame(parent, bg=self.colors['sidebar_bg'])
        toggle_frame.pack(side='bottom', fill='x', padx=15, pady=15)
        
        # Icono según tema actual
        icon = "🌙" if self.current_theme == 'dark' else "☀️"
        text = "Modo Claro" if self.current_theme == 'dark' else "Modo Oscuro"
        
        toggle_btn = ModernButton(toggle_frame,
                                  text=f"{icon}  {text}",
                                  style='secondary',
                                  command=self.toggle_theme)
        toggle_btn.pack(fill='x')
        
        return toggle_frame
    
    def _create_main_layout(self):
        """Crea el layout principal de la aplicación"""
        # Container principal
        self.main_container = ttk.Frame(self.root, style='Main.TFrame')
        self.main_container.grid(row=0, column=0, sticky='nsew')
        # Columna 0 tiene weight para login/wizard; al entrar a la app se reconfigura
        self.main_container.grid_columnconfigure(0, weight=1)
        self.main_container.grid_columnconfigure(1, weight=0)
        self.main_container.grid_rowconfigure(0, weight=1)
        self.main_container.grid_rowconfigure(1, weight=0)
        
        # Header (será visible después del login)
        self.header_frame = ttk.Frame(self.main_container, style='Header.TFrame', height=60)
        # No lo mostramos aún
        
        # Sidebar (será visible después del login)
        self.sidebar_frame = ttk.Frame(self.main_container, style='Sidebar.TFrame',
                                      width=self.config.get('ui', {}).get('sidebar_width', 220))
        # No lo mostramos aún
        
        # Content area
        self.content_frame = ttk.Frame(self.main_container, style='Main.TFrame')
        self.content_frame.grid(row=0, column=0, rowspan=2, sticky='nsew', padx=0, pady=0)
        self.content_frame.grid_columnconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(0, weight=1)
    
    def _create_header(self):
        """Crea el header ELITE de la aplicación"""
        self.header_frame.grid(row=0, column=0, columnspan=2, sticky='ew')
        self.header_frame.grid_columnconfigure(1, weight=1)
        
        # Fondo con efecto
        self.header_frame.configure(style='Header.TFrame')
        
        # Logo/Título con efecto glow
        title_container = tk.Frame(self.header_frame, bg=self.colors['secondary_dark'])
        title_container.grid(row=0, column=0, padx=25, pady=18)
        
        _biz = self.config.get('business', {})
        _biz_emoji = _biz.get('tipo_emoji', '🏪')
        _biz_nombre = _biz.get('nombre', 'Sistema TPV Elite')
        _biz_tipo_label = _biz.get('tipo_label', '')

        icon_label = tk.Label(title_container, text=_biz_emoji,
                             font=('Segoe UI', 24),
                             bg=self.colors['secondary_dark'],
                             fg=self.colors['accent_blue'])
        icon_label.pack(side='left', padx=(0, 10))
        
        title_label = tk.Label(title_container,
                              text=_biz_nombre,
                              font=('Segoe UI', 18, 'bold'),
                              bg=self.colors['secondary_dark'],
                              fg=self.colors['text_primary'])
        title_label.pack(side='left')
        
        subtitle_label = tk.Label(title_container,
                                 text=_biz_tipo_label,
                                 font=('Segoe UI', 9),
                                 bg=self.colors['secondary_dark'],
                                 fg=self.colors['text_muted'])
        subtitle_label.pack(side='left', padx=(10, 0))
        
        # Info usuario con diseño moderno
        if self.auth_service.current_user:
            user_container = tk.Frame(self.header_frame, bg=self.colors['bg_hover'],
                                     highlightbackground=self.colors['border'],
                                     highlightthickness=1)
            user_container.grid(row=0, column=1, padx=20, pady=12, sticky='e')
            
            user_icon = tk.Label(user_container, text="👤",
                               font=('Segoe UI', 14),
                               bg=self.colors['bg_hover'],
                               fg=self.colors['accent_blue'])
            user_icon.pack(side='left', padx=(10, 5))
            
            user_name = tk.Label(user_container,
                                text=self.auth_service.current_user.nombre,
                                font=('Segoe UI', 11, 'bold'),
                                bg=self.colors['bg_hover'],
                                fg=self.colors['text_primary'])
            user_name.pack(side='left', padx=(0, 5))
            
            user_role = tk.Label(user_container,
                               text=f"• {self.auth_service.current_user.rol}",
                               font=('Segoe UI', 9),
                               bg=self.colors['bg_hover'],
                               fg=self.colors['text_muted'])
            user_role.pack(side='left', padx=(0, 10))
            
            # Botón cerrar sesión moderno
            logout_btn = tk.Button(self.header_frame,
                                  text="⏻  Cerrar Sesión",
                                  command=self.logout,
                                  font=('Segoe UI', 10, 'bold'),
                                  bg=self.colors['danger'],
                                  fg='white',
                                  bd=0,
                                  padx=20,
                                  pady=10,
                                  cursor='hand2',
                                  relief='flat',
                                  activebackground=self.colors['accent_red'],
                                  activeforeground='white')
            logout_btn.grid(row=0, column=2, padx=25, pady=12)
            
            # Efecto hover
            logout_btn.bind('<Enter>', lambda e: logout_btn.configure(bg=self.colors['accent_red']))
            logout_btn.bind('<Leave>', lambda e: logout_btn.configure(bg=self.colors['danger']))
    
    def _create_sidebar(self):
        """Crea el sidebar ELITE con navegación moderna"""
        self.sidebar_frame.grid(row=1, column=0, sticky='ns')
        self.sidebar_frame.grid_propagate(False)
        
        # Frame superior con logo - SÚPER COMPACTO
        top_frame = tk.Frame(self.sidebar_frame, bg=self.colors['sidebar_bg'])
        top_frame.pack(fill='x', pady=10, padx=12)  # Reducido de 15 a 10
        
        # Logo con efecto - MÁS PEQUEÑO
        logo_container = tk.Frame(top_frame, bg=self.colors['sidebar_active'],
                                 highlightbackground=self.colors['accent_purple'],
                                 highlightthickness=2)
        logo_container.pack(pady=5)  # Reducido de 8 a 5
        
        logo_label = tk.Label(logo_container,
                             text=self.config.get('business', {}).get('tipo_emoji', '🏪'),
                             font=('Segoe UI', 22),  # Reducido de 28 a 22
                             bg=self.colors['sidebar_active'],
                             fg=self.colors['accent_blue'])
        logo_label.pack(padx=10, pady=10)  # Reducido de 12 a 10
        
        # Título del sidebar con estilo - SÚPER COMPACTO
        sidebar_title = tk.Label(self.sidebar_frame,
                                text="NAVEGACIÓN",
                                font=('Segoe UI', 8, 'bold'),  # Reducido de 9 a 8
                                bg=self.colors['sidebar_bg'],
                                fg=self.colors['text_muted'])
        sidebar_title.pack(pady=(8, 5), padx=15, anchor='w')  # Reducido pady
        
        # Contenedor de botones - SIN EXPAND para ver todos
        buttons_frame = tk.Frame(self.sidebar_frame, bg=self.colors['sidebar_bg'])
        buttons_frame.pack(fill='x', padx=8)
        
        # Botones de navegación con iconos mejorados
        menu_items = [
            ("🏠", "Inicio", self.show_dashboard, self.colors['accent_blue']),
            ("🛒", "Nueva Venta", self.show_pos_screen, self.colors['success']),
            ("📊", "Ventas", self.show_sales_history, self.colors['accent_purple']),
            ("📦", "Productos", self.show_products_manager, self.colors['accent_orange']),
            ("📋", "Pedidos", self.show_orders_manager, self.colors['info']),
            ("📈", "Reportes", self.show_reports, self.colors['accent_pink']),
            ("👥", "Usuarios", self.show_users_manager, self.colors['warning']),
            ("⚙️", "Configuración", self.show_settings, self.colors['text_muted']),
            ("❓", "Ayuda", self.show_help, self.colors['accent_blue']),
        ]
        
        self.sidebar_buttons = {}
        
        for icon, text, command, color in menu_items:
            # Frame para cada botón
            btn_frame = tk.Frame(buttons_frame, bg=self.colors['sidebar_bg'])
            btn_frame.pack(fill='x', pady=1)  # Reducido de 2 a 1
            
            # Indicador lateral (inicialmente oculto)
            indicator = tk.Frame(btn_frame, bg=color, width=4)
            indicator.pack(side='left', fill='y')
            indicator.pack_forget()  # Ocultar inicialmente
            
            # Botón principal - SÚPER COMPACTO
            btn = tk.Button(btn_frame,
                           text=f"{icon}  {text}",
                           command=lambda cmd=command, b=btn_frame, ind=indicator: self._navigate_with_animation(cmd, b, ind),
                           font=('Segoe UI', 9),  # Reducido de 10 a 9
                           bg=self.colors['sidebar_bg'],
                           fg=self.colors['text_secondary'],
                           bd=0,
                           padx=15,  # Reducido de 18 a 15
                           pady=8,  # Reducido de 11 a 8
                           cursor='hand2',
                           anchor='w',
                           relief='flat',
                           activebackground=self.colors['sidebar_hover'],
                           activeforeground=self.colors['text_primary'])
            btn.pack(side='left', fill='both', expand=True)
            
            # Efectos hover
            def on_enter(e, button=btn, frame=btn_frame):
                button.configure(bg=self.colors['sidebar_hover'],
                               fg=self.colors['text_primary'])
            
            def on_leave(e, button=btn, frame=btn_frame, txt=text):
                if self.current_screen != txt.lower().replace(' ', '_'):
                    button.configure(bg=self.colors['sidebar_bg'],
                                   fg=self.colors['text_secondary'])
            
            btn.bind('<Enter>', on_enter)
            btn.bind('<Leave>', on_leave)
            
            self.sidebar_buttons[text] = (btn, indicator, btn_frame)
        
        # Separador elegante - SÚPER COMPACTO
        separator_frame = tk.Frame(self.sidebar_frame, bg=self.colors['sidebar_bg'], height=1)
        separator_frame.pack(fill='x', padx=15, pady=8)  # Reducido de 15 a 8
        
        sep_line = tk.Frame(separator_frame, bg=self.colors['border'], height=1)
        sep_line.pack(fill='x')
        
        # ✨ TOGGLE DE TEMA
        self._create_theme_toggle(self.sidebar_frame)
    
    def _navigate_with_animation(self, command, btn_frame, indicator):
        """Navega con animación suave y actualiza indicadores"""
        # Ocultar todos los indicadores
        for btn, ind, frame in self.sidebar_buttons.values():
            ind.pack_forget()
            btn.configure(bg=self.colors['sidebar_bg'],
                         fg=self.colors['text_secondary'])
        
        # Mostrar indicador del botón actual
        indicator.pack(side='left', fill='y')
        
        # Cambiar color del botón activo
        for widget in btn_frame.winfo_children():
            if isinstance(widget, tk.Button):
                widget.configure(bg=self.colors['sidebar_active'],
                               fg=self.colors['accent_blue'])
        
        # Transición suave: fade out → ejecutar comando → fade in
        self._smooth_transition(command)
    
    def _smooth_transition(self, command):
        """Ejecuta una transición suave entre pantallas"""
        # Si hay contenido, hacer fade out
        children = self.content_frame.winfo_children()
        if children:
            # Fade out del contenido actual
            for widget in children:
                AnimationHelper.fade_out(widget, duration=200)
            
            # Esperar a que termine el fade out antes de limpiar
            self.root.after(250, lambda: self._execute_transition(command))
        else:
            # Si no hay contenido, ejecutar directamente
            command()
    
    def _execute_transition(self, command):
        """Ejecuta el comando después del fade out"""
        # Limpiar contenido
        self.clear_content()
        
        # Ejecutar el comando (carga la nueva pantalla)
        command()
        
        # El fade in se maneja automáticamente en cada pantalla
        # (ya tienen AnimationHelper.fade_in en su código)
    
    def clear_content(self):
        """Limpia el área de contenido"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()
    
    def show_modern_confirm(self, title, message, type='info'):
        """Muestra un diálogo de confirmación moderno"""
        dialog = ModernDialog(self.root, title=title, width=500, height=300)
        content = dialog.get_content()
        
        # Icono y mensaje
        message_frame = tk.Frame(content, bg=ELITE_COLORS['bg_primary'])
        message_frame.pack(expand=True, fill='both', padx=30, pady=30)
        
        # Icono grande
        icons = {
            'info': ('ℹ️', ELITE_COLORS['info']),
            'warning': ('⚠️', ELITE_COLORS['warning']),
            'error': ('✕', ELITE_COLORS['danger']),
            'success': ('✓', ELITE_COLORS['success']),
            'question': ('❓', ELITE_COLORS['accent_purple'])
        }
        
        icon, color = icons.get(type, icons['info'])
        
        icon_container = tk.Frame(message_frame,
                                 bg=ELITE_COLORS['bg_secondary'],
                                 highlightbackground=color,
                                 highlightthickness=2)
        icon_container.pack(pady=(0, 20))
        
        tk.Label(icon_container,
                text=icon,
                font=('Segoe UI', 48),
                bg=ELITE_COLORS['bg_secondary'],
                fg=color).pack(padx=25, pady=25)
        
        # Mensaje
        tk.Label(message_frame,
                text=message,
                font=('Segoe UI', 12),
                bg=ELITE_COLORS['bg_primary'],
                fg=ELITE_COLORS['text_primary'],
                wraplength=400,
                justify='center').pack()
        
        # Botones en footer
        footer = dialog.get_footer()
        button_container = tk.Frame(footer, bg=ELITE_COLORS['bg_secondary'])
        button_container.pack(pady=15)
        
        ModernButton(button_container,
                    text="✓  Aceptar",
                    style='success',
                    command=lambda: dialog.accept(True),
                    padx=30).pack(side='left', padx=5)
        
        ModernButton(button_container,
                    text="✕  Cancelar",
                    style='secondary',
                    command=dialog.cancel,
                    padx=30).pack(side='left', padx=5)
        
        return dialog.wait_result()
    
    def show_modern_alert(self, title, message, type='info'):
        """Muestra una alerta moderna"""
        dialog = ModernDialog(self.root, title=title, width=450, height=280)
        content = dialog.get_content()
        
        # Icono y mensaje
        message_frame = tk.Frame(content, bg=ELITE_COLORS['bg_primary'])
        message_frame.pack(expand=True, fill='both', padx=30, pady=30)
        
        # Icono
        icons = {
            'info': ('ℹ️', ELITE_COLORS['info']),
            'warning': ('⚠️', ELITE_COLORS['warning']),
            'error': ('✕', ELITE_COLORS['danger']),
            'success': ('✓', ELITE_COLORS['success'])
        }
        
        icon, color = icons.get(type, icons['info'])
        
        icon_container = tk.Frame(message_frame,
                                 bg=ELITE_COLORS['bg_secondary'],
                                 highlightbackground=color,
                                 highlightthickness=2)
        icon_container.pack(pady=(0, 20))
        
        tk.Label(icon_container,
                text=icon,
                font=('Segoe UI', 42),
                bg=ELITE_COLORS['bg_secondary'],
                fg=color).pack(padx=22, pady=22)
        
        # Mensaje
        tk.Label(message_frame,
                text=message,
                font=('Segoe UI', 11),
                bg=ELITE_COLORS['bg_primary'],
                fg=ELITE_COLORS['text_primary'],
                wraplength=380,
                justify='center').pack()
        
        # Botón en footer
        footer = dialog.get_footer()
        button_container = tk.Frame(footer, bg=ELITE_COLORS['bg_secondary'])
        button_container.pack(pady=15)
        
        ModernButton(button_container,
                    text="✓  Entendido",
                    style='primary',
                    command=lambda: dialog.accept(True),
                    padx=40).pack()
        
        return dialog.wait_result()
    
    def show_toast(self, message, type='info', duration=3000):
        """Muestra una notificación toast"""
        ToastNotification.show(self.root, message, type, duration)
    
    def show_loading(self, text="Procesando..."):
        """Muestra un spinner de carga"""
        spinner = LoadingSpinner(self.content_frame, text=text)
        spinner.show()
        return spinner
    
    # ──────────────────────────────────────────────────────────────────────────
    # WIZARD DE PRIMERA EJECUCIÓN
    # ──────────────────────────────────────────────────────────────────────────

    # Tipos de negocio disponibles: (clave, emoji, etiqueta, color acento)
    BUSINESS_TYPES = [
        ("heladeria",     "🍦", "Heladería",          "#00BCD4"),
        ("restaurante",   "🍕", "Restaurante",         "#FF7043"),
        ("ferreteria",    "🔧", "Ferretería",          "#78909C"),
        ("ropa",          "👕", "Ropa / Boutique",     "#AB47BC"),
        ("farmacia",      "💊", "Farmacia",            "#26A69A"),
        ("tecnologia",    "📱", "Tecnología",          "#42A5F5"),
        ("supermercado",  "🛒", "Supermercado",        "#66BB6A"),
        ("belleza",       "💄", "Belleza / Estética",  "#EC407A"),
        ("libreria",      "📚", "Librería",            "#FFA726"),
        ("veterinaria",   "🐾", "Veterinaria",         "#8D6E63"),
        ("jugueteria",    "🎮", "Juguetería",          "#5C6BC0"),
        ("otro",          "🏪", "Otro negocio",        "#607D8B"),
    ]

    def show_setup_wizard(self):
        """Wizard de configuración inicial del negocio"""
        self.clear_content()

        bg = tk.Canvas(self.content_frame, bg=self.colors['bg_primary'], highlightthickness=0)
        bg.pack(fill='both', expand=True)

        # Estado del wizard
        state = {
            'step': 1,              # 1 = tipo negocio, 2 = datos, 3 = confirmar
            'tipo': None,
            'tipo_emoji': None,
            'tipo_label': None,
            'tipo_color': None,
        }

        # ── contenedor central ────────────────────────────────────────────────
        outer = tk.Frame(bg, bg=self.colors['bg_card'],
                         highlightbackground=self.colors['border'],
                         highlightthickness=1)
        _wid = [None]

        def _center(event=None):
            bg.update_idletasks()
            w, h = bg.winfo_width(), bg.winfo_height()
            if w < 10 or h < 10:
                return
            if _wid[0] is None:
                _wid[0] = bg.create_window(w // 2, h // 2, window=outer, anchor='center')
            else:
                bg.coords(_wid[0], w // 2, h // 2)

        bg.bind('<Configure>', _center)
        bg.after(50, _center)

        # ── área de contenido variable ────────────────────────────────────────
        content_area = tk.Frame(outer, bg=self.colors['bg_card'])
        content_area.pack(fill='both', expand=True, padx=50, pady=40)

        def _clear_content_area():
            for w in content_area.winfo_children():
                w.destroy()

        # ── barra de pasos ─────────────────────────────────────────────────────
        def _draw_steps(active_step):
            steps_bar = tk.Frame(content_area, bg=self.colors['bg_card'])
            steps_bar.pack(pady=(0, 24))
            labels = ["Tipo negocio", "Datos", "Tu cuenta", "Confirmar"]
            for i, lbl in enumerate(labels, 1):
                done   = i < active_step
                active = i == active_step
                col = (self.colors['success'] if done else
                       self.colors['accent_blue'] if active else
                       self.colors['text_muted'])
                marker = '✓' if done else str(i)
                circle = tk.Frame(steps_bar, bg=col, width=28, height=28)
                circle.pack(side='left')
                circle.pack_propagate(False)
                tk.Label(circle, text=marker, font=('Segoe UI', 9, 'bold'),
                         bg=col, fg='white').place(relx=.5, rely=.5, anchor='center')
                tk.Label(steps_bar, text=lbl,
                         font=('Segoe UI', 9, 'bold' if active else 'normal'),
                         bg=self.colors['bg_card'], fg=col).pack(side='left', padx=(4, 0))
                if i < len(labels):
                    tk.Frame(steps_bar, bg=self.colors['border'], width=30, height=1
                             ).pack(side='left', padx=6)

        # ══════════════════════════════════════════════════════════════════════
        # PASO 1 — Selección de tipo de negocio
        # ══════════════════════════════════════════════════════════════════════
        def _show_step1():
            _clear_content_area()
            outer.configure(highlightbackground=self.colors['border'])

            # Encabezado
            tk.Label(content_area, text="⚙️  Configuración inicial",
                     font=('Segoe UI', 22, 'bold'),
                     bg=self.colors['bg_card'], fg=self.colors['text_primary']).pack(pady=(0, 4))
            tk.Label(content_area, text="Bienvenido — es la primera vez que inicia la aplicación.",
                     font=('Segoe UI', 11),
                     bg=self.colors['bg_card'], fg=self.colors['text_muted']).pack(pady=(0, 20))

            _draw_steps(1)

            tk.Label(content_area, text="¿Qué tipo de negocio es?",
                     font=('Segoe UI', 13, 'bold'),
                     bg=self.colors['bg_card'], fg=self.colors['text_primary']).pack(pady=(0, 16))

            # Grilla de tarjetas de tipo
            grid_frame = tk.Frame(content_area, bg=self.colors['bg_card'])
            grid_frame.pack()

            selected_frames = {}

            def _select_type(key, emoji, label, color, card):
                state['tipo'] = key
                state['tipo_emoji'] = emoji
                state['tipo_label'] = label
                state['tipo_color'] = color
                for k, f in selected_frames.items():
                    f.configure(highlightbackground=self.colors['border'],
                                highlightthickness=1)
                card.configure(highlightbackground=color, highlightthickness=2)

            cols = 4
            for idx, (key, emoji, label, color) in enumerate(self.BUSINESS_TYPES):
                row, col = divmod(idx, cols)
                card = tk.Frame(grid_frame, bg=self.colors['bg_secondary'],
                                highlightbackground=self.colors['border'],
                                highlightthickness=1,
                                cursor='hand2', width=130, height=90)
                card.grid(row=row, column=col, padx=6, pady=6)
                card.grid_propagate(False)
                selected_frames[key] = card

                tk.Label(card, text=emoji, font=('Segoe UI', 26),
                         bg=self.colors['bg_secondary'],
                         fg=color).place(relx=0.5, rely=0.35, anchor='center')
                tk.Label(card, text=label, font=('Segoe UI', 8, 'bold'),
                         bg=self.colors['bg_secondary'],
                         fg=self.colors['text_secondary'],
                         wraplength=115, justify='center').place(relx=0.5, rely=0.78, anchor='center')

                card.bind('<Button-1>', lambda e, k=key, em=emoji, lb=label, co=color, c=card:
                          _select_type(k, em, lb, co, c))
                for child in card.winfo_children():
                    child.bind('<Button-1>', lambda e, k=key, em=emoji, lb=label, co=color, c=card:
                               _select_type(k, em, lb, co, c))

                # Pre-seleccionar si ya había selección
                if state['tipo'] == key:
                    card.configure(highlightbackground=color, highlightthickness=2)

            # Botón siguiente
            def _go_step2():
                if not state['tipo']:
                    messagebox.showwarning("Advertencia", "Seleccioná un tipo de negocio para continuar.")
                    return
                _show_step2()

            btn_frame = tk.Frame(content_area, bg=self.colors['bg_card'])
            btn_frame.pack(pady=(20, 0))
            _btn(btn_frame, "Siguiente  →", _go_step2)

        # ══════════════════════════════════════════════════════════════════════
        # PASO 2 — Datos del negocio
        # ══════════════════════════════════════════════════════════════════════
        def _show_step2():
            _clear_content_area()

            tk.Label(content_area, text=f"{state['tipo_emoji']}  Datos del negocio",
                     font=('Segoe UI', 22, 'bold'),
                     bg=self.colors['bg_card'], fg=self.colors['text_primary']).pack(pady=(0, 4))
            tk.Label(content_area, text="Completá los datos de tu negocio. Los campos con * son obligatorios.",
                     font=('Segoe UI', 10),
                     bg=self.colors['bg_card'], fg=self.colors['text_muted']).pack(pady=(0, 16))

            _draw_steps(2)

            form = tk.Frame(content_area, bg=self.colors['bg_card'])
            form.pack(fill='x')

            fields = {}
            existing = self.config.get('business', {})

            def _field(parent, label, key, placeholder='', required=False, row=0, col=0, colspan=1):
                lbl_text = f"{'*' if required else '  '} {label}"
                tk.Label(parent, text=lbl_text,
                         font=('Segoe UI', 9, 'bold' if required else 'normal'),
                         bg=self.colors['bg_card'],
                         fg=self.colors['accent_blue'] if required else self.colors['text_secondary'],
                         anchor='w').grid(row=row*2, column=col, columnspan=colspan,
                                          sticky='w', padx=(0, 20), pady=(8, 2))
                entry = tk.Entry(parent, width=28 * colspan,
                                 font=('Segoe UI', 11),
                                 bg=self.colors['bg_secondary'],
                                 fg=self.colors['text_primary'],
                                 insertbackground=self.colors['accent_blue'],
                                 bd=0, relief='flat')
                entry.grid(row=row*2+1, column=col, columnspan=colspan,
                           sticky='ew', padx=(0, 20), ipady=9, ipadx=8)
                val = existing.get(key, '')
                if val:
                    entry.insert(0, str(val))
                elif placeholder:
                    entry.insert(0, placeholder)
                    entry.configure(fg=self.colors['text_muted'])
                    def _clear(e, en=entry, ph=placeholder):
                        if en.get() == ph:
                            en.delete(0, 'end')
                            en.configure(fg=self.colors['text_primary'])
                    def _restore(e, en=entry, ph=placeholder):
                        if not en.get():
                            en.insert(0, ph)
                            en.configure(fg=self.colors['text_muted'])
                    entry.bind('<FocusIn>', _clear)
                    entry.bind('<FocusOut>', _restore)
                tk.Frame(parent, bg=self.colors['border'], height=1).grid(
                    row=row*2+2, column=col, columnspan=colspan, sticky='ew', padx=(0, 20))
                fields[key] = (entry, placeholder)
                return entry

            form.grid_columnconfigure(0, weight=1)
            form.grid_columnconfigure(1, weight=1)

            _field(form, "Nombre del negocio", 'nombre', required=True, row=0, col=0, colspan=2)
            _field(form, "Propietario / Responsable", 'propietario', row=1, col=0)
            _field(form, "CUIT / RUC / NIT / RFC", 'cuit', placeholder='Ej: 20-12345678-9', row=1, col=1)
            _field(form, "Dirección", 'direccion', placeholder='Calle, número, ciudad', row=2, col=0, colspan=2)
            _field(form, "Teléfono", 'telefono', placeholder='Ej: +54 9 11 1234-5678', row=3, col=0)
            _field(form, "Email", 'email', placeholder='contacto@negocio.com', row=3, col=1)
            _field(form, "Moneda", 'moneda', placeholder='Ej: ARS, USD, MXN', row=4, col=0)
            _field(form, "Símbolo moneda", 'simbolo_moneda', placeholder='Ej: $, €, S/', row=4, col=1)
            _field(form, "Mensaje en tickets / recibos", 'mensaje_ticket',
                   placeholder='Ej: ¡Gracias por su compra!', row=5, col=0, colspan=2)

            def _go_step3():
                # Leer valores
                data = {}
                for key, (entry, ph) in fields.items():
                    val = entry.get().strip()
                    data[key] = '' if val == ph else val

                if not data.get('nombre'):
                    messagebox.showwarning("Advertencia", "El nombre del negocio es obligatorio.")
                    return

                data['tipo'] = state['tipo']
                data['tipo_label'] = state['tipo_label']
                data['tipo_emoji'] = state['tipo_emoji']
                data['configured'] = False
                state['form_data'] = data
                _show_step3_user(data)

            btn_frame = tk.Frame(content_area, bg=self.colors['bg_card'])
            btn_frame.pack(pady=(20, 0))
            _btn(btn_frame, "← Atrás", _show_step1, secondary=True)
            _btn(btn_frame, "Siguiente  →", _go_step3)

        # ══════════════════════════════════════════════════════════════════════
        # PASO 3 — Crear usuario administrador
        # ══════════════════════════════════════════════════════════════════════
        def _show_step3_user(data):
            _clear_content_area()

            tk.Label(content_area, text="👤  Creá tu cuenta de administrador",
                     font=('Segoe UI', 20, 'bold'),
                     bg=self.colors['bg_card'], fg=self.colors['text_primary']).pack(pady=(0, 4))
            tk.Label(content_area,
                     text="Estos serán tus datos de acceso. Guardá la contraseña en un lugar seguro.",
                     font=('Segoe UI', 10),
                     bg=self.colors['bg_card'], fg=self.colors['text_muted']).pack(pady=(0, 18))

            _draw_steps(3)

            form = tk.Frame(content_area, bg=self.colors['bg_card'])
            form.pack(fill='x')
            form.grid_columnconfigure(0, weight=1)
            form.grid_columnconfigure(1, weight=1)

            def _entry_row(parent, label, row, col=0, colspan=1, show=''):
                tk.Label(parent, text=label,
                         font=('Segoe UI', 9, 'bold'),
                         bg=self.colors['bg_card'],
                         fg=self.colors['accent_blue'], anchor='w'
                         ).grid(row=row*3, column=col, columnspan=colspan,
                                sticky='w', padx=(0, 20), pady=(10, 2))
                e = tk.Entry(parent, show=show, width=30,
                             font=('Segoe UI', 12),
                             bg=self.colors['bg_secondary'],
                             fg=self.colors['text_primary'],
                             insertbackground=self.colors['accent_blue'],
                             bd=0, relief='flat')
                e.grid(row=row*3+1, column=col, columnspan=colspan,
                       sticky='ew', padx=(0, 20), ipady=10, ipadx=10)
                tk.Frame(parent, bg=self.colors['border'], height=1
                         ).grid(row=row*3+2, column=col, columnspan=colspan,
                                sticky='ew', padx=(0, 20))
                return e

            nombre_e   = _entry_row(form, "* Nombre completo", 0, col=0, colspan=2)
            usuario_e  = _entry_row(form, "* Usuario (para iniciar sesión)", 1, col=0)
            email_e    = _entry_row(form, "Email (opcional)", 1, col=1)
            pass_e     = _entry_row(form, "* Contraseña", 2, col=0, show='●')
            pass2_e    = _entry_row(form, "* Confirmar contraseña", 2, col=1, show='●')

            # Pre-rellenar nombre desde datos del negocio
            propietario = data.get('propietario', '')
            if propietario:
                nombre_e.insert(0, propietario)

            def _go_step4():
                nombre   = nombre_e.get().strip()
                usuario  = usuario_e.get().strip()
                email    = email_e.get().strip()
                password = pass_e.get()
                password2 = pass2_e.get()

                if not nombre:
                    messagebox.showwarning("Advertencia", "Ingresá tu nombre completo.")
                    return
                if not usuario:
                    messagebox.showwarning("Advertencia", "Elegí un nombre de usuario.")
                    return
                if len(usuario) < 3:
                    messagebox.showwarning("Advertencia", "El usuario debe tener al menos 3 caracteres.")
                    return
                if not password:
                    messagebox.showwarning("Advertencia", "Ingresá una contraseña.")
                    return
                if len(password) < 6:
                    messagebox.showwarning("Advertencia", "La contraseña debe tener al menos 6 caracteres.")
                    return
                if password != password2:
                    messagebox.showwarning("Advertencia", "Las contraseñas no coinciden.")
                    return

                state['user_data'] = {
                    'nombre': nombre,
                    'usuario': usuario,
                    'email': email,
                    'password': password,
                }
                _show_step4_confirm(data)

            btn_frame = tk.Frame(content_area, bg=self.colors['bg_card'])
            btn_frame.pack(pady=(20, 0))
            _btn(btn_frame, "← Atrás", _show_step2, secondary=True)
            _btn(btn_frame, "Siguiente  →", _go_step4)

        # ══════════════════════════════════════════════════════════════════════
        # PASO 4 — Confirmación y guardado
        # ══════════════════════════════════════════════════════════════════════
        def _show_step4_confirm(data):
            _clear_content_area()

            color = state['tipo_color']
            emoji = state['tipo_emoji']
            ud    = state.get('user_data', {})

            tk.Label(content_area, text=f"{emoji}  Confirmá la configuración",
                     font=('Segoe UI', 20, 'bold'),
                     bg=self.colors['bg_card'], fg=self.colors['text_primary']).pack(pady=(0, 4))
            tk.Label(content_area, text="Revisá todo antes de guardar. Luego podrás cambiar estos datos en Configuración.",
                     font=('Segoe UI', 10),
                     bg=self.colors['bg_card'], fg=self.colors['text_muted']).pack(pady=(0, 16))

            _draw_steps(4)

            # Dos columnas de resumen
            cols_frame = tk.Frame(content_area, bg=self.colors['bg_card'])
            cols_frame.pack(fill='x', pady=(0, 16))
            cols_frame.grid_columnconfigure(0, weight=1)
            cols_frame.grid_columnconfigure(1, weight=1)

            def _summary_card(parent, title, rows, col):
                card = tk.Frame(parent, bg=self.colors['bg_secondary'],
                                highlightbackground=color, highlightthickness=1)
                card.grid(row=0, column=col, sticky='nsew', padx=(0 if col else 0, 8 if col == 0 else 0))
                tk.Label(card, text=title, font=('Segoe UI', 10, 'bold'),
                         bg=self.colors['bg_secondary'], fg=color).pack(anchor='w', padx=14, pady=(10, 6))
                for lbl, val in rows:
                    row_f = tk.Frame(card, bg=self.colors['bg_secondary'])
                    row_f.pack(fill='x', padx=14, pady=2)
                    tk.Label(row_f, text=f"{lbl}:", font=('Segoe UI', 8, 'bold'),
                             bg=self.colors['bg_secondary'],
                             fg=self.colors['text_muted'], width=14, anchor='w').pack(side='left')
                    tk.Label(row_f, text=val or '—', font=('Segoe UI', 9),
                             bg=self.colors['bg_secondary'],
                             fg=self.colors['text_primary'], anchor='w').pack(side='left')
                tk.Frame(card, height=10, bg=self.colors['bg_secondary']).pack()

            _summary_card(cols_frame, "🏪 Negocio", [
                ("Tipo",      f"{emoji} {state['tipo_label']}"),
                ("Nombre",    data.get('nombre')),
                ("Propiet.",  data.get('propietario')),
                ("CUIT/RUC",  data.get('cuit')),
                ("Dirección", data.get('direccion')),
                ("Teléfono",  data.get('telefono')),
                ("Moneda",    f"{data.get('moneda','')} {data.get('simbolo_moneda','')}".strip()),
            ], col=0)

            _summary_card(cols_frame, "👤 Cuenta admin", [
                ("Nombre",  ud.get('nombre')),
                ("Usuario", ud.get('usuario')),
                ("Email",   ud.get('email') or '—'),
                ("Clave",   '●' * len(ud.get('password', ''))),
            ], col=1)

            def _save_and_finish():
                import bcrypt
                data['configured'] = True
                self.config['business'] = data
                self.config['app']['name'] = f"Sistema TPV — {data['nombre']}"
                if data.get('simbolo_moneda'):
                    self.config.setdefault('currency', {})['symbol'] = data['simbolo_moneda']
                save_config(self.config, 'config.json')

                # Crear usuario admin en la base de datos
                try:
                    pwd = ud['password'].encode('utf-8')
                    hashed = bcrypt.hashpw(pwd, bcrypt.gensalt()).decode('utf-8')
                    # Desactivar usuario admin por defecto si existe
                    self.db.execute_query(
                        "UPDATE usuarios SET activo=0 WHERE username='admin'")
                    # Crear nuevo usuario
                    self.db.execute_query(
                        """INSERT INTO usuarios (username, password_hash, nombre, email, rol, activo)
                           VALUES (?, ?, ?, ?, 'admin', 1)""",
                        (ud['usuario'], hashed, ud['nombre'], ud.get('email') or None))
                except Exception as e:
                    logger.error(f"Error creando usuario: {e}")
                    messagebox.showerror("Error",
                        f"No se pudo crear el usuario: {e}\n\nPodés usar admin / admin123 para entrar.")

                self._apply_business_branding()
                messagebox.showinfo("¡Todo listo! 🎉",
                    f"{emoji} {data['nombre']} configurado correctamente.\n\n"
                    f"Iniciá sesión con:\n  Usuario: {ud['usuario']}\n  Contraseña: tu contraseña")
                outer.destroy()
                self.show_login_screen()

            btn_frame = tk.Frame(content_area, bg=self.colors['bg_card'])
            btn_frame.pack(pady=(12, 0))
            _btn(btn_frame, "← Editar cuenta",
                 lambda: _show_step3_user(state['form_data']), secondary=True)
            _btn(btn_frame, "✅  Guardar y comenzar", _save_and_finish, color=self.colors['success'])

        # Helper para botones del wizard
        def _btn(parent, text, command, secondary=False, color=None):
            bg = color or (self.colors['bg_secondary'] if secondary else self.colors['accent_blue'])
            fg = self.colors['text_secondary'] if secondary else 'white'
            tk.Button(parent, text=text, command=command,
                      font=('Segoe UI', 10 if secondary else 11, 'bold'),
                      bg=bg, fg=fg, bd=0, padx=20 if secondary else 30,
                      pady=8 if secondary else 10, cursor='hand2',
                      relief='flat').pack(side='left', padx=(0, 10))

        # Arrancar en paso 1
        _show_step1()

    def _apply_business_branding(self):
        """Actualiza título de ventana e ícono según el tipo de negocio configurado"""
        biz = self.config.get('business', {})
        nombre = biz.get('nombre', '')
        emoji = biz.get('tipo_emoji', '🏪')
        if nombre:
            self.root.title(f"{emoji} {nombre} — Sistema TPV Elite")
        else:
            self.root.title("🏪 Sistema TPV Elite")

    def show_login_screen(self):
        """Pantalla de login split: panel de marca a la izquierda, formulario a la derecha"""
        self.clear_content()

        biz        = self.config.get('business', {})
        biz_nombre = biz.get('nombre', 'Sistema TPV Elite')
        biz_emoji  = biz.get('tipo_emoji', '🏪')
        biz_color  = biz.get('tipo_color') or self.colors['accent_purple']
        configured = biz.get('configured', False)

        # Contenedor raíz full-screen
        root_frame = tk.Frame(self.content_frame, bg=self.colors['bg_primary'])
        root_frame.place(relx=0, rely=0, relwidth=1, relheight=1)

        # ── Panel izquierdo (marca) ──────────────────────────────────────────
        left = tk.Frame(root_frame, bg=self.colors['secondary_dark'])
        left.place(relx=0, rely=0, relwidth=0.42, relheight=1)

        brand_box = tk.Frame(left, bg=self.colors['secondary_dark'])
        brand_box.place(relx=0.5, rely=0.5, anchor='center')

        # Emoji grande del negocio
        tk.Label(brand_box, text=biz_emoji,
                 font=('Segoe UI', 80),
                 bg=self.colors['secondary_dark'],
                 fg=biz_color).pack(pady=(0, 16))

        # Nombre del negocio
        tk.Label(brand_box, text=biz_nombre,
                 font=('Segoe UI', 22, 'bold'),
                 bg=self.colors['secondary_dark'],
                 fg=self.colors['text_primary'],
                 wraplength=280, justify='center').pack()

        # Subtítulo
        tk.Label(brand_box, text="Sistema de Punto de Venta Elite",
                 font=('Segoe UI', 11),
                 bg=self.colors['secondary_dark'],
                 fg=self.colors['text_muted']).pack(pady=(6, 30))

        # Separador decorativo
        sep = tk.Frame(brand_box, bg=biz_color, height=3, width=60)
        sep.pack()

        # Versión
        tk.Label(brand_box, text="v2.0  •  Gestión profesional",
                 font=('Segoe UI', 9),
                 bg=self.colors['secondary_dark'],
                 fg=self.colors['text_muted']).pack(pady=(14, 0))

        # ── Panel derecho (formulario) ───────────────────────────────────────
        right = tk.Frame(root_frame, bg=self.colors['bg_card'])
        right.place(relx=0.42, rely=0, relwidth=0.58, relheight=1)

        form_box = tk.Frame(right, bg=self.colors['bg_card'])
        form_box.place(relx=0.5, rely=0.5, anchor='center')

        # Encabezado del form
        tk.Label(form_box, text="Iniciar sesión",
                 font=('Segoe UI', 26, 'bold'),
                 bg=self.colors['bg_card'],
                 fg=self.colors['text_primary']).pack(anchor='w', pady=(0, 4))
        tk.Label(form_box, text="Ingresá tus credenciales para continuar",
                 font=('Segoe UI', 11),
                 bg=self.colors['bg_card'],
                 fg=self.colors['text_muted']).pack(anchor='w', pady=(0, 30))

        # ── Campo usuario ────────────────────────────────────────────────────
        def _input_block(label_text, show=''):
            block = tk.Frame(form_box, bg=self.colors['bg_card'])
            block.pack(fill='x', pady=(0, 18))
            tk.Label(block, text=label_text,
                     font=('Segoe UI', 9, 'bold'),
                     bg=self.colors['bg_card'],
                     fg=self.colors['text_muted']).pack(anchor='w', pady=(0, 6))
            entry_wrap = tk.Frame(block, bg=self.colors['bg_secondary'],
                                  highlightbackground=self.colors['border'],
                                  highlightthickness=1)
            entry_wrap.pack(fill='x')
            entry = tk.Entry(entry_wrap, show=show, width=34,
                             font=('Segoe UI', 12),
                             bg=self.colors['bg_secondary'],
                             fg=self.colors['text_primary'],
                             insertbackground=biz_color,
                             bd=0, relief='flat')
            entry.pack(fill='x', ipady=12, ipadx=14)
            # Resaltar borde al enfocar
            entry.bind('<FocusIn>',  lambda e: entry_wrap.configure(highlightbackground=biz_color))
            entry.bind('<FocusOut>', lambda e: entry_wrap.configure(highlightbackground=self.colors['border']))
            return entry

        username_entry = _input_block("USUARIO")
        username_entry.focus()
        password_entry = _input_block("CONTRASEÑA", show='●')

        # ── Botón login ──────────────────────────────────────────────────────
        login_btn = tk.Button(form_box,
                              text="  INICIAR SESIÓN",
                              font=('Segoe UI', 12, 'bold'),
                              bg=biz_color, fg='white',
                              bd=0, padx=0, pady=14,
                              cursor='hand2', relief='flat',
                              activebackground=self.colors['accent_blue'],
                              activeforeground='white')
        login_btn.pack(fill='x', pady=(10, 6))

        login_btn.bind('<Enter>', lambda e: login_btn.configure(bg=self.colors['accent_blue']))
        login_btn.bind('<Leave>', lambda e: login_btn.configure(bg=biz_color))

        # ── Info de credenciales ─────────────────────────────────────────────
        info_frame = tk.Frame(form_box, bg=self.colors['bg_secondary'],
                              highlightbackground=self.colors['border'],
                              highlightthickness=1)
        info_frame.pack(fill='x', pady=(16, 0))
        if not configured:
            hint = "Usuario por defecto: admin  |  Contraseña: admin123"
        else:
            hint = f"Accedé con el usuario creado durante la configuración"
        tk.Label(info_frame, text=f"ℹ  {hint}",
                 font=('Segoe UI', 9),
                 bg=self.colors['bg_secondary'],
                 fg=self.colors['text_muted']).pack(padx=14, pady=10)

        # ── Lógica de login ──────────────────────────────────────────────────
        def attempt_login():
            username = username_entry.get().strip()
            password = password_entry.get()
            if not username or not password:
                AnimationHelper.shake(form_box, duration=400)
                messagebox.showwarning("Advertencia", "Ingresá usuario y contraseña")
                return
            success, message, user = self.auth_service.login(username, password)
            if success:
                logger.info(f"✅ Login exitoso: {username}")
                self.show_main_app()
            else:
                logger.warning(f"❌ Login fallido: {username}")
                AnimationHelper.shake(form_box, duration=400)
                messagebox.showerror("Error de Acceso", message)
                password_entry.delete(0, 'end')

        login_btn.configure(command=attempt_login)
        username_entry.bind('<Return>', lambda e: password_entry.focus())
        password_entry.bind('<Return>', lambda e: attempt_login())

        # Animación de entrada
        AnimationHelper.fade_in(right, duration=500)

    def show_main_app(self):
        """Muestra la aplicación principal después del login"""
        # Limpiar content frame
        self.clear_content()

        # Reconfigurar grid del main_container para modo app (header + sidebar + content)
        self.main_container.grid_columnconfigure(0, weight=0)  # sidebar fijo
        self.main_container.grid_columnconfigure(1, weight=1)  # content expande
        self.main_container.grid_rowconfigure(0, weight=0)     # header fijo
        self.main_container.grid_rowconfigure(1, weight=1)     # content expande

        # Reubicar content_frame a su posición dentro del app (col 1, row 1)
        self.content_frame.grid(row=1, column=1, sticky='nsew', padx=0, pady=0)

        # Mostrar header y sidebar
        self._create_header()
        self._create_sidebar()

        # Mostrar dashboard
        self.show_dashboard()
    
    def logout(self):
        """Cierra sesión"""
        if messagebox.askyesno("Cerrar Sesión", "¿Está seguro que desea cerrar sesión?"):
            self.auth_service.logout()
            
            # Ocultar header y sidebar
            self.header_frame.grid_forget()
            self.sidebar_frame.grid_forget()

            # Restaurar grid del main_container para modo login (content ocupa todo)
            self.main_container.grid_columnconfigure(0, weight=1)
            self.main_container.grid_columnconfigure(1, weight=0)
            self.main_container.grid_rowconfigure(0, weight=1)
            self.main_container.grid_rowconfigure(1, weight=0)
            self.content_frame.grid(row=0, column=0, rowspan=2, sticky='nsew', padx=0, pady=0)

            # Mostrar login
            self.show_login_screen()

            logger.info("Sesión cerrada")
    
    # ══════════════════════════════════════════════════════════════════════════
    # HELPERS DE UI — usados por todas las secciones
    # ══════════════════════════════════════════════════════════════════════════

    def _ui_page_header(self, parent, icon, title, subtitle=None):
        """Encabezado de sección consistente con icono, título y subtítulo."""
        bar = tk.Frame(parent, bg=self.colors['bg_primary'])
        bar.pack(fill='x', padx=24, pady=(20, 0))

        left = tk.Frame(bar, bg=self.colors['bg_primary'])
        left.pack(side='left')

        tk.Label(left, text=icon, font=('Segoe UI', 26),
                 bg=self.colors['bg_primary'],
                 fg=self.colors['accent_blue']).pack(side='left', padx=(0, 12))

        titles = tk.Frame(left, bg=self.colors['bg_primary'])
        titles.pack(side='left')
        tk.Label(titles, text=title, font=('Segoe UI', 22, 'bold'),
                 bg=self.colors['bg_primary'],
                 fg=self.colors['text_primary']).pack(anchor='w')
        if subtitle:
            tk.Label(titles, text=subtitle, font=('Segoe UI', 10),
                     bg=self.colors['bg_primary'],
                     fg=self.colors['text_muted']).pack(anchor='w')

        # Línea separadora
        tk.Frame(parent, bg=self.colors['border'], height=1).pack(
            fill='x', padx=24, pady=(10, 0))
        return bar

    def _ui_section_card(self, parent, bg=None):
        """Card contenedor con borde sutil."""
        bg = bg or self.colors['bg_card']
        card = tk.Frame(parent, bg=bg,
                        highlightbackground=self.colors['border'],
                        highlightthickness=1)
        return card

    def _ui_search_bar(self, parent, placeholder='Buscar...', on_change=None, on_enter=None):
        """Barra de búsqueda consistente. Devuelve el Entry."""
        wrap = tk.Frame(parent, bg=self.colors['bg_secondary'],
                        highlightbackground=self.colors['accent_blue'],
                        highlightthickness=2)
        wrap.pack(fill='x')

        tk.Label(wrap, text='🔍', font=('Segoe UI', 13),
                 bg=self.colors['bg_secondary'],
                 fg=self.colors['accent_blue']).pack(side='left', padx=(12, 4))

        e = tk.Entry(wrap, font=('Segoe UI', 11),
                     bg=self.colors['bg_secondary'],
                     fg=self.colors['text_muted'],
                     insertbackground=self.colors['accent_blue'],
                     bd=0, relief='flat')
        e.insert(0, placeholder)
        e.pack(side='left', fill='x', expand=True, pady=10, padx=(0, 10))

        def _focus_in(ev):
            if e.get() == placeholder:
                e.delete(0, 'end')
                e.configure(fg=self.colors['text_primary'])
        def _focus_out(ev):
            if not e.get():
                e.insert(0, placeholder)
                e.configure(fg=self.colors['text_muted'])

        e.bind('<FocusIn>',  _focus_in)
        e.bind('<FocusOut>', _focus_out)
        if on_change:
            e.bind('<KeyRelease>', lambda ev: on_change(e.get() if e.get() != placeholder else ''))
        if on_enter:
            e.bind('<Return>', lambda ev: on_enter(e.get() if e.get() != placeholder else ''))

        wrap.bind('<FocusIn>',  lambda ev: wrap.configure(highlightbackground=self.colors['accent_blue']))
        return e

    def _ui_action_btn(self, parent, text, command, style='primary', side='left', padx=(0, 8)):
        """Botón de acción consistente."""
        colors_map = {
            'primary':  (self.colors['accent_blue'],   'white'),
            'success':  (self.colors['success'],       self.colors['primary_dark']),
            'danger':   (self.colors['danger'],        'white'),
            'warning':  (self.colors['warning'],       self.colors['primary_dark']),
            'secondary':(self.colors['bg_secondary'],  self.colors['text_secondary']),
            'purple':   (self.colors['accent_purple'], 'white'),
        }
        bg, fg = colors_map.get(style, colors_map['primary'])
        btn = tk.Button(parent, text=text, command=command,
                        font=('Segoe UI', 10, 'bold'),
                        bg=bg, fg=fg, bd=0, padx=16, pady=8,
                        cursor='hand2', relief='flat',
                        activebackground=self.colors['accent_blue'],
                        activeforeground='white')
        btn.pack(side=side, padx=padx)
        btn.bind('<Enter>', lambda e: btn.configure(bg=self.colors['accent_blue'], fg='white'))
        btn.bind('<Leave>', lambda e: btn.configure(bg=bg, fg=fg))
        return btn

    def _ui_apply_stripes(self, tree):
        """Aplica colores alternados a un Treeview ya cargado."""
        even_bg = self.colors.get('bg_secondary', '#1a2035')
        odd_bg  = self.colors.get('bg_card',      '#1e2a3a')
        tree.tag_configure('even', background=even_bg)
        tree.tag_configure('odd',  background=odd_bg)
        for i, child in enumerate(tree.get_children()):
            current_tags = list(tree.item(child, 'tags'))
            current_tags = [t for t in current_tags if t not in ('even', 'odd')]
            current_tags.append('even' if i % 2 == 0 else 'odd')
            tree.item(child, tags=current_tags)

    def _ui_pagination_bar(self, parent, on_prev, on_next):
        """Barra de paginación consistente. Devuelve (frame, label, {prev, next})."""
        bar = tk.Frame(parent, bg=self.colors['bg_card'])
        bar.pack(fill='x', padx=16, pady=(8, 12))

        prev_btn = tk.Button(bar, text='◀  Anterior',
                             command=on_prev,
                             font=('Segoe UI', 9, 'bold'),
                             bg=self.colors['bg_secondary'],
                             fg=self.colors['text_secondary'],
                             bd=0, padx=12, pady=6, cursor='hand2', relief='flat')
        prev_btn.pack(side='left')

        lbl = tk.Label(bar, text='Página 1 de 1',
                       font=('Segoe UI', 9),
                       bg=self.colors['bg_card'],
                       fg=self.colors['text_muted'])
        lbl.pack(side='left', expand=True)

        next_btn = tk.Button(bar, text='Siguiente  ▶',
                             command=on_next,
                             font=('Segoe UI', 9, 'bold'),
                             bg=self.colors['bg_secondary'],
                             fg=self.colors['text_secondary'],
                             bd=0, padx=12, pady=6, cursor='hand2', relief='flat')
        next_btn.pack(side='right')

        return bar, lbl, {'prev': prev_btn, 'next': next_btn}

    def _ui_styled_tree(self, parent, columns, col_configs=None, height=None):
        """Treeview estilizado listo para usar. col_configs = {col: (heading, width, anchor)}."""
        frame = tk.Frame(parent, bg=self.colors['bg_card'])
        frame.pack(fill='both', expand=True)
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(0, weight=1)

        kw = dict(columns=columns, show='headings', style='Modern.Treeview')
        if height:
            kw['height'] = height
        tree = ttk.Treeview(frame, **kw)

        col_configs = col_configs or {}
        for col in columns:
            cfg = col_configs.get(col, {})
            tree.heading(col, text=cfg.get('text', col))
            tree.column(col,
                        width=cfg.get('width', 120),
                        anchor=cfg.get('anchor', 'w'),
                        minwidth=cfg.get('minwidth', 40))

        tree.grid(row=0, column=0, sticky='nsew')

        sb = ttk.Scrollbar(frame, orient='vertical', command=tree.yview,
                           style='Elite.Vertical.TScrollbar')
        sb.grid(row=0, column=1, sticky='ns')
        tree.configure(yscrollcommand=sb.set)

        # Tags de color
        self._ui_apply_stripes(tree)  # preconfigura los tags aunque esté vacío
        return tree

    # ══════════════════════════════════════════════════════════════════════════

    def show_dashboard(self):
        """Muestra el dashboard principal ELITE"""
        self.clear_content()
        self.current_screen = "dashboard"
        
        # Crear canvas con scrollbar elite
        canvas = tk.Canvas(self.content_frame, bg=self.colors['bg_primary'],
                          highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.content_frame, orient="vertical",
                                 command=canvas.yview, style='Elite.Vertical.TScrollbar')
        scrollable_frame = tk.Frame(canvas, bg=self.colors['bg_primary'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Smooth scroll con rueda del mouse
        def _on_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except:
                pass
        
        # Usar bind en lugar de bind_all para evitar conflictos
        canvas.bind("<MouseWheel>", _on_mousewheel)
        scrollable_frame.bind("<MouseWheel>", _on_mousewheel)
        
        # Limpiar binding cuando se destruye el canvas
        def _on_destroy(event):
            try:
                canvas.unbind("<MouseWheel>")
                scrollable_frame.unbind("<MouseWheel>")
            except:
                pass
        canvas.bind("<Destroy>", _on_destroy)
        
        # Header del dashboard
        header_frame = tk.Frame(scrollable_frame, bg=self.colors['bg_primary'])
        header_frame.pack(fill='x', padx=30, pady=(25, 20))
        
        # Título con efecto
        title_container = tk.Frame(header_frame, bg=self.colors['bg_primary'])
        title_container.pack(side='left')
        
        icon_title = tk.Label(title_container,
                             text="📊",
                             font=('Segoe UI', 28),
                             bg=self.colors['bg_primary'],
                             fg=self.colors['accent_blue'])
        icon_title.pack(side='left', padx=(0, 15))
        
        title = tk.Label(title_container,
                        text="Panel de Control",
                        font=('Segoe UI', 26, 'bold'),
                        bg=self.colors['bg_primary'],
                        fg=self.colors['text_primary'])
        title.pack(side='left')
        
        subtitle = tk.Label(title_container,
                           text="Vista general del sistema",
                           font=('Segoe UI', 11),
                           bg=self.colors['bg_primary'],
                           fg=self.colors['text_muted'])
        subtitle.pack(side='left', padx=(15, 0))
        
        # Fecha actual
        fecha_label = tk.Label(header_frame,
                              text=datetime.now().strftime("%d/%m/%Y • %H:%M"),
                              font=('Segoe UI', 12),
                              bg=self.colors['bg_primary'],
                              fg=self.colors['text_secondary'])
        fecha_label.pack(side='right')
        
        # Contenedor de tarjetas con grid
        cards_container = tk.Frame(scrollable_frame, bg=self.colors['bg_primary'])
        cards_container.pack(fill='both', expand=True, padx=20, pady=10)
        cards_container.grid_columnconfigure((0, 1, 2), weight=1)
        
        # Obtener estadísticas
        stats = self._get_dashboard_stats()
        
        # ========== PRIMERA FILA: KPIs PRINCIPALES ==========
        # Ventas de hoy con tendencia
        self._create_enhanced_kpi_card(
            cards_container,
            title="Ventas de Hoy",
            value=format_currency(stats['ventas_hoy']),
            subtitle=f"{stats['cant_ventas_hoy']} transacciones • Promedio: {format_currency(stats['ventas_hoy']/stats['cant_ventas_hoy'] if stats['cant_ventas_hoy'] > 0 else 0)}",
            trend=stats['tendencia_dia'],
            icon="💰",
            color=self.colors['success'],
            row=0, col=0
        )
        
        # Ventas del mes
        self._create_enhanced_kpi_card(
            cards_container,
            title="Ventas del Mes",
            value=format_currency(stats['ventas_mes']),
            subtitle=f"{stats['cant_ventas_mes']} ventas • Promedio diario: {format_currency(stats['promedio_diario'])}",
            trend=None,
            icon="📊",
            color=self.colors['accent_purple'],
            row=0, col=1
        )
        
        # Productos bajo stock con alertas
        stock_warning = "⚠️ CRÍTICO" if stats['productos_criticos'] > 0 else "✓ Normal"
        self._create_enhanced_kpi_card(
            cards_container,
            title="Control de Stock",
            value=str(stats['productos_bajo_stock']),
            subtitle=f"{stats['productos_criticos']} productos críticos • {stock_warning}",
            trend=None,
            icon="📦",
            color=self.colors['danger'] if stats['productos_criticos'] > 0 else self.colors['warning'],
            row=0, col=2
        )
        
        # ========== SEGUNDA FILA: TOP PRODUCTOS ==========
        self._create_top_products_widget(cards_container, stats['top_productos'], row=1, col=0, colspan=3)
        
        # ========== TERCERA FILA: GRÁFICOS ==========
        # Gráfico de ventas últimos 7 días
        self._create_sales_chart(cards_container, stats['ventas_ultimos_7_dias'])
        
        # 💳 Gráfico de métodos de pago
        self._create_payment_methods_dashboard_chart(cards_container, stats['metodos_pago'])
        
        # Animación de entrada
        AnimationHelper.fade_in(scrollable_frame, duration=400)
    
    def _get_dashboard_stats(self):
        """Obtiene las estadísticas para el dashboard"""
        try:
            hoy = datetime.now().date()
            ayer = hoy - timedelta(days=1)
            inicio_mes = datetime(hoy.year, hoy.month, 1)
            
            # Ventas de hoy
            ventas_hoy_rows = self.db.fetch_all(
                "SELECT SUM(total) as total, COUNT(*) as cantidad FROM ventas WHERE DATE(fecha) = ? AND estado = 'completada'",
                (hoy,)
            )
            ventas_hoy = ventas_hoy_rows[0]['total'] if ventas_hoy_rows and ventas_hoy_rows[0]['total'] else 0
            cant_ventas_hoy = ventas_hoy_rows[0]['cantidad'] if ventas_hoy_rows and ventas_hoy_rows[0]['cantidad'] else 0
            
            # Ventas de ayer (para comparación)
            ventas_ayer_rows = self.db.fetch_all(
                "SELECT SUM(total) as total FROM ventas WHERE DATE(fecha) = ? AND estado = 'completada'",
                (ayer,)
            )
            ventas_ayer = ventas_ayer_rows[0]['total'] if ventas_ayer_rows and ventas_ayer_rows[0]['total'] else 0
            
            # Calcular tendencia diaria
            if ventas_ayer > 0:
                tendencia_dia = ((ventas_hoy - ventas_ayer) / ventas_ayer) * 100
            else:
                tendencia_dia = 100 if ventas_hoy > 0 else 0
            
            # Ventas del mes
            ventas_mes_rows = self.db.fetch_all(
                "SELECT SUM(total) as total, COUNT(*) as cantidad FROM ventas WHERE fecha >= ? AND estado = 'completada'",
                (inicio_mes,)
            )
            ventas_mes = ventas_mes_rows[0]['total'] if ventas_mes_rows and ventas_mes_rows[0]['total'] else 0
            cant_ventas_mes = ventas_mes_rows[0]['cantidad'] if ventas_mes_rows and ventas_mes_rows[0]['cantidad'] else 0
            
            # Promedio de venta diaria del mes
            dias_transcurridos = (datetime.now() - inicio_mes).days + 1
            promedio_diario = ventas_mes / dias_transcurridos if dias_transcurridos > 0 else 0
            
            # Productos bajo stock
            productos_bajo_stock = len(self.producto_service.productos_bajo_stock())
            productos_criticos = len([p for p in self.producto_service.productos_bajo_stock() if p.stock <= 5])
            
            # Top 5 productos más vendidos del mes
            top_productos = self.db.fetch_all("""
                SELECT 
                    p.nombre,
                    SUM(dv.cantidad) as total_vendido,
                    SUM(dv.subtotal) as ingresos
                FROM detalles_venta dv
                JOIN productos p ON dv.producto_id = p.id
                JOIN ventas v ON dv.venta_id = v.id
                WHERE v.fecha >= ? AND v.estado = 'completada'
                GROUP BY p.id, p.nombre
                ORDER BY total_vendido DESC
                LIMIT 5
            """, (inicio_mes,))
            
            # Ventas últimos 7 días
            ventas_7_dias = []
            for i in range(6, -1, -1):
                fecha = hoy - timedelta(days=i)
                rows = self.db.fetch_all(
                    "SELECT SUM(total) as total FROM ventas WHERE DATE(fecha) = ? AND estado = 'completada'",
                    (fecha,)
                )
                total = rows[0]['total'] if rows and rows[0]['total'] else 0
                ventas_7_dias.append({
                    'fecha': fecha.strftime('%d/%m'),
                    'total': total
                })
            
            # 💳 Ventas por método de pago (del mes actual)
            metodos_pago = self.db.fetch_all("""
                SELECT 
                    metodo_pago,
                    COUNT(*) as cantidad,
                    SUM(total) as total
                FROM ventas 
                WHERE fecha >= ? AND estado = 'completada'
                GROUP BY metodo_pago
                ORDER BY total DESC
            """, (inicio_mes,))
            
            return {
                'ventas_hoy': ventas_hoy,
                'cant_ventas_hoy': cant_ventas_hoy,
                'tendencia_dia': tendencia_dia,
                'ventas_mes': ventas_mes,
                'cant_ventas_mes': cant_ventas_mes,
                'promedio_diario': promedio_diario,
                'productos_bajo_stock': productos_bajo_stock,
                'productos_criticos': productos_criticos,
                'top_productos': top_productos if top_productos else [],
                'ventas_ultimos_7_dias': ventas_7_dias,
                'metodos_pago': metodos_pago if metodos_pago else []
            }
        except Exception as e:
            logger.error(f"Error al obtener estadísticas: {e}")
            return {
                'ventas_hoy': 0,
                'cant_ventas_hoy': 0,
                'tendencia_dia': 0,
                'ventas_mes': 0,
                'cant_ventas_mes': 0,
                'promedio_diario': 0,
                'productos_bajo_stock': 0,
                'productos_criticos': 0,
                'top_productos': [],
                'ventas_ultimos_7_dias': [],
                'metodos_pago': []
            }
    
    def _create_stat_card(self, parent, title, value, color, row, col):
        """Crea una tarjeta de estadística ELITE con animaciones"""
        # Frame externo con sombra simulada
        outer_frame = tk.Frame(parent, bg=self.colors['bg_primary'])
        outer_frame.grid(row=row, column=col, padx=15, pady=15, sticky='nsew')
        
        # Card principal con efecto hover
        card = tk.Frame(outer_frame,
                       bg=self.colors['bg_card'],
                       highlightbackground=self.colors['border'],
                       highlightthickness=1)
        card.pack(fill='both', expand=True, padx=2, pady=2)
        
        # Frame interior
        inner = tk.Frame(card, bg=self.colors['bg_card'])
        inner.pack(fill='both', expand=True, padx=25, pady=25)
        
        # Icono según el título
        icons = {
            "Ventas Hoy": "💰",
            "Ventas del Mes": "📊",
            "Productos Bajo Stock": "⚠️"
        }
        icon = icons.get(title, "📈")
        
        # Contenedor superior con icono
        top_frame = tk.Frame(inner, bg=self.colors['bg_card'])
        top_frame.pack(fill='x')
        
        icon_container = tk.Frame(top_frame,
                                 bg=self.colors['bg_secondary'],
                                 highlightbackground=color,
                                 highlightthickness=2)
        icon_container.pack(side='left')
        
        icon_label = tk.Label(icon_container,
                             text=icon,
                             font=('Segoe UI', 24),
                             bg=self.colors['bg_secondary'],
                             fg=color)
        icon_label.pack(padx=12, pady=12)
        
        # Título
        title_label = tk.Label(inner,
                              text=title.upper(),
                              font=('Segoe UI', 10, 'bold'),
                              bg=self.colors['bg_card'],
                              fg=self.colors['text_muted'],
                              anchor='w')
        title_label.pack(fill='x', pady=(15, 5))
        
        # Valor con efecto brillante
        value_label = tk.Label(inner,
                              text=value,
                              font=('Segoe UI', 32, 'bold'),
                              bg=self.colors['bg_card'],
                              fg=color,
                              anchor='w')
        value_label.pack(fill='x', pady=(0, 10))
        
        # Barra decorativa inferior
        bottom_bar = tk.Frame(inner, bg=color, height=3)
        bottom_bar.pack(fill='x', side='bottom')
        
        # Efectos hover
        def on_enter(e):
            card.configure(highlightbackground=color, highlightthickness=2)
            AnimationHelper.pulse(card, duration=400, scale=1.02)
        
        def on_leave(e):
            card.configure(highlightbackground=self.colors['border'], highlightthickness=1)
        
        card.bind('<Enter>', on_enter)
        card.bind('<Leave>', on_leave)
        
        # Animación de entrada
        AnimationHelper.fade_in(card, duration=500)
    
    def _create_enhanced_kpi_card(self, parent, title, value, subtitle, trend, icon, color, row, col):
        """Crea una tarjeta KPI mejorada con tendencia y detalles"""
        # Frame externo
        outer_frame = tk.Frame(parent, bg=self.colors['bg_primary'])
        outer_frame.grid(row=row, column=col, padx=12, pady=12, sticky='nsew')
        
        # Card principal
        card = tk.Frame(outer_frame,
                       bg=self.colors['bg_card'],
                       highlightbackground=self.colors['border'],
                       highlightthickness=1)
        card.pack(fill='both', expand=True, padx=2, pady=2)
        
        # Frame interior
        inner = tk.Frame(card, bg=self.colors['bg_card'])
        inner.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Header con icono y tendencia
        header_frame = tk.Frame(inner, bg=self.colors['bg_card'])
        header_frame.pack(fill='x', pady=(0, 10))
        
        # Icono con fondo
        icon_bg = tk.Frame(header_frame, bg=color, width=48, height=48)
        icon_bg.pack(side='left')
        icon_bg.pack_propagate(False)
        
        icon_label = tk.Label(icon_bg,
                             text=icon,
                             font=('Segoe UI', 22),
                             bg=color,
                             fg='white')
        icon_label.pack(expand=True)
        
        # Tendencia (si existe)
        if trend is not None:
            trend_frame = tk.Frame(header_frame, bg=self.colors['bg_secondary'])
            trend_frame.pack(side='right')
            
            trend_color = self.colors['success'] if trend >= 0 else self.colors['danger']
            trend_icon = "↑" if trend >= 0 else "↓"
            
            tk.Label(trend_frame,
                    text=f"{trend_icon} {abs(trend):.1f}%",
                    font=('Segoe UI', 10, 'bold'),
                    bg=self.colors['bg_secondary'],
                    fg=trend_color).pack(padx=10, pady=5)
        
        # Título
        tk.Label(inner,
                text=title.upper(),
                font=('Segoe UI', 9, 'bold'),
                bg=self.colors['bg_card'],
                fg=self.colors['text_muted'],
                anchor='w').pack(fill='x', pady=(10, 5))
        
        # Valor principal
        tk.Label(inner,
                text=value,
                font=('Segoe UI', 28, 'bold'),
                bg=self.colors['bg_card'],
                fg=color,
                anchor='w').pack(fill='x')
        
        # Subtítulo con info adicional
        if subtitle:
            tk.Label(inner,
                    text=subtitle,
                    font=('Segoe UI', 9),
                    bg=self.colors['bg_card'],
                    fg=self.colors['text_secondary'],
                    anchor='w').pack(fill='x', pady=(5, 0))
        
        # Barra decorativa
        tk.Frame(inner, bg=color, height=3).pack(fill='x', side='bottom', pady=(10, 0))
        
        # Efectos hover
        def on_enter(e):
            card.configure(highlightbackground=color, highlightthickness=2)
            AnimationHelper.pulse(card, duration=300, scale=1.02)
        
        def on_leave(e):
            card.configure(highlightbackground=self.colors['border'], highlightthickness=1)
        
        card.bind('<Enter>', on_enter)
        card.bind('<Leave>', on_leave)
        
        # Animación de entrada con delay
        self.root.after(row * 100 + col * 50, lambda: AnimationHelper.fade_in(card, duration=400))
    
    def _create_top_products_widget(self, parent, products, row, col, colspan=3):
        """Crea un widget con los top productos más vendidos"""
        # Frame externo
        outer_frame = tk.Frame(parent, bg=self.colors['bg_primary'])
        outer_frame.grid(row=row, column=col, columnspan=colspan, padx=12, pady=12, sticky='nsew')
        
        # Card principal
        card = tk.Frame(outer_frame,
                       bg=self.colors['bg_card'],
                       highlightbackground=self.colors['border'],
                       highlightthickness=1)
        card.pack(fill='both', expand=True, padx=2, pady=2)
        
        # Header
        header = tk.Frame(card, bg=self.colors['bg_secondary'])
        header.pack(fill='x')
        
        tk.Label(header,
                text="🏆  TOP 5 PRODUCTOS MÁS VENDIDOS",
                font=('Segoe UI', 12, 'bold'),
                bg=self.colors['bg_secondary'],
                fg=self.colors['text_primary']).pack(side='left', padx=20, pady=15)
        
        # Contenido
        content = tk.Frame(card, bg=self.colors['bg_card'])
        content.pack(fill='both', expand=True, padx=20, pady=15)
        
        if not products:
            tk.Label(content,
                    text="No hay datos de ventas disponibles",
                    font=('Segoe UI', 11),
                    bg=self.colors['bg_card'],
                    fg=self.colors['text_muted']).pack(pady=40)
            return
        
        # Lista de productos con barras de progreso
        max_vendido = max([p['total_vendido'] for p in products]) if products else 1
        
        for i, producto in enumerate(products):
            # Frame del producto
            prod_frame = tk.Frame(content, bg=self.colors['bg_card'])
            prod_frame.pack(fill='x', pady=8)
            
            # Ranking
            ranking_colors = [
                self.colors['accent_orange'],  # Oro
                self.colors['text_muted'],      # Plata
                self.colors['accent_orange'],   # Bronce
                self.colors['accent_blue'],
                self.colors['accent_purple']
            ]
            
            tk.Label(prod_frame,
                    text=f"#{i+1}",
                    font=('Segoe UI', 14, 'bold'),
                    bg=self.colors['bg_card'],
                    fg=ranking_colors[i],
                    width=3).pack(side='left')
            
            # Info del producto
            info_frame = tk.Frame(prod_frame, bg=self.colors['bg_card'])
            info_frame.pack(side='left', fill='both', expand=True, padx=10)
            
            # Nombre y cantidad
            top_info = tk.Frame(info_frame, bg=self.colors['bg_card'])
            top_info.pack(fill='x')
            
            tk.Label(top_info,
                    text=producto['nombre'],
                    font=('Segoe UI', 11, 'bold'),
                    bg=self.colors['bg_card'],
                    fg=self.colors['text_primary'],
                    anchor='w').pack(side='left')
            
            tk.Label(top_info,
                    text=f"{int(producto['total_vendido'])} unidades • {format_currency(producto['ingresos'])}",
                    font=('Segoe UI', 9),
                    bg=self.colors['bg_card'],
                    fg=self.colors['text_secondary'],
                    anchor='e').pack(side='right')
            
            # Barra de progreso
            progress_bg = tk.Frame(info_frame, bg=self.colors['bg_secondary'], height=8)
            progress_bg.pack(fill='x', pady=(5, 0))
            
            progress_width = (producto['total_vendido'] / max_vendido)
            progress_bar = tk.Frame(progress_bg, bg=ranking_colors[i], height=8)
            progress_bar.place(relwidth=progress_width, rely=0, relheight=1)
        
        # Animación de entrada
        AnimationHelper.fade_in(card, duration=500)
    
    def _create_sales_chart(self, parent, data):
        """Crea un gráfico de ventas"""
        chart_card = ttk.Frame(parent, style='Card.TFrame')
        chart_card.grid(row=2, column=0, columnspan=3, padx=10, pady=10, sticky='nsew')
        
        # Título
        title = ttk.Label(chart_card,
                         text="Ventas de los últimos 7 días",
                         style='Subtitle.TLabel')
        title.pack(pady=15)
        
        # Verificar si matplotlib está disponible
        if not MATPLOTLIB_AVAILABLE or Figure is None or FigureCanvasTkAgg is None:
            # Mostrar mensaje alternativo
            msg = ttk.Label(chart_card,
                          text="Gráficos no disponibles (matplotlib no instalado)",
                          style='Info.TLabel')
            msg.pack(pady=50)
            return
        
        try:
            # Crear figura
            fig = Figure(figsize=(10, 4), dpi=80)
            ax = fig.add_subplot(111)
            
            # Datos
            fechas = [d['fecha'] for d in data]
            totales = [d['total'] for d in data]
            
            # Gráfico de barras
            ax.bar(fechas, totales, color=self.colors.get('accent', '#3498DB'), alpha=0.8)
            ax.set_xlabel('Fecha', fontsize=10)
            ax.set_ylabel('Total ($)', fontsize=10)
            ax.grid(axis='y', alpha=0.3)
            
            # Formatear ejes
            ax.tick_params(axis='both', labelsize=9)
            fig.tight_layout()
            fig.patch.set_facecolor('white')
            ax.set_facecolor('white')
            
            # Integrar con tkinter
            canvas = FigureCanvasTkAgg(fig, chart_card)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True, padx=20, pady=(0, 20))
        except Exception as e:
            logger.error(f"Error al crear gráfico: {e}")
            msg = ttk.Label(chart_card,
                          text="Error al crear el gráfico",
                          style='Info.TLabel')
            msg.pack(pady=50)
    
    def _create_payment_methods_dashboard_chart(self, parent, data):
        """Crea un gráfico de métodos de pago en el dashboard"""
        if not data:
            return  # No mostrar nada si no hay datos
        
        chart_card = ttk.Frame(parent, style='Card.TFrame')
        chart_card.grid(row=3, column=0, columnspan=3, padx=10, pady=10, sticky='nsew')
        
        # Título
        title = ttk.Label(chart_card,
                         text="💳 Ventas por Método de Pago (Mes Actual)",
                         style='Subtitle.TLabel')
        title.pack(pady=15)
        
        # Verificar si matplotlib está disponible
        if not MATPLOTLIB_AVAILABLE or Figure is None or FigureCanvasTkAgg is None:
            msg = ttk.Label(chart_card,
                          text="Gráficos no disponibles (matplotlib no instalado)",
                          style='Info.TLabel')
            msg.pack(pady=30)
            return
        
        try:
            # Diccionario de nombres amigables
            metodo_nombres = {
                'efectivo': 'Efectivo',
                'tarjeta_credito': 'T. Crédito',
                'tarjeta_debito': 'T. Débito',
                'transferencia': 'Transferencia',
                'otro': 'Otro'
            }
            
            # Preparar datos
            metodos = [metodo_nombres.get(d['metodo_pago'], d['metodo_pago']) for d in data]
            totales = [d['total'] for d in data]
            cantidades = [d['cantidad'] for d in data]
            
            # Crear figura con 2 subgráficos
            fig = Figure(figsize=(10, 4), dpi=80)
            
            # Gráfico de barras - Montos
            ax1 = fig.add_subplot(121)
            colors = ['#27AE60', '#3498DB', '#E74C3C', '#F39C12', '#9B59B6']
            bars = ax1.bar(metodos, totales, color=colors[:len(metodos)], alpha=0.8)
            ax1.set_title('Ventas por Monto', fontsize=11, fontweight='bold', pad=10)
            ax1.set_ylabel('Total ($)', fontsize=9)
            ax1.tick_params(axis='x', rotation=15, labelsize=8)
            ax1.tick_params(axis='y', labelsize=8)
            ax1.grid(axis='y', alpha=0.3)
            
            # Agregar valores en las barras
            for bar in bars:
                height = bar.get_height()
                ax1.text(bar.get_x() + bar.get_width()/2., height,
                        f'${height:,.0f}',
                        ha='center', va='bottom', fontsize=8, fontweight='bold')
            
            # Gráfico circular - Cantidad de transacciones
            ax2 = fig.add_subplot(122)
            wedges, texts, autotexts = ax2.pie(cantidades, labels=metodos, autopct='%1.1f%%',
                                               colors=colors[:len(metodos)], startangle=90,
                                               textprops={'fontsize': 8})
            ax2.set_title('Cantidad de Transacciones', fontsize=11, fontweight='bold', pad=10)
            
            # Mejorar legibilidad de porcentajes
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
                autotext.set_fontsize(9)
            
            fig.tight_layout()
            fig.patch.set_facecolor('white')
            ax1.set_facecolor('white')
            
            # Integrar con tkinter
            canvas = FigureCanvasTkAgg(fig, chart_card)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True, padx=20, pady=(0, 20))
            
        except Exception as e:
            logger.error(f"Error al crear gráfico de métodos de pago: {e}")
            msg = ttk.Label(chart_card,
                          text=f"Error al crear el gráfico: {str(e)}",
                          style='Info.TLabel')
            msg.pack(pady=30)
    
    def show_pos_screen(self):
        """Muestra la pantalla de punto de venta ELITE"""
        self.clear_content()
        self.current_screen = "pos"
        self.current_cart = []
        
        # Frame principal con fondo oscuro
        main_container = tk.Frame(self.content_frame, bg=self.colors['bg_primary'])
        main_container.pack(fill='both', expand=True)
        
        # Header del POS - MÁS COMPACTO
        header = tk.Frame(main_container, bg=self.colors['bg_secondary'],
                         highlightbackground=self.colors['border'],
                         highlightthickness=1)
        header.pack(fill='x', padx=20, pady=(15, 10))
        
        header_content = tk.Frame(header, bg=self.colors['bg_secondary'])
        header_content.pack(fill='x', padx=20, pady=12)
        
        # Título con icono - MÁS PEQUEÑO
        title_frame = tk.Frame(header_content, bg=self.colors['bg_secondary'])
        title_frame.pack(side='left')
        
        tk.Label(title_frame,
                text="🛒",
                font=('Segoe UI', 18),
                bg=self.colors['bg_secondary'],
                fg=self.colors['success']).pack(side='left', padx=(0, 10))
        
        tk.Label(title_frame,
                text="Punto de Venta",
                font=('Segoe UI', 16, 'bold'),
                bg=self.colors['bg_secondary'],
                fg=self.colors['text_primary']).pack(side='left')
        
        # Contenedor principal con 2 columnas
        content_container = tk.Frame(main_container, bg=self.colors['bg_primary'])
        content_container.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        content_container.grid_columnconfigure(0, weight=2)
        content_container.grid_columnconfigure(1, weight=1)
        content_container.grid_rowconfigure(0, weight=1)
        
        # === PANEL IZQUIERDO - PRODUCTOS ===
        left_panel = tk.Frame(content_container,
                             bg=self.colors['bg_card'],
                             highlightbackground=self.colors['border'],
                             highlightthickness=1)
        left_panel.grid(row=0, column=0, sticky='nsew', padx=(0, 15))
        
        # Header del panel
        left_header = tk.Frame(left_panel, bg=self.colors['bg_secondary'])
        left_header.pack(fill='x')
        
        tk.Label(left_header,
                text="📦  PRODUCTOS DISPONIBLES",
                font=('Segoe UI', 12, 'bold'),
                bg=self.colors['bg_secondary'],
                fg=self.colors['text_primary']).pack(side='left', padx=20, pady=15)
        
        # Barra de búsqueda moderna
        search_container = tk.Frame(left_panel, bg=self.colors['bg_card'])
        search_container.pack(fill='x', padx=20, pady=15)
        
        search_frame = tk.Frame(search_container,
                               bg=self.colors['bg_secondary'],
                               highlightbackground=self.colors['accent_blue'],
                               highlightthickness=2)
        search_frame.pack(fill='x')
        
        tk.Label(search_frame,
                text="🔍",
                font=('Segoe UI', 14),
                bg=self.colors['bg_secondary'],
                fg=self.colors['accent_blue']).pack(side='left', padx=(15, 5))
        
        search_entry = tk.Entry(search_frame,
                               font=('Segoe UI', 12),
                               bg=self.colors['bg_secondary'],
                               fg=self.colors['text_primary'],
                               bd=0,
                               insertbackground=self.colors['accent_blue'],
                               relief='flat')
        search_entry.pack(side='left', fill='x', expand=True, padx=5, pady=12)
        
        # Checkbox código de barras
        barcode_frame = tk.Frame(search_container, bg=self.colors['bg_card'])
        barcode_frame.pack(fill='x', pady=(10, 0))
        
        barcode_search_var = tk.BooleanVar(value=False)
        barcode_check = tk.Checkbutton(barcode_frame,
                                      text="📟  Buscar por código de barras",
                                      variable=barcode_search_var,
                                      font=('Segoe UI', 10),
                                      bg=self.colors['bg_card'],
                                      fg=self.colors['text_secondary'],
                                      selectcolor=self.colors['bg_secondary'],
                                      activebackground=self.colors['bg_card'],
                                      activeforeground=self.colors['accent_blue'],
                                      bd=0,
                                      highlightthickness=0)
        barcode_check.pack(side='left')
        
        def search_products():
            search_term = search_entry.get().lower()
            search_by_barcode = barcode_search_var.get()
            self._update_products_tree(products_tree, search_term, search_by_barcode)
        
        search_entry.bind('<Return>', lambda e: search_products())
        search_entry.bind('<KeyRelease>', lambda e: search_products())
        
        # Treeview de productos con estilo elite
        tree_container = tk.Frame(left_panel, bg=self.colors['bg_card'])
        tree_container.pack(fill='both', expand=True, padx=20, pady=(10, 20))
        
        products_tree = ttk.Treeview(tree_container,
                                    columns=('Nombre', 'Precio', 'Stock'),
                                    show='headings',
                                    style='Modern.Treeview',
                                    selectmode='browse')
        
        products_tree.heading('Nombre', text='PRODUCTO')
        products_tree.heading('Precio', text='PRECIO')
        products_tree.heading('Stock', text='STOCK')
        
        products_tree.column('Nombre', width=250)
        products_tree.column('Precio', width=120)
        products_tree.column('Stock', width=100)
        
        products_tree.pack(side='left', fill='both', expand=True)
        
        # Scrollbar elite
        scrollbar = ttk.Scrollbar(tree_container, orient='vertical',
                                 command=products_tree.yview,
                                 style='Elite.Vertical.TScrollbar')
        scrollbar.pack(side='right', fill='y')
        products_tree.configure(yscrollcommand=scrollbar.set)
        
        # Cargar productos
        self._update_products_tree(products_tree)
        
        # Doble click para agregar
        def add_to_cart_double_click(event):
            self._add_to_cart_from_tree(products_tree, cart_tree, total_label)
            toast = ToastNotification(self.root, "Producto agregado al carrito", 'success', 2000)
            toast.show()
        
        products_tree.bind('<Double-1>', add_to_cart_double_click)
        
        # === PANEL DERECHO - CARRITO ===
        right_panel = tk.Frame(content_container,
                              bg=self.colors['bg_card'],
                              highlightbackground=self.colors['border'],
                              highlightthickness=1)
        right_panel.grid(row=0, column=1, sticky='nsew')
        
        # Header del carrito
        cart_header = tk.Frame(right_panel, bg=self.colors['success'])
        cart_header.pack(fill='x')
        
        tk.Label(cart_header,
                text="🛍️  CARRITO",
                font=('Segoe UI', 11, 'bold'),
                bg=self.colors['success'],
                fg=self.colors['primary_dark']).pack(side='left', padx=15, pady=10)
        
        cart_count = tk.Label(cart_header,
                             text="0 items",
                             font=('Segoe UI', 9),
                             bg=self.colors['success'],
                             fg=self.colors['primary_dark'])
        cart_count.pack(side='right', padx=15)
        
        # Treeview del carrito - SÚPER COMPACTO
        cart_container = tk.Frame(right_panel, bg=self.colors['bg_card'])
        cart_container.pack(fill='x', padx=12, pady=(10, 5))
        
        cart_tree = ttk.Treeview(cart_container,
                                columns=('Producto', 'Cant', 'Precio', 'Total'),
                                show='headings',
                                style='Modern.Treeview',
                                height=3)  # Reducido a 3 filas para ahorrar espacio
        
        cart_tree.heading('Producto', text='PRODUCTO')
        cart_tree.heading('Cant', text='CANT')
        cart_tree.heading('Precio', text='PRECIO')
        cart_tree.heading('Total', text='TOTAL')
        
        cart_tree.column('Producto', width=150)
        cart_tree.column('Cant', width=60)
        cart_tree.column('Precio', width=80)
        cart_tree.column('Total', width=90)
        
        cart_tree.pack(side='left', fill='both', expand=True)
        
        cart_scrollbar = ttk.Scrollbar(cart_container, orient='vertical',
                                       command=cart_tree.yview,
                                       style='Elite.Vertical.TScrollbar')
        cart_scrollbar.pack(side='right', fill='y')
        cart_tree.configure(yscrollcommand=cart_scrollbar.set)
        
        # Acciones del carrito - ESTILO NAVBAR (2 COLUMNAS)
        actions_frame = tk.Frame(right_panel, bg=self.colors['bg_card'])
        actions_frame.pack(fill='x', padx=12, pady=(5, 5))
        
        # Fila 1: Agregar y Quitar
        row1 = tk.Frame(actions_frame, bg=self.colors['bg_card'])
        row1.pack(fill='x', pady=2)
        
        add_btn = ModernButton(row1,
                              text="➕ Agregar",
                              style='success',
                              command=lambda: self._add_to_cart_from_tree(products_tree, cart_tree, total_label),
                              padx=10,
                              pady=6)
        add_btn.pack(side='left', fill='x', expand=True, padx=(0, 3))
        
        remove_btn = ModernButton(row1,
                                 text="➖ Quitar",
                                 style='warning',
                                 command=lambda: self._remove_from_cart(cart_tree, total_label, cart_count),
                                 padx=10,
                                 pady=6)
        remove_btn.pack(side='right', fill='x', expand=True, padx=(3, 0))
        
        # Fila 2: Vaciar
        row2 = tk.Frame(actions_frame, bg=self.colors['bg_card'])
        row2.pack(fill='x', pady=2)
        
        clear_cart_btn = ModernButton(row2,
                                     text="🗑️ Vaciar Carrito",
                                     style='danger',
                                     command=lambda: self._clear_cart(cart_tree, total_label, cart_count),
                                     padx=10,
                                     pady=6)
        clear_cart_btn.pack(fill='x')
        
        # Separador compacto
        separator = tk.Frame(right_panel, bg=self.colors['accent_blue'], height=2)
        separator.pack(fill='x', padx=12, pady=(5, 5))
        
        # Total destacado - COMPACTO
        total_container = tk.Frame(right_panel,
                                   bg=self.colors['bg_secondary'],
                                   highlightbackground=self.colors['success'],
                                   highlightthickness=2)
        total_container.pack(fill='x', padx=12, pady=(5, 8))
        
        total_inner = tk.Frame(total_container, bg=self.colors['bg_secondary'])
        total_inner.pack(fill='x', padx=10, pady=8)
        
        tk.Label(total_inner,
                text="TOTAL A PAGAR",
                font=('Segoe UI', 9, 'bold'),
                bg=self.colors['bg_secondary'],
                fg=self.colors['text_muted']).pack(pady=(0, 2))
        
        total_label = tk.Label(total_inner,
                              text="$0.00",
                              font=('Segoe UI', 22, 'bold'),
                              bg=self.colors['bg_secondary'],
                              fg=self.colors['success'])
        total_label.pack(pady=(2, 2))
        
        # Botón PROCESAR VENTA - DESTACADO Y VISIBLE
        process_btn = ModernButton(right_panel,
                                   text="💳 PROCESAR VENTA",
                                   command=lambda: self._process_sale(cart_tree),
                                   style='success',
                                   padx=20,
                                   pady=10)
        process_btn.pack(fill='x', padx=12, pady=(8, 12))
        
        # Hacer el botón más destacado
        process_btn.configure(
            font=('Segoe UI', 12, 'bold'),
            cursor='hand2'
        )
        
        # Animación de entrada
        AnimationHelper.fade_in(main_container, duration=400)
        
        # Guardar referencias
        self.products_tree = products_tree
        self.cart_tree = cart_tree
        self.total_label = total_label
    
    def _update_products_tree(self, tree, search_term='', search_by_barcode=False):
        """Actualiza el árbol de productos con búsqueda y color de stock"""
        for item in tree.get_children():
            tree.delete(item)

        # Configurar tags de color de stock
        tree.tag_configure('sin_stock',    foreground=self.colors['danger'])
        tree.tag_configure('stock_bajo',   foreground=self.colors['warning'])
        tree.tag_configure('stock_normal', foreground=self.colors['text_primary'])
        tree.tag_configure('even', background=self.colors.get('bg_secondary', '#1a2035'))
        tree.tag_configure('odd',  background=self.colors.get('bg_card',      '#1e2a3a'))

        productos = self.producto_service.listar_productos(solo_activos=True)

        idx = 0
        for producto in productos:
            match = (
                search_term == '' or
                (search_by_barcode and producto.codigo_barras and
                 search_term in producto.codigo_barras.lower()) or
                (not search_by_barcode and search_term in producto.nombre.lower())
            )
            if not match:
                continue

            stripe = 'even' if idx % 2 == 0 else 'odd'
            if producto.stock <= 0:
                stock_tag = 'sin_stock'
            elif producto.stock_minimo and producto.stock <= producto.stock_minimo:
                stock_tag = 'stock_bajo'
            else:
                stock_tag = 'stock_normal'

            tree.insert('', 'end',
                        values=(producto.nombre,
                                format_currency(producto.precio),
                                producto.stock),
                        tags=(producto.id, stripe, stock_tag))
            idx += 1
    
    def _add_to_cart_from_tree(self, products_tree, cart_tree, total_label, clear_btn=None):
        """Agrega un producto al carrito desde el árbol - VERSION ELITE"""
        selection = products_tree.selection()
        if not selection:
            ToastNotification.show(self.root, "Por favor seleccione un producto", 'warning', 2000)
            return
        
        # Obtener datos del producto
        item = products_tree.item(selection[0])
        # El ID del producto está en tags (primer valor numérico)
        producto_id = next(int(t) for t in item['tags'] if str(t).isdigit())
        nombre = item['values'][0]
        precio_str = item['values'][1].replace('$', '').replace(',', '')
        precio = float(precio_str)
        stock = int(item['values'][2])
        
        if stock <= 0:
            ToastNotification.show(self.root, f"Sin stock: {nombre}", 'error', 2500)
            AnimationHelper.shake(products_tree, duration=400)
            return
        
        # Verificar si ya está en el carrito
        for cart_item in self.current_cart:
            if cart_item['id'] == producto_id:
                if cart_item['cantidad'] >= stock:
                    ToastNotification.show(self.root, f"Stock máximo alcanzado: {nombre}", 'warning', 2500)
                    return
                cart_item['cantidad'] += 1
                cart_item['total'] = cart_item['cantidad'] * cart_item['precio']
                break
        else:
            # Agregar nuevo item
            self.current_cart.append({
                'id': producto_id,
                'nombre': nombre,
                'precio': precio,
                'cantidad': 1,
                'total': precio
            })
        
        # Actualizar vista del carrito con animación
        self._update_cart_tree_elite(cart_tree, total_label)
        AnimationHelper.pulse(cart_tree, duration=300)
    
    def _remove_from_cart(self, cart_tree, total_label, cart_count=None):
        """Quita un producto del carrito - VERSION ELITE"""
        selection = cart_tree.selection()
        if not selection:
            ToastNotification.show(self.root, "Seleccione un item del carrito", 'warning', 2000)
            return
        
        # Obtener índice
        item_index = cart_tree.index(selection[0])
        item_name = self.current_cart[item_index]['nombre']
        
        # Quitar del carrito
        del self.current_cart[item_index]
        
        # Actualizar vista
        self._update_cart_tree_elite(cart_tree, total_label, cart_count)
        ToastNotification.show(self.root, f"Eliminado: {item_name}", 'info', 2000)
    
    def _clear_cart(self, cart_tree, total_label, cart_count=None):
        """Vacía el carrito - VERSION ELITE"""
        if not self.current_cart:
            ToastNotification.show(self.root, "El carrito ya está vacío", 'info', 2000)
            return
        
        if messagebox.askyesno("Confirmar", "¿Desea vaciar el carrito?"):
            self.current_cart.clear()
            self._update_cart_tree_elite(cart_tree, total_label, cart_count)
            ToastNotification.show(self.root, "Carrito vaciado", 'success', 2000)
    
    def _update_cart_tree_elite(self, cart_tree, total_label, cart_count=None):
        """Actualiza la vista del carrito - VERSION ELITE"""
        # Limpiar
        for item in cart_tree.get_children():
            cart_tree.delete(item)
        
        # Agregar items
        total = 0
        for item in self.current_cart:
            cart_tree.insert('', 'end',
                           values=(item['nombre'],
                                 item['cantidad'],
                                 format_currency(item['precio']),
                                 format_currency(item['total'])))
            total += item['total']
        
        # Actualizar total con animación
        total_label.configure(text=format_currency(total))
        if total > 0:
            AnimationHelper.pulse(total_label, duration=400)
        
        # Actualizar contador si existe
        if cart_count:
            count = len(self.current_cart)
            cart_count.configure(text=f"{count} item{'s' if count != 1 else ''}")
    
    def _process_sale(self, cart_tree):
        """Procesa la venta"""
        if not self.current_cart:
            messagebox.showwarning("Carrito Vacío", "Agregue productos al carrito primero")
            return
        
        # ✅ VALIDACIÓN DE STOCK ANTES DE PROCESAR
        productos_sin_stock = []
        for item in self.current_cart:
            producto = self.producto_service.obtener_producto(item['id'])
            if not producto:
                messagebox.showerror("Error", f"Producto {item['nombre']} no encontrado")
                return
            
            if producto.stock < item['cantidad']:
                productos_sin_stock.append({
                    'nombre': item['nombre'],
                    'solicitado': item['cantidad'],
                    'disponible': producto.stock
                })
        
        # Si hay productos sin stock suficiente, mostrar mensaje detallado
        if productos_sin_stock:
            mensaje = "Los siguientes productos no tienen stock suficiente:\n\n"
            for p in productos_sin_stock:
                mensaje += f"• {p['nombre']}: Solicitado {p['solicitado']}, Disponible {p['disponible']}\n"
            mensaje += "\nPor favor ajuste las cantidades en el carrito."
            messagebox.showerror("Stock Insuficiente", mensaje)
            return
        
        # 💳 SELECCIONAR MÉTODO DE PAGO
        metodo_pago = self._select_payment_method()
        if not metodo_pago:
            return  # Usuario canceló
        
        # Crear venta
        venta = Venta()
        venta.usuario_id = self.auth_service.current_user.id
        venta.fecha = datetime.now()
        venta.metodo_pago = metodo_pago
        
        # Agregar detalles
        for item in self.current_cart:
            detalle = DetalleVenta()
            detalle.producto_id = item['id']
            detalle.producto_nombre = item['nombre']
            detalle.cantidad = item['cantidad']
            detalle.precio_unitario = item['precio']
            detalle.calcular_subtotal()
            venta.detalles.append(detalle)
        
        # Procesar venta
        success, message, venta_id = self.venta_service.crear_venta(
            venta,
            self.auth_service.current_user.id
        )
        
        if success:
            # Preguntar si desea imprimir ticket
            imprimir = messagebox.askyesno(
                "Venta Exitosa", 
                f"Venta procesada correctamente\nNúmero: {venta.numero_venta}\n\n¿Desea imprimir el ticket?"
            )
            
            if imprimir:
                self._print_ticket(venta_id, venta.numero_venta)
            
            # Limpiar carrito
            self.current_cart.clear()
            self._update_cart_tree_elite(cart_tree, self.total_label)
            # Actualizar productos
            self._update_products_tree(self.products_tree)
        else:
            messagebox.showerror("Error", message)
    
    def _select_payment_method(self):
        """Muestra diálogo para seleccionar método de pago"""
        payment_dialog = tk.Toplevel(self.root)
        payment_dialog.title("Método de Pago")
        payment_dialog.geometry("450x420")
        payment_dialog.transient(self.root)
        payment_dialog.grab_set()
        
        # Centrar diálogo
        payment_dialog.update_idletasks()
        x = (payment_dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (payment_dialog.winfo_screenheight() // 2) - (420 // 2)
        payment_dialog.geometry(f"450x420+{x}+{y}")
        
        main_frame = ttk.Frame(payment_dialog, padding=30)
        main_frame.pack(fill='both', expand=True)
        
        # Título
        ttk.Label(main_frame, text="Seleccione el Método de Pago",
                 font=('Segoe UI', 16, 'bold')).pack(pady=(0, 20))
        
        # Variable para almacenar la selección
        selected_method = tk.StringVar()
        
        # Frame para los botones de método de pago
        methods_frame = ttk.Frame(main_frame)
        methods_frame.pack(expand=True, fill='both')
        
        # Métodos de pago con iconos
        metodos = [
            ('💵 Efectivo', 'efectivo'),
            ('💳 Tarjeta de Crédito', 'tarjeta_credito'),
            ('💳 Tarjeta de Débito', 'tarjeta_debito'),
            ('🏦 Transferencia', 'transferencia')
        ]
        
        def select_and_close(metodo):
            selected_method.set(metodo)
            payment_dialog.destroy()
        
        # Crear botones grandes para cada método
        for i, (texto, valor) in enumerate(metodos):
            btn = ttk.Button(methods_frame,
                           text=texto,
                           command=lambda v=valor: select_and_close(v),
                           style='Success.TButton')
            btn.pack(pady=8, fill='x', ipady=10)
        
        # Botón cancelar
        ttk.Button(main_frame,
                  text="Cancelar",
                  command=payment_dialog.destroy,
                  style='Danger.TButton').pack(pady=(20, 0))
        
        # Esperar a que se cierre el diálogo
        payment_dialog.wait_window()
        
        return selected_method.get() if selected_method.get() else None
    
    def _print_ticket(self, venta_id, numero_venta):
        """Imprime el ticket de una venta recién procesada"""
        try:
            # Obtener información de la venta
            venta = self.db.fetch_one(
                """SELECT v.*, u.nombre as usuario_nombre
                   FROM ventas v
                   LEFT JOIN usuarios u ON v.usuario_id = u.id
                   WHERE v.id = ?""",
                (venta_id,)
            )
            
            if not venta:
                messagebox.showerror("Error", "Venta no encontrada")
                return
            
            # Obtener detalles con nombre del producto
            detalles = self.db.fetch_all(
                """SELECT dv.*, p.nombre as producto_nombre
                   FROM detalles_venta dv
                   LEFT JOIN productos p ON dv.producto_id = p.id
                   WHERE dv.venta_id = ?""",
                (venta_id,)
            )
            
            # Generar contenido del ticket
            ticket_content = self._generate_ticket_content(venta, detalles)
            
            # Mostrar opciones de impresión
            print_dialog = tk.Toplevel(self.root)
            print_dialog.title("Imprimir Ticket")
            print_dialog.geometry("700x800")
            print_dialog.transient(self.root)
            print_dialog.grab_set()
            
            main_frame = ttk.Frame(print_dialog, padding=20)
            main_frame.pack(fill='both', expand=True)
            
            ttk.Label(main_frame, text="Vista Previa del Ticket",
                     font=('Segoe UI', 14, 'bold')).pack(pady=(0, 10))
            
            # Texto del ticket
            text_frame = ttk.Frame(main_frame)
            text_frame.pack(fill='both', expand=True)
            
            ticket_text = tk.Text(text_frame, font=('Courier New', 9),
                                 wrap='none', bg='white', fg='black')
            ticket_text.pack(side='left', fill='both', expand=True)
            
            scrollbar = ttk.Scrollbar(text_frame, command=ticket_text.yview)
            scrollbar.pack(side='right', fill='y')
            ticket_text.config(yscrollcommand=scrollbar.set)
            
            ticket_text.insert('1.0', ticket_content)
            ticket_text.config(state='disabled')
            
            # Botones
            btn_frame = ttk.Frame(main_frame)
            btn_frame.pack(pady=15)
            
            def save_ticket():
                filename = filedialog.asksaveasfilename(
                    defaultextension=".txt",
                    filetypes=[("Archivo de texto", "*.txt"), ("Todos los archivos", "*.*")],
                    initialfile=f"ticket_{numero_venta}.txt"
                )
                if filename:
                    try:
                        with open(filename, 'w', encoding='utf-8') as f:
                            f.write(ticket_content)
                        messagebox.showinfo("Éxito", "Ticket guardado correctamente")
                    except Exception as e:
                        messagebox.showerror("Error", f"No se pudo guardar: {str(e)}")
            
            def print_to_printer():
                # Intentar imprimir en Windows usando notepad
                try:
                    import tempfile
                    import subprocess
                    
                    # Crear archivo temporal
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt',
                                                    delete=False, encoding='utf-8') as f:
                        f.write(ticket_content)
                        temp_file = f.name
                    
                    # Imprimir usando notepad (solo Windows)
                    subprocess.run(['notepad', '/p', temp_file], check=True)
                    
                    # Limpiar archivo temporal después de un delay
                    import os
                    import time
                    time.sleep(2)
                    try:
                        os.unlink(temp_file)
                    except:
                        pass
                    
                    messagebox.showinfo("Información", "Ticket enviado a la impresora predeterminada")
                    print_dialog.destroy()
                    
                except Exception as e:
                    logger.error(f"Error al imprimir: {e}")
                    messagebox.showerror("Error", 
                                       f"No se pudo imprimir directamente.\n" +
                                       "Use 'Guardar como...' y luego imprima el archivo manualmente.")
            
            ttk.Button(btn_frame, text="🖨️ Imprimir", command=print_to_printer,
                      style='Success.TButton').pack(side='left', padx=5)
            ttk.Button(btn_frame, text="💾 Guardar como...", command=save_ticket,
                      style='Primary.TButton').pack(side='left', padx=5)
            ttk.Button(btn_frame, text="Cerrar", command=print_dialog.destroy,
                      style='Danger.TButton').pack(side='left', padx=5)
            
        except Exception as e:
            logger.error(f"Error al preparar impresión: {e}")
            messagebox.showerror("Error", f"No se pudo preparar la impresión: {str(e)}")
    
    def show_sales_history(self):
        """Muestra el historial de ventas — diseño moderno"""
        self.clear_content()
        self.current_screen = "sales"

        self.sales_current_page = 1
        self.sales_items_per_page = 50
        self.sales_filters = {
            'from_date': datetime.now().strftime('%Y-%m-%d'),
            'to_date':   datetime.now().strftime('%Y-%m-%d'),
        }

        # Wrapper scrollable
        outer = tk.Frame(self.content_frame, bg=self.colors['bg_primary'])
        outer.pack(fill='both', expand=True)

        # ── Encabezado ──────────────────────────────────────────────────────
        self._ui_page_header(outer, '📋', 'Historial de Ventas',
                             'Consultá y gestioná todas las ventas registradas')

        # ── Barra de filtros ─────────────────────────────────────────────────
        filter_card = self._ui_section_card(outer)
        filter_card.pack(fill='x', padx=24, pady=(14, 0))

        filter_inner = tk.Frame(filter_card, bg=self.colors['bg_card'])
        filter_inner.pack(fill='x', padx=16, pady=12)

        def _lbl(text):
            return tk.Label(filter_inner, text=text, font=('Segoe UI', 9, 'bold'),
                            bg=self.colors['bg_card'], fg=self.colors['text_muted'])
        def _entry(default, w=14):
            e = tk.Entry(filter_inner, width=w, font=('Segoe UI', 10),
                         bg=self.colors['bg_secondary'],
                         fg=self.colors['text_primary'],
                         insertbackground=self.colors['accent_blue'],
                         bd=0, relief='flat')
            e.insert(0, default)
            return e

        _lbl('Desde:').pack(side='left', padx=(0, 4))
        from_date = _entry(self.sales_filters['from_date'])
        from_date.pack(side='left', ipady=6, ipadx=6, padx=(0, 16))

        _lbl('Hasta:').pack(side='left', padx=(0, 4))
        to_date = _entry(self.sales_filters['to_date'])
        to_date.pack(side='left', ipady=6, ipadx=6, padx=(0, 16))

        _lbl('Por página:').pack(side='left', padx=(0, 4))
        items_per_page_var = tk.StringVar(value='50')
        combo = ttk.Combobox(filter_inner, textvariable=items_per_page_var,
                             width=7, state='readonly',
                             font=('Segoe UI', 10))
        combo['values'] = ('10', '25', '50', '100', '200')
        combo.pack(side='left', padx=(0, 16))

        def apply_filters():
            self.sales_filters['from_date'] = from_date.get()
            self.sales_filters['to_date']   = to_date.get()
            self.sales_items_per_page = int(items_per_page_var.get())
            self.sales_current_page   = 1
            self._load_sales_paginated(sales_tree, pagination_label, nav_buttons)

        self._ui_action_btn(filter_inner, '🔍  Buscar', apply_filters)

        # ── Tabla de ventas ──────────────────────────────────────────────────
        table_card = self._ui_section_card(outer)
        table_card.pack(fill='both', expand=True, padx=24, pady=14)

        sales_tree = self._ui_styled_tree(table_card,
            columns=('ID', 'Número', 'Fecha', 'Usuario', 'Total', 'Método', 'Estado'),
            col_configs={
                'ID':      {'text': 'ID',          'width': 50,  'anchor': 'center'},
                'Número':  {'text': 'N° Venta',    'width': 130, 'anchor': 'center'},
                'Fecha':   {'text': 'Fecha',        'width': 155, 'anchor': 'center'},
                'Usuario': {'text': 'Usuario',      'width': 120},
                'Total':   {'text': 'Total',        'width': 110, 'anchor': 'e'},
                'Método':  {'text': 'Método Pago',  'width': 120},
                'Estado':  {'text': 'Estado',       'width': 100, 'anchor': 'center'},
            })

        # Tags de estado
        sales_tree.tag_configure('completada', foreground=self.colors['success'])
        sales_tree.tag_configure('anulada',    foreground=self.colors['danger'])
        sales_tree.tag_configure('pendiente',  foreground=self.colors['warning'])

        # ── Paginación ───────────────────────────────────────────────────────
        _, pagination_label, nav_buttons = self._ui_pagination_bar(
            table_card,
            on_prev=lambda: self._sales_prev_page(sales_tree, pagination_label, nav_buttons),
            on_next=lambda: self._sales_next_page(sales_tree, pagination_label, nav_buttons),
        )

        # ── Barra de acciones ────────────────────────────────────────────────
        action_card = self._ui_section_card(outer)
        action_card.pack(fill='x', padx=24, pady=(0, 20))

        action_inner = tk.Frame(action_card, bg=self.colors['bg_card'])
        action_inner.pack(fill='x', padx=16, pady=10)

        self._ui_action_btn(action_inner, '📄  Ver Detalles',
                            lambda: self._show_sale_details(sales_tree))
        self._ui_action_btn(action_inner, '🖨️  Imprimir',
                            lambda: self._print_sale(sales_tree), style='success')
        self._ui_action_btn(action_inner, '❌  Anular Venta',
                            lambda: self._cancel_sale(sales_tree), style='danger')

        # Cargar datos
        self._load_sales_paginated(sales_tree, pagination_label, nav_buttons)

    def _sales_prev_page(self, tree, label, buttons):
        """Ir a la página anterior"""
        if self.sales_current_page > 1:
            self.sales_current_page -= 1
            self._load_sales_paginated(tree, label, buttons)

    def _sales_next_page(self, tree, label, buttons):
        """Ir a la página siguiente"""
        total_items = self._get_total_sales_count()
        total_pages = (total_items + self.sales_items_per_page - 1) // self.sales_items_per_page
        if self.sales_current_page < total_pages:
            self.sales_current_page += 1
            self._load_sales_paginated(tree, label, buttons)

    def _get_total_sales_count(self):
        """Obtiene el total de ventas según los filtros"""
        try:
            result = self.db.fetch_one(
                """SELECT COUNT(*) as total FROM ventas
                   WHERE DATE(fecha) BETWEEN ? AND ?""",
                (self.sales_filters['from_date'], self.sales_filters['to_date'])
            )
            return result['total'] if result else 0
        except Exception as e:
            logger.error(f"Error al contar ventas: {e}")
            return 0
    
    def _load_sales_paginated(self, tree, label, buttons):
        """Carga las ventas con paginación"""
        for item in tree.get_children():
            tree.delete(item)

        try:
            offset = (self.sales_current_page - 1) * self.sales_items_per_page

            rows = self.db.fetch_all(
                """SELECT v.*, u.nombre as usuario_nombre
                   FROM ventas v
                   LEFT JOIN usuarios u ON v.usuario_id = u.id
                   WHERE DATE(v.fecha) BETWEEN ? AND ?
                   ORDER BY v.fecha DESC
                   LIMIT ? OFFSET ?""",
                (self.sales_filters['from_date'], self.sales_filters['to_date'],
                 self.sales_items_per_page, offset)
            )

            for i, row in enumerate(rows):
                fecha  = datetime.fromisoformat(row['fecha']).strftime('%Y-%m-%d %H:%M')
                estado = row['estado']
                # Tags: stripe + estado
                stripe = 'even' if i % 2 == 0 else 'odd'
                tags   = (stripe, estado)
                tree.insert('', 'end',
                            values=(row['id'],
                                    row['numero_venta'],
                                    fecha,
                                    row['usuario_nombre'] or 'N/A',
                                    format_currency(row['total']),
                                    row['metodo_pago'].title(),
                                    estado.title()),
                            tags=tags)

            total_items = self._get_total_sales_count()
            total_pages = max(1, (total_items + self.sales_items_per_page - 1) // self.sales_items_per_page)
            label.config(text=f"Página {self.sales_current_page} de {total_pages}  ({total_items} registros)")

            buttons['prev'].config(state='normal' if self.sales_current_page > 1 else 'disabled')
            buttons['next'].config(state='normal' if self.sales_current_page < total_pages else 'disabled')

        except Exception as e:
            logger.error(f"Error al cargar ventas paginadas: {e}")
            messagebox.showerror("Error", f"Error al cargar ventas: {str(e)}")
            label.config(text="Error al cargar datos")
            buttons['prev'].config(state='disabled')
            buttons['next'].config(state='disabled')
    
    def _load_sales(self, tree, from_date, to_date):
        """Carga las ventas en el árbol"""
        # Limpiar árbol
        for item in tree.get_children():
            tree.delete(item)
        
        try:
            # Consultar ventas
            rows = self.db.fetch_all(
                """SELECT v.*, u.nombre as usuario_nombre 
                   FROM ventas v 
                   LEFT JOIN usuarios u ON v.usuario_id = u.id
                   WHERE DATE(v.fecha) BETWEEN ? AND ?
                   ORDER BY v.fecha DESC""",
                (from_date, to_date)
            )
            
            for row in rows:
                fecha = datetime.fromisoformat(row['fecha']).strftime('%Y-%m-%d %H:%M:%S')
                tree.insert('', 'end',
                           values=(row['id'],
                                  row['numero_venta'],
                                  fecha,
                                  row['usuario_nombre'] or 'N/A',
                                  format_currency(row['total']),
                                  row['metodo_pago'].title(),
                                  row['estado'].title()))
            
        except Exception as e:
            logger.error(f"Error al cargar ventas: {e}")
            messagebox.showerror("Error", f"Error al cargar ventas: {str(e)}")
    
    def _show_sale_details(self, tree):
        """Muestra los detalles de una venta"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione una venta")
            return
        
        venta_id = tree.item(selection[0])['values'][0]
        
        # Obtener detalles
        try:
            # Modificar consulta para incluir nombre del producto
            detalles = self.db.fetch_all(
                """SELECT dv.*, p.nombre as producto_nombre 
                   FROM detalles_venta dv
                   LEFT JOIN productos p ON dv.producto_id = p.id
                   WHERE dv.venta_id = ?""",
                (venta_id,)
            )
            
            # Crear ventana de detalles
            details_window = tk.Toplevel(self.root)
            details_window.title(f"Detalles de Venta #{venta_id}")
            details_window.geometry("600x400")
            details_window.transient(self.root)
            
            # Título
            ttk.Label(details_window,
                     text=f"Detalles de Venta #{venta_id}",
                     font=('Segoe UI', 16, 'bold')).pack(pady=20)
            
            # Treeview de detalles
            details_tree = ttk.Treeview(details_window,
                                       columns=('Producto', 'Cantidad', 'Precio', 'Subtotal'),
                                       show='headings',
                                       style='Modern.Treeview')
            
            details_tree.heading('Producto', text='Producto')
            details_tree.heading('Cantidad', text='Cantidad')
            details_tree.heading('Precio', text='Precio Unit.')
            details_tree.heading('Subtotal', text='Subtotal')
            
            details_tree.pack(fill='both', expand=True, padx=20, pady=10)
            
            total = 0
            for detalle in detalles:
                # Manejar producto_nombre que puede ser None si el producto fue eliminado
                producto_nombre = detalle['producto_nombre'] if detalle['producto_nombre'] else 'Producto Desconocido'
                details_tree.insert('', 'end',
                                   values=(producto_nombre,
                                          detalle['cantidad'],
                                          format_currency(detalle['precio_unitario']),
                                          format_currency(detalle['subtotal'])))
                total += detalle['subtotal']
            
            # Total
            ttk.Label(details_window,
                     text=f"TOTAL: {format_currency(total)}",
                     font=('Segoe UI', 14, 'bold')).pack(pady=10)
            
            # Botón cerrar
            ttk.Button(details_window,
                      text="Cerrar",
                      command=details_window.destroy,
                      style='Primary.TButton').pack(pady=10)
            
        except Exception as e:
            logger.error(f"Error al mostrar detalles: {e}")
            messagebox.showerror("Error", f"Error al mostrar detalles: {str(e)}")
    
    def _print_sale(self, tree):
        """Imprime un ticket de venta"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione una venta")
            return
        
        try:
            venta_id = tree.item(selection[0])['values'][0]
            
            # Obtener información de la venta
            venta = self.db.fetch_one(
                """SELECT v.*, u.nombre as usuario_nombre
                   FROM ventas v
                   LEFT JOIN usuarios u ON v.usuario_id = u.id
                   WHERE v.id = ?""",
                (venta_id,)
            )
            
            if not venta:
                messagebox.showerror("Error", "Venta no encontrada")
                return
            
            # Obtener detalles con nombre del producto
            detalles = self.db.fetch_all(
                """SELECT dv.*, p.nombre as producto_nombre
                   FROM detalles_venta dv
                   LEFT JOIN productos p ON dv.producto_id = p.id
                   WHERE dv.venta_id = ?""",
                (venta_id,)
            )
            
            # Generar contenido del ticket
            ticket_content = self._generate_ticket_content(venta, detalles)
            
            # Mostrar opciones de impresión
            print_dialog = tk.Toplevel(self.root)
            print_dialog.title("Imprimir Ticket")
            print_dialog.geometry("700x800")
            print_dialog.transient(self.root)
            
            main_frame = ttk.Frame(print_dialog, padding=20)
            main_frame.pack(fill='both', expand=True)
            
            ttk.Label(main_frame, text="Vista Previa del Ticket",
                     font=('Segoe UI', 14, 'bold')).pack(pady=(0, 10))
            
            # Texto del ticket
            text_frame = ttk.Frame(main_frame)
            text_frame.pack(fill='both', expand=True)
            
            ticket_text = tk.Text(text_frame, font=('Courier New', 9),
                                 wrap='none', bg='white', fg='black')
            ticket_text.pack(side='left', fill='both', expand=True)
            
            scrollbar = ttk.Scrollbar(text_frame, command=ticket_text.yview)
            scrollbar.pack(side='right', fill='y')
            ticket_text.config(yscrollcommand=scrollbar.set)
            
            ticket_text.insert('1.0', ticket_content)
            ticket_text.config(state='disabled')
            
            # Botones
            btn_frame = ttk.Frame(main_frame)
            btn_frame.pack(pady=15)
            
            def save_ticket():
                filename = filedialog.asksaveasfilename(
                    defaultextension=".txt",
                    filetypes=[("Archivo de texto", "*.txt"), ("Todos los archivos", "*.*")],
                    initialfile=f"ticket_{venta['numero_venta']}.txt"
                )
                if filename:
                    try:
                        with open(filename, 'w', encoding='utf-8') as f:
                            f.write(ticket_content)
                        messagebox.showinfo("Éxito", "Ticket guardado correctamente")
                    except Exception as e:
                        messagebox.showerror("Error", f"No se pudo guardar: {str(e)}")
            
            def print_to_printer():
                # Intentar imprimir en Windows usando notepad
                try:
                    import tempfile
                    import subprocess
                    
                    # Crear archivo temporal
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt',
                                                    delete=False, encoding='utf-8') as f:
                        f.write(ticket_content)
                        temp_file = f.name
                    
                    # Imprimir usando notepad (solo Windows)
                    subprocess.run(['notepad', '/p', temp_file], check=True)
                    
                    # Limpiar archivo temporal después de un delay
                    import os
                    import time
                    time.sleep(2)
                    try:
                        os.unlink(temp_file)
                    except:
                        pass
                    
                    messagebox.showinfo("Información", "Ticket enviado a la impresora predeterminada")
                    
                except Exception as e:
                    logger.error(f"Error al imprimir: {e}")
                    messagebox.showerror("Error", 
                                       f"No se pudo imprimir directamente.\n" +
                                       "Use 'Guardar como...' y luego imprima el archivo manualmente.")
            
            ttk.Button(btn_frame, text="🖨️ Imprimir", command=print_to_printer,
                      style='Success.TButton').pack(side='left', padx=5)
            ttk.Button(btn_frame, text="💾 Guardar como...", command=save_ticket,
                      style='Primary.TButton').pack(side='left', padx=5)
            ttk.Button(btn_frame, text="Cerrar", command=print_dialog.destroy,
                      style='Danger.TButton').pack(side='left', padx=5)
            
        except Exception as e:
            logger.error(f"Error al preparar impresión: {e}")
            messagebox.showerror("Error", f"No se pudo preparar la impresión: {str(e)}")
    
    def _generate_ticket_content(self, venta, detalles):
        """Genera el contenido del ticket de venta"""
        from datetime import datetime
        
        # Encabezado
        ticket = "=" * 48 + "\n"
        ticket += "           SISTEMA VENTAS PROFECIONAL\n"
        ticket += "         Sistema TPV - Ticket de Venta\n"
        ticket += "=" * 48 + "\n\n"
        
        # Información de la venta
        fecha = datetime.fromisoformat(venta['fecha']).strftime('%d/%m/%Y %H:%M:%S')
        
        # Diccionario de nombres amigables para métodos de pago
        metodo_nombres = {
            'efectivo': 'Efectivo',
            'tarjeta_credito': 'Tarjeta de Crédito',
            'tarjeta_debito': 'Tarjeta de Débito',
            'transferencia': 'Transferencia',
            'otro': 'Otro'
        }
        metodo_pago = metodo_nombres.get(venta['metodo_pago'], venta['metodo_pago'].title())
        
        ticket += f"Ticket No:    {venta['numero_venta']}\n"
        ticket += f"Fecha:        {fecha}\n"
        ticket += f"Cajero:       {venta['usuario_nombre']}\n"
        ticket += f"Método Pago:  {metodo_pago}\n"
        
        # Verificar si existe cliente_nombre y no es None
        try:
            if venta['cliente_nombre']:
                ticket += f"Cliente:      {venta['cliente_nombre']}\n"
        except (KeyError, TypeError):
            pass
        
        ticket += "\n" + "-" * 48 + "\n"
        ticket += "PRODUCTO                   CANT   PRECIO   TOTAL\n"
        ticket += "-" * 48 + "\n"
        
        # Detalles de productos
        subtotal_total = 0
        for detalle in detalles:
            # Manejar producto_nombre que puede ser None
            nombre_producto = detalle['producto_nombre'] if detalle['producto_nombre'] else 'Producto'
            nombre = nombre_producto[:23]  # Máximo 23 caracteres
            cantidad = detalle['cantidad']
            precio = detalle['precio_unitario']
            subtotal = detalle['subtotal']
            subtotal_total += subtotal
            
            ticket += f"{nombre:<24} {cantidad:>4} {precio:>8.2f} {subtotal:>8.2f}\n"
        
        # Totales
        ticket += "-" * 48 + "\n"
        ticket += f"{'Subtotal:':<36} {format_currency(venta['subtotal']):>11}\n"
        
        # Verificar descuento
        try:
            if venta['descuento'] and venta['descuento'] > 0:
                ticket += f"{'Descuento:':<36} -{format_currency(venta['descuento']):>11}\n"
        except (KeyError, TypeError):
            pass
        
        # Verificar impuestos
        try:
            if venta['impuestos'] and venta['impuestos'] > 0:
                ticket += f"{'Impuestos:':<36} {format_currency(venta['impuestos']):>11}\n"
        except (KeyError, TypeError):
            pass
        
        ticket += "=" * 48 + "\n"
        ticket += f"{'TOTAL:':<36} {format_currency(venta['total']):>11}\n"
        ticket += "=" * 48 + "\n\n"
        
        # Pie de página
        ticket += "        ¡Gracias por su compra!\n"
        ticket += "         Vuelva pronto :)\n\n"
        ticket += f"   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
        ticket += "=" * 48 + "\n"
        
        return ticket
    
    def _cancel_sale(self, tree):
        """Anula una venta y restaura el stock de los productos"""
        # Validar que hay una selección
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Por favor, seleccione una venta para anular")
            return
        
        # Obtener el ID de la venta
        venta_id = tree.item(selection[0])['values'][0]
        
        try:
            # Obtener información de la venta
            venta = self.db.fetch_one("""
                SELECT v.*, u.username as usuario_nombre, u.nombre as usuario_nombre_completo
                FROM ventas v
                LEFT JOIN usuarios u ON v.usuario_id = u.id
                WHERE v.id = ?
            """, (venta_id,))
            
            if not venta:
                messagebox.showerror("Error", "No se encontró la venta seleccionada")
                return
            
            # Verificar que la venta no esté ya anulada
            if venta['estado'] == 'cancelada':
                messagebox.showwarning("Advertencia", "Esta venta ya fue anulada previamente")
                return
            
            # Confirmar la anulación
            from datetime import datetime
            fecha = datetime.fromisoformat(venta['fecha']).strftime('%d/%m/%Y %H:%M')
            
            confirmar = messagebox.askyesno(
                "Confirmar Anulación",
                f"¿Está seguro de anular la siguiente venta?\n\n"
                f"Ticket No: {venta['numero_venta']}\n"
                f"Fecha: {fecha}\n"
                f"Total: {format_currency(venta['total'])}\n"
                f"Método: {venta['metodo_pago'].title()}\n\n"
                "El stock de los productos será restaurado automáticamente."
            )
            
            if not confirmar:
                return
            
            # Obtener los detalles de la venta para restaurar el stock
            detalles = self.db.fetch_all("""
                SELECT dv.*, p.nombre as producto_nombre
                FROM detalles_venta dv
                LEFT JOIN productos p ON dv.producto_id = p.id
                WHERE dv.venta_id = ?
            """, (venta_id,))
            
            # Preparar transacción
            operations = []
            
            # 1. Actualizar estado de la venta
            operations.append((
                "UPDATE ventas SET estado = 'cancelada' WHERE id = ?",
                (venta_id,)
            ))
            
            # 2. Restaurar stock de cada producto
            for detalle in detalles:
                # Incrementar stock
                operations.append((
                    "UPDATE productos SET stock = stock + ? WHERE id = ?",
                    (detalle['cantidad'], detalle['producto_id'])
                ))
                
                # Registrar movimiento de inventario
                operations.append((
                    """INSERT INTO movimientos_inventario 
                       (producto_id, tipo, cantidad, usuario_id, referencia, fecha)
                       VALUES (?, 'devolucion', ?, ?, ?, datetime('now'))""",
                    (detalle['producto_id'], 
                     detalle['cantidad'],
                     self.auth_service.current_user.id,
                     f"Anulación venta {venta['numero_venta']}")
                ))
            
            # Ejecutar transacción
            if self.db.execute_transaction(operations):
                messagebox.showinfo(
                    "Éxito",
                    f"La venta {venta['numero_venta']} ha sido anulada correctamente.\n\n"
                    f"Stock restaurado para {len(detalles)} producto(s)."
                )
                
                # Log de la operación
                logger.info(
                    f"Venta {venta['numero_venta']} (ID: {venta_id}) anulada por "
                    f"{self.auth_service.current_user.username}. "
                    f"Total: {venta['total']}"
                )
                
                # Recargar la pantalla de ventas completa
                self.show_sales_history()
                    
            else:
                messagebox.showerror(
                    "Error",
                    "No se pudo anular la venta. Por favor, intente nuevamente."
                )
                logger.error(f"Error al anular venta ID: {venta_id}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al anular la venta: {str(e)}")
            logger.error(f"Error al anular venta ID {venta_id}: {str(e)}")
    
    def show_products_manager(self):
        """Muestra el gestor de productos — diseño moderno"""
        self.clear_content()
        self.current_screen = "products"

        self.products_current_page  = 1
        self.products_items_per_page = 50
        self.products_search_term   = ''

        outer = tk.Frame(self.content_frame, bg=self.colors['bg_primary'])
        outer.pack(fill='both', expand=True)

        # ── Encabezado ──────────────────────────────────────────────────────
        self._ui_page_header(outer, '📦', 'Gestión de Productos',
                             'Administrá el catálogo, precios y stock')

        # ── Barra de herramientas ────────────────────────────────────────────
        toolbar_card = self._ui_section_card(outer)
        toolbar_card.pack(fill='x', padx=24, pady=(14, 0))

        toolbar = tk.Frame(toolbar_card, bg=self.colors['bg_card'])
        toolbar.pack(fill='x', padx=16, pady=10)

        # Búsqueda
        search_wrap = tk.Frame(toolbar, bg=self.colors['bg_card'])
        search_wrap.pack(side='left', fill='x', expand=True, padx=(0, 16))
        search_entry_raw = self._ui_search_bar(
            search_wrap, placeholder='Buscar por nombre o categoría...',
            on_enter=lambda _: do_search())

        # Por página
        tk.Label(toolbar, text='Por página:', font=('Segoe UI', 9, 'bold'),
                 bg=self.colors['bg_card'],
                 fg=self.colors['text_muted']).pack(side='left', padx=(0, 4))
        items_per_page_var = tk.StringVar(value='50')
        combo = ttk.Combobox(toolbar, textvariable=items_per_page_var,
                             width=6, state='readonly', font=('Segoe UI', 10))
        combo['values'] = ('25', '50', '100', '200')
        combo.pack(side='left', padx=(0, 16))

        def do_search():
            raw = search_entry_raw.get()
            self.products_search_term   = '' if raw == 'Buscar por nombre o categoría...' else raw
            self.products_items_per_page = int(items_per_page_var.get())
            self.products_current_page  = 1
            self._load_products_paginated(products_tree, pagination_label, nav_buttons)

        self._ui_action_btn(toolbar, '🔍  Buscar', do_search)
        self._ui_action_btn(toolbar, '➕  Nuevo Producto',
                            lambda: self._new_product_dialog(products_tree),
                            style='success', side='right', padx=(0, 0))

        # ── Tabla de productos ───────────────────────────────────────────────
        table_card = self._ui_section_card(outer)
        table_card.pack(fill='both', expand=True, padx=24, pady=14)

        cols = ('ID', 'Nombre', 'Categoría', 'Precio', 'Costo', 'Stock', 'Mín', 'Estado')
        products_tree = self._ui_styled_tree(table_card,
            columns=cols,
            col_configs={
                'ID':        {'text': 'ID',        'width': 50,  'anchor': 'center'},
                'Nombre':    {'text': 'Nombre',     'width': 200},
                'Categoría': {'text': 'Categoría',  'width': 120},
                'Precio':    {'text': 'Precio',     'width': 100, 'anchor': 'e'},
                'Costo':     {'text': 'Costo',      'width': 100, 'anchor': 'e'},
                'Stock':     {'text': 'Stock',      'width': 80,  'anchor': 'center'},
                'Mín':       {'text': 'Stock Mín',  'width': 80,  'anchor': 'center'},
                'Estado':    {'text': 'Estado',     'width': 90,  'anchor': 'center'},
            })

        # Tags de stock
        products_tree.tag_configure('stock_ok',       foreground=self.colors['success'])
        products_tree.tag_configure('stock_low',      foreground=self.colors['warning'])
        products_tree.tag_configure('stock_critical', foreground=self.colors['danger'])
        products_tree.tag_configure('inactivo',       foreground=self.colors['text_muted'])

        # ── Paginación ───────────────────────────────────────────────────────
        _, pagination_label, nav_buttons = self._ui_pagination_bar(
            table_card,
            on_prev=lambda: self._products_prev_page(products_tree, pagination_label, nav_buttons),
            on_next=lambda: self._products_next_page(products_tree, pagination_label, nav_buttons),
        )

        # ── Barra de acciones ────────────────────────────────────────────────
        action_card = self._ui_section_card(outer)
        action_card.pack(fill='x', padx=24, pady=(0, 20))

        action_inner = tk.Frame(action_card, bg=self.colors['bg_card'])
        action_inner.pack(fill='x', padx=16, pady=10)

        self._ui_action_btn(action_inner, '✏️  Editar',
                            lambda: self._edit_product_dialog(products_tree))
        self._ui_action_btn(action_inner, '📦  Ajustar Stock',
                            lambda: self._adjust_stock_dialog(products_tree), style='success')
        self._ui_action_btn(action_inner, '🗑️  Eliminar',
                            lambda: self._delete_product(products_tree), style='danger')

        # Cargar datos
        self._load_products_paginated(products_tree, pagination_label, nav_buttons)


    def _products_prev_page(self, tree, label, buttons):
        """Ir a la página anterior de productos"""
        if self.products_current_page > 1:
            self.products_current_page -= 1
            self._load_products_paginated(tree, label, buttons)
    
    def _products_next_page(self, tree, label, buttons):
        """Ir a la página siguiente de productos"""
        total_items = self._get_total_products_count()
        total_pages = (total_items + self.products_items_per_page - 1) // self.products_items_per_page
        
        if self.products_current_page < total_pages:
            self.products_current_page += 1
            self._load_products_paginated(tree, label, buttons)
    
    def _get_total_products_count(self):
        """Obtiene el total de productos según filtro"""
        try:
            productos = self.producto_service.listar_productos(solo_activos=False)
            if not self.products_search_term:
                return len(productos)
            
            search_lower = self.products_search_term.lower()
            count = sum(1 for p in productos 
                       if search_lower in p.nombre.lower() or 
                          (p.categoria and search_lower in p.categoria.lower()))
            return count
        except Exception as e:
            logger.error(f"Error al contar productos: {e}")
            return 0
    
    def _load_products_paginated(self, tree, label, buttons):
        """Carga productos con paginación, colorea stock"""
        for item in tree.get_children():
            tree.delete(item)

        try:
            productos = self.producto_service.listar_productos(solo_activos=False)

            if self.products_search_term:
                sl = self.products_search_term.lower()
                productos = [p for p in productos
                             if sl in p.nombre.lower() or
                                (p.categoria and sl in p.categoria.lower())]

            total_items = len(productos)
            start_idx   = (self.products_current_page - 1) * self.products_items_per_page
            page        = productos[start_idx: start_idx + self.products_items_per_page]

            for i, p in enumerate(page):
                estado = 'Activo' if p.activo else 'Inactivo'
                stripe = 'even' if i % 2 == 0 else 'odd'

                # Color según stock vs mínimo
                if not p.activo:
                    stock_tag = 'inactivo'
                elif p.stock <= 0:
                    stock_tag = 'stock_critical'
                elif p.stock_minimo and p.stock <= p.stock_minimo:
                    stock_tag = 'stock_low'
                else:
                    stock_tag = 'stock_ok'

                tree.insert('', 'end',
                            values=(p.id, p.nombre,
                                    p.categoria or 'N/A',
                                    format_currency(p.precio),
                                    format_currency(p.costo),
                                    p.stock, p.stock_minimo, estado),
                            tags=(stripe, stock_tag))

            total_pages = max(1, (total_items + self.products_items_per_page - 1) // self.products_items_per_page)
            label.config(text=f"Página {self.products_current_page} de {total_pages}  ({total_items} registros)")

            buttons['prev'].config(state='normal' if self.products_current_page > 1 else 'disabled')
            buttons['next'].config(state='normal' if self.products_current_page < total_pages else 'disabled')

        except Exception as e:
            logger.error(f"Error al cargar productos paginados: {e}")
            messagebox.showerror("Error", f"Error al cargar productos: {str(e)}")
            label.config(text="Error al cargar datos")
            buttons['prev'].config(state='disabled')
            buttons['next'].config(state='disabled')
    
    def _load_all_products(self, tree):
        """Carga todos los productos"""
        for item in tree.get_children():
            tree.delete(item)
        
        productos = self.producto_service.listar_productos(solo_activos=False)
        for producto in productos:
            estado = 'Activo' if producto.activo else 'Inactivo'
            tree.insert('', 'end',
                       values=(producto.id,
                              producto.nombre,
                              producto.categoria or 'N/A',
                              format_currency(producto.precio),
                              format_currency(producto.costo),
                              producto.stock,
                              producto.stock_minimo,
                              estado))
    
    def _filter_products(self, tree, search_term):
        """Filtra productos por término de búsqueda"""
        for item in tree.get_children():
            tree.delete(item)
        
        productos = self.producto_service.listar_productos(solo_activos=False)
        search_lower = search_term.lower()
        
        for producto in productos:
            if search_lower in producto.nombre.lower() or (producto.categoria and search_lower in producto.categoria.lower()):
                estado = 'Activo' if producto.activo else 'Inactivo'
                tree.insert('', 'end',
                           values=(producto.id,
                                  producto.nombre,
                                  producto.categoria or 'N/A',
                                  format_currency(producto.precio),
                                  format_currency(producto.costo),
                                  producto.stock,
                                  producto.stock_minimo,
                                  estado))
    
    def _new_product_dialog(self, tree):
        """Diálogo para crear nuevo producto"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Nuevo Producto")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Formulario
        form_frame = ttk.Frame(dialog, padding=20)
        form_frame.pack(fill='both', expand=True)
        
        fields = {}
        row = 0
        
        # Nombre
        ttk.Label(form_frame, text="Nombre:*").grid(row=row, column=0, sticky='w', pady=5)
        fields['nombre'] = ttk.Entry(form_frame, width=40)
        fields['nombre'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Descripción
        ttk.Label(form_frame, text="Descripción:").grid(row=row, column=0, sticky='w', pady=5)
        fields['descripcion'] = ttk.Entry(form_frame, width=40)
        fields['descripcion'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Categoría
        ttk.Label(form_frame, text="Categoría:").grid(row=row, column=0, sticky='w', pady=5)
        fields['categoria'] = ttk.Entry(form_frame, width=40)
        fields['categoria'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Precio
        ttk.Label(form_frame, text="Precio:*").grid(row=row, column=0, sticky='w', pady=5)
        fields['precio'] = ttk.Entry(form_frame, width=40)
        fields['precio'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Costo
        ttk.Label(form_frame, text="Costo:").grid(row=row, column=0, sticky='w', pady=5)
        fields['costo'] = ttk.Entry(form_frame, width=40)
        fields['costo'].insert(0, "0")
        fields['costo'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Stock
        ttk.Label(form_frame, text="Stock Inicial:").grid(row=row, column=0, sticky='w', pady=5)
        fields['stock'] = ttk.Entry(form_frame, width=40)
        fields['stock'].insert(0, "0")
        fields['stock'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Stock mínimo
        ttk.Label(form_frame, text="Stock Mínimo:").grid(row=row, column=0, sticky='w', pady=5)
        fields['stock_minimo'] = ttk.Entry(form_frame, width=40)
        fields['stock_minimo'].insert(0, "5")
        fields['stock_minimo'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Unidad de medida
        ttk.Label(form_frame, text="Unidad:").grid(row=row, column=0, sticky='w', pady=5)
        fields['unidad_medida'] = ttk.Entry(form_frame, width=40)
        fields['unidad_medida'].insert(0, "unidad")
        fields['unidad_medida'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        def save_product():
            try:
                from models import Producto
                producto = Producto()
                producto.nombre = fields['nombre'].get().strip()
                producto.descripcion = fields['descripcion'].get().strip()
                producto.categoria = fields['categoria'].get().strip()
                producto.precio = float(fields['precio'].get())
                producto.costo = float(fields['costo'].get())
                producto.stock = int(fields['stock'].get())
                producto.stock_minimo = int(fields['stock_minimo'].get())
                producto.unidad_medida = fields['unidad_medida'].get().strip()
                producto.activo = True
                
                success, message, producto_id = self.producto_service.crear_producto(
                    producto,
                    self.auth_service.current_user.id
                )
                
                if success:
                    messagebox.showinfo("Éxito", message)
                    self._load_all_products(tree)
                    dialog.destroy()
                else:
                    messagebox.showerror("Error", message)
            except ValueError:
                messagebox.showerror("Error", "Verifique que los valores numéricos sean correctos")
        
        # Botones
        btn_frame = ttk.Frame(form_frame)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=20)
        
        ttk.Button(btn_frame, text="💾 Guardar", command=save_product, style='Success.TButton').pack(side='left', padx=5)
        ttk.Button(btn_frame, text="✕ Cancelar", command=dialog.destroy, style='Danger.TButton').pack(side='left', padx=5)
    
    def _edit_product_dialog(self, tree):
        """Diálogo para editar producto"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un producto")
            return
        
        producto_id = tree.item(selection[0])['values'][0]
        producto = self.producto_service.obtener_producto(producto_id)
        
        if not producto:
            messagebox.showerror("Error", "Producto no encontrado")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Editar Producto #{producto_id}")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Formulario (similar a _new_product_dialog pero con valores precargados)
        form_frame = ttk.Frame(dialog, padding=20)
        form_frame.pack(fill='both', expand=True)
        
        fields = {}
        row = 0
        
        # Nombre
        ttk.Label(form_frame, text="Nombre:*").grid(row=row, column=0, sticky='w', pady=5)
        fields['nombre'] = ttk.Entry(form_frame, width=40)
        fields['nombre'].insert(0, producto.nombre)
        fields['nombre'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Descripción
        ttk.Label(form_frame, text="Descripción:").grid(row=row, column=0, sticky='w', pady=5)
        fields['descripcion'] = ttk.Entry(form_frame, width=40)
        fields['descripcion'].insert(0, producto.descripcion or '')
        fields['descripcion'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Categoría
        ttk.Label(form_frame, text="Categoría:").grid(row=row, column=0, sticky='w', pady=5)
        fields['categoria'] = ttk.Entry(form_frame, width=40)
        fields['categoria'].insert(0, producto.categoria or '')
        fields['categoria'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Precio
        ttk.Label(form_frame, text="Precio:*").grid(row=row, column=0, sticky='w', pady=5)
        fields['precio'] = ttk.Entry(form_frame, width=40)
        fields['precio'].insert(0, str(producto.precio))
        fields['precio'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Costo
        ttk.Label(form_frame, text="Costo:").grid(row=row, column=0, sticky='w', pady=5)
        fields['costo'] = ttk.Entry(form_frame, width=40)
        fields['costo'].insert(0, str(producto.costo))
        fields['costo'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Stock mínimo
        ttk.Label(form_frame, text="Stock Mínimo:").grid(row=row, column=0, sticky='w', pady=5)
        fields['stock_minimo'] = ttk.Entry(form_frame, width=40)
        fields['stock_minimo'].insert(0, str(producto.stock_minimo))
        fields['stock_minimo'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        # Unidad de medida
        ttk.Label(form_frame, text="Unidad:").grid(row=row, column=0, sticky='w', pady=5)
        fields['unidad_medida'] = ttk.Entry(form_frame, width=40)
        fields['unidad_medida'].insert(0, producto.unidad_medida or 'unidad')
        fields['unidad_medida'].grid(row=row, column=1, pady=5, padx=10)
        row += 1
        
        def update_product():
            try:
                producto.nombre = fields['nombre'].get().strip()
                producto.descripcion = fields['descripcion'].get().strip()
                producto.categoria = fields['categoria'].get().strip()
                producto.precio = float(fields['precio'].get())
                producto.costo = float(fields['costo'].get())
                producto.stock_minimo = int(fields['stock_minimo'].get())
                producto.unidad_medida = fields['unidad_medida'].get().strip()
                
                success, message = self.producto_service.actualizar_producto(producto_id, producto)
                
                if success:
                    messagebox.showinfo("Éxito", message)
                    self._load_all_products(tree)
                    dialog.destroy()
                else:
                    messagebox.showerror("Error", message)
            except ValueError:
                messagebox.showerror("Error", "Verifique que los valores numéricos sean correctos")
        
        # Botones
        btn_frame = ttk.Frame(form_frame)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=20)
        
        ttk.Button(btn_frame, text="💾 Guardar", command=update_product, style='Success.TButton').pack(side='left', padx=5)
        ttk.Button(btn_frame, text="✕ Cancelar", command=dialog.destroy, style='Danger.TButton').pack(side='left', padx=5)
    
    def _adjust_stock_dialog(self, tree):
        """Diálogo para ajustar stock"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un producto")
            return
        
        producto_id = tree.item(selection[0])['values'][0]
        producto = self.producto_service.obtener_producto(producto_id)
        
        if not producto:
            messagebox.showerror("Error", "Producto no encontrado")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Ajustar Stock - {producto.nombre}")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        form_frame = ttk.Frame(dialog, padding=20)
        form_frame.pack(fill='both', expand=True)
        
        # Stock actual
        ttk.Label(form_frame,
                 text=f"Stock actual: {producto.stock}",
                 font=('Segoe UI', 14, 'bold')).pack(pady=10)
        
        # Tipo de ajuste
        ttk.Label(form_frame, text="Tipo de ajuste:").pack(pady=5)
        ajuste_tipo = ttk.Combobox(form_frame,
                                   values=['Entrada (Compra)', 'Salida (Ajuste)', 'Corrección'],
                                   state='readonly',
                                   width=30)
        ajuste_tipo.set('Entrada (Compra)')
        ajuste_tipo.pack(pady=5)
        
        # Cantidad
        ttk.Label(form_frame, text="Cantidad:").pack(pady=5)
        cantidad_entry = ttk.Entry(form_frame, width=30)
        cantidad_entry.pack(pady=5)
        
        # Motivo
        ttk.Label(form_frame, text="Motivo:").pack(pady=5)
        motivo_entry = ttk.Entry(form_frame, width=30)
        motivo_entry.pack(pady=5)
        
        def apply_adjustment():
            try:
                cantidad = int(cantidad_entry.get())
                motivo = motivo_entry.get().strip()
                tipo = ajuste_tipo.get()
                
                if cantidad <= 0:
                    messagebox.showerror("Error", "La cantidad debe ser mayor a 0")
                    return
                
                # Determinar tipo de movimiento
                if 'Entrada' in tipo:
                    tipo_mov = 'ajuste'
                elif 'Salida' in tipo:
                    tipo_mov = 'merma'
                else:
                    tipo_mov = 'ajuste'
                
                success, message = self.producto_service.ajustar_stock(
                    producto_id,
                    cantidad if 'Entrada' in tipo else -cantidad,
                    tipo_mov,
                    self.auth_service.current_user.id,
                    notas=motivo
                )
                
                if success:
                    messagebox.showinfo("Éxito", message)
                    self._load_all_products(tree)
                    dialog.destroy()
                else:
                    messagebox.showerror("Error", message)
            except ValueError:
                messagebox.showerror("Error", "La cantidad debe ser un número entero")
        
        # Botones
        btn_frame = ttk.Frame(form_frame)
        btn_frame.pack(pady=20)
        
        ttk.Button(btn_frame, text="Aplicar", command=apply_adjustment, style='Success.TButton').pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=dialog.destroy, style='Danger.TButton').pack(side='left', padx=5)
    
    def _delete_product(self, tree):
        """Elimina (desactiva) un producto"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un producto")
            return
        
        producto_id = tree.item(selection[0])['values'][0]
        nombre = tree.item(selection[0])['values'][1]
        
        if messagebox.askyesno("Confirmar", f"¿Desea eliminar el producto '{nombre}'?"):
            success = self.db.execute_query(
                "UPDATE productos SET activo = 0 WHERE id = ?",
                (producto_id,)
            )
            
            if success:
                messagebox.showinfo("Éxito", "Producto eliminado correctamente")
                self._load_all_products(tree)
            else:
                messagebox.showerror("Error", "No se pudo eliminar el producto")
    
    def show_orders_manager(self):
        """Muestra el gestor de pedidos"""
        self.clear_content()
        self.current_screen = "orders"
        
        # Configurar grid
        self.content_frame.grid_columnconfigure(0, weight=1)
        
        # Título
        title = ttk.Label(self.content_frame,
                         text="Gestión de Pedidos a Proveedores",
                         font=(self.config.get('fonts', {}).get('family', 'Segoe UI'), 20, 'bold'),
                         background=self.colors.get('bg_light', '#ECF0F1'),
                         foreground=self.colors.get('primary', '#2C3E50'))
        title.pack(pady=20)
        
        # Frame principal con tabs
        tab_control = ttk.Notebook(self.content_frame)
        tab_control.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Tab de Pedidos
        pedidos_tab = ttk.Frame(tab_control, style='Main.TFrame')
        tab_control.add(pedidos_tab, text='📦 Pedidos')
        
        # Tab de Proveedores
        proveedores_tab = ttk.Frame(tab_control, style='Main.TFrame')
        tab_control.add(proveedores_tab, text='🏢 Proveedores')
        
        # Configurar tabs
        self._setup_pedidos_tab(pedidos_tab)
        self._setup_proveedores_tab(proveedores_tab)
    
    def _setup_pedidos_tab(self, parent):
        """Configura el tab de pedidos"""
        # Variables de paginación
        self.pedidos_current_page = 1
        self.pedidos_items_per_page = 50
        
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)
        
        # Barra de herramientas
        toolbar = ttk.Frame(parent, style='Card.TFrame')
        toolbar.grid(row=0, column=0, sticky='ew', padx=10, pady=10)
        
        ttk.Button(toolbar,
                  text="➕ Nuevo Pedido",
                  command=self._new_pedido_dialog,
                  style='Success.TButton').pack(side='left', padx=5, pady=10)
        
        ttk.Button(toolbar,
                  text="👁️ Ver Detalles",
                  command=lambda: self._view_pedido_details(),
                  style='Primary.TButton').pack(side='left', padx=5)
        
        ttk.Button(toolbar,
                  text="📥 Recibir Pedido",
                  command=lambda: self._receive_pedido(),
                  style='Primary.TButton').pack(side='left', padx=5)
        
        ttk.Button(toolbar,
                  text="❌ Cancelar",
                  command=lambda: self._cancel_pedido(),
                  style='Danger.TButton').pack(side='left', padx=5)
        
        # Registros por página
        ttk.Label(toolbar, text="Por página:", style='Info.TLabel').pack(side='right', padx=5)
        items_per_page_var = tk.StringVar(value=str(self.pedidos_items_per_page))
        items_combo = ttk.Combobox(toolbar, textvariable=items_per_page_var,
                                   width=8, state='readonly')
        items_combo['values'] = ('25', '50', '100', '200')
        items_combo.pack(side='right', padx=5)
        
        def change_page_size(event=None):
            self.pedidos_items_per_page = int(items_per_page_var.get())
            self.pedidos_current_page = 1
            self._load_pedidos_paginated(pedidos_tree, pagination_label, nav_buttons)
        
        items_combo.bind('<<ComboboxSelected>>', change_page_size)
        
        ttk.Button(toolbar,
                  text="🔄 Actualizar",
                  command=lambda: self._load_pedidos_paginated(pedidos_tree, pagination_label, nav_buttons),
                  style='Primary.TButton').pack(side='right', padx=5)
        
        # Frame para el TreeView
        tree_frame = ttk.Frame(parent, style='Card.TFrame')
        tree_frame.grid(row=1, column=0, sticky='nsew', padx=10, pady=10)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)
        
        # TreeView de pedidos
        columns = ('numero', 'proveedor', 'fecha', 'entrega', 'estado', 'total')
        pedidos_tree = ttk.Treeview(tree_frame, columns=columns, show='headings',
                                    style='Modern.Treeview')
        
        pedidos_tree.heading('numero', text='N° Pedido')
        pedidos_tree.heading('proveedor', text='Proveedor')
        pedidos_tree.heading('fecha', text='Fecha Pedido')
        pedidos_tree.heading('entrega', text='Entrega Estimada')
        pedidos_tree.heading('estado', text='Estado')
        pedidos_tree.heading('total', text='Total')
        
        pedidos_tree.column('numero', width=100)
        pedidos_tree.column('proveedor', width=200)
        pedidos_tree.column('fecha', width=120)
        pedidos_tree.column('entrega', width=120)
        pedidos_tree.column('estado', width=100)
        pedidos_tree.column('total', width=100)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=pedidos_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=pedidos_tree.xview)
        pedidos_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        pedidos_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        # Frame de paginación
        pagination_frame = ttk.Frame(tree_frame, style='Card.TFrame')
        pagination_frame.grid(row=2, column=0, columnspan=2, sticky='ew', pady=(10, 0))
        pagination_frame.grid_columnconfigure(1, weight=1)
        
        # Botón anterior
        prev_btn = ttk.Button(pagination_frame,
                             text="◀ Anterior",
                             command=lambda: self._pedidos_prev_page(pedidos_tree, pagination_label, nav_buttons),
                             style='Secondary.TButton',
                             width=12)
        prev_btn.pack(side='left', padx=5)
        
        # Label de página
        pagination_label = ttk.Label(pagination_frame,
                                    text="Página 1 de 1 (0 registros)",
                                    font=('Segoe UI', 10),
                                    foreground=self.colors.get('text', '#2C3E50'))
        pagination_label.pack(side='left', expand=True)
        
        # Botón siguiente
        next_btn = ttk.Button(pagination_frame,
                             text="Siguiente ▶",
                             command=lambda: self._pedidos_next_page(pedidos_tree, pagination_label, nav_buttons),
                             style='Secondary.TButton',
                             width=12)
        next_btn.pack(side='left', padx=5)
        
        # Guardar referencias
        nav_buttons = {'prev': prev_btn, 'next': next_btn}
        self.pedidos_pagination_label = pagination_label
        self.pedidos_nav_buttons = nav_buttons
        
        # Guardar referencia
        self.pedidos_tree = pedidos_tree
        
        # Cargar pedidos con paginación
        self._load_pedidos_paginated(pedidos_tree, pagination_label, nav_buttons)
    
    def _setup_proveedores_tab(self, parent):
        """Configura el tab de proveedores"""
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)
        
        # Barra de herramientas
        toolbar = ttk.Frame(parent, style='Card.TFrame')
        toolbar.grid(row=0, column=0, sticky='ew', padx=10, pady=10)
        
        ttk.Button(toolbar,
                  text="➕ Nuevo Proveedor",
                  command=self._new_proveedor_dialog,
                  style='Success.TButton').pack(side='left', padx=5, pady=10)
        
        ttk.Button(toolbar,
                  text="✏️ Editar",
                  command=lambda: self._edit_proveedor(),
                  style='Primary.TButton').pack(side='left', padx=5)
        
        ttk.Button(toolbar,
                  text="❌ Eliminar",
                  command=lambda: self._delete_proveedor(),
                  style='Danger.TButton').pack(side='left', padx=5)
        
        ttk.Button(toolbar,
                  text="🔄 Actualizar",
                  command=lambda: self._load_proveedores(proveedores_tree),
                  style='Primary.TButton').pack(side='right', padx=5)
        
        # Frame para el TreeView
        tree_frame = ttk.Frame(parent, style='Card.TFrame')
        tree_frame.grid(row=1, column=0, sticky='nsew', padx=10, pady=10)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)
        
        # TreeView de proveedores
        columns = ('nombre', 'contacto', 'telefono', 'email', 'activo')
        proveedores_tree = ttk.Treeview(tree_frame, columns=columns, show='headings',
                                       style='Modern.Treeview')
        
        proveedores_tree.heading('nombre', text='Nombre')
        proveedores_tree.heading('contacto', text='Contacto')
        proveedores_tree.heading('telefono', text='Teléfono')
        proveedores_tree.heading('email', text='Email')
        proveedores_tree.heading('activo', text='Estado')
        
        proveedores_tree.column('nombre', width=200)
        proveedores_tree.column('contacto', width=150)
        proveedores_tree.column('telefono', width=120)
        proveedores_tree.column('email', width=180)
        proveedores_tree.column('activo', width=80)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=proveedores_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=proveedores_tree.xview)
        proveedores_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        proveedores_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        # Guardar referencia
        self.proveedores_tree = proveedores_tree
        
        # Cargar proveedores
        self._load_proveedores(proveedores_tree)
    
    def _load_pedidos(self, tree):
        """Carga los pedidos en el TreeView (versión simple para compatibilidad)"""
        # Si tenemos referencias de paginación, usar versión paginada
        if hasattr(self, 'pedidos_pagination_label') and hasattr(self, 'pedidos_nav_buttons'):
            self._load_pedidos_paginated(tree, self.pedidos_pagination_label, self.pedidos_nav_buttons)
            return
        
        # Versión sin paginación (fallback)
        # Limpiar TreeView
        for item in tree.get_children():
            tree.delete(item)
        
        try:
            pedidos = self.db.fetch_all("""
                SELECT p.id, p.numero_pedido, p.proveedor,
                       p.fecha, p.fecha_entrega, p.estado, p.total
                FROM pedidos p
                ORDER BY p.fecha DESC
            """)
            
            for pedido in pedidos:
                fecha_pedido = pedido['fecha'][:10] if pedido['fecha'] else ''
                fecha_entrega = pedido['fecha_entrega'] if pedido['fecha_entrega'] else 'Sin fecha'
                
                tree.insert('', 'end', iid=pedido['id'], values=(
                    pedido['numero_pedido'],
                    pedido['proveedor'] or 'Sin proveedor',
                    fecha_pedido,
                    fecha_entrega,
                    pedido['estado'].upper(),
                    format_currency(pedido['total'])
                ))
        except Exception as e:
            logger.error(f"Error al cargar pedidos: {e}")
            messagebox.showerror("Error", f"No se pudieron cargar los pedidos: {str(e)}")
    
    def _pedidos_prev_page(self, tree, label, buttons):
        """Ir a la página anterior de pedidos"""
        if self.pedidos_current_page > 1:
            self.pedidos_current_page -= 1
            self._load_pedidos_paginated(tree, label, buttons)
    
    def _pedidos_next_page(self, tree, label, buttons):
        """Ir a la página siguiente de pedidos"""
        total_items = self._get_total_pedidos_count()
        total_pages = (total_items + self.pedidos_items_per_page - 1) // self.pedidos_items_per_page
        
        if self.pedidos_current_page < total_pages:
            self.pedidos_current_page += 1
            self._load_pedidos_paginated(tree, label, buttons)
    
    def _get_total_pedidos_count(self):
        """Obtiene el total de pedidos"""
        try:
            result = self.db.fetch_one("SELECT COUNT(*) as total FROM pedidos")
            return result['total'] if result else 0
        except Exception as e:
            logger.error(f"Error al contar pedidos: {e}")
            return 0
    
    def _load_pedidos_paginated(self, tree, label, buttons):
        """Carga los pedidos con paginación"""
        # Limpiar TreeView
        for item in tree.get_children():
            tree.delete(item)
        
        try:
            # Calcular offset
            offset = (self.pedidos_current_page - 1) * self.pedidos_items_per_page
            
            # Consultar con LIMIT y OFFSET - campos correctos de la tabla actual
            pedidos = self.db.fetch_all("""
                SELECT p.id, p.numero_pedido, p.proveedor,
                       p.fecha, p.fecha_entrega, p.estado, p.total
                FROM pedidos p
                ORDER BY p.fecha DESC
                LIMIT ? OFFSET ?
            """, (self.pedidos_items_per_page, offset))
            
            for pedido in pedidos:
                fecha_pedido = pedido['fecha'][:10] if pedido['fecha'] else ''
                fecha_entrega = pedido['fecha_entrega'] if pedido['fecha_entrega'] else 'Sin fecha'
                
                tree.insert('', 'end', iid=pedido['id'], values=(
                    pedido['numero_pedido'],
                    pedido['proveedor'] or 'Sin proveedor',
                    fecha_pedido,
                    fecha_entrega,
                    pedido['estado'].upper(),
                    format_currency(pedido['total'])
                ))
            
            # Actualizar label de paginación
            total_items = self._get_total_pedidos_count()
            total_pages = max(1, (total_items + self.pedidos_items_per_page - 1) // self.pedidos_items_per_page)
            label.config(text=f"Página {self.pedidos_current_page} de {total_pages} ({total_items} registros)")
            
            # Actualizar estado de botones
            buttons['prev'].config(state='normal' if self.pedidos_current_page > 1 else 'disabled')
            buttons['next'].config(state='normal' if self.pedidos_current_page < total_pages else 'disabled')
            
        except Exception as e:
            logger.error(f"Error al cargar pedidos paginados: {e}")
            messagebox.showerror("Error", f"No se pudieron cargar los pedidos: {str(e)}")
            label.config(text="Error al cargar datos")
            buttons['prev'].config(state='disabled')
            buttons['next'].config(state='disabled')
    
    def _load_proveedores(self, tree):
        """Carga los proveedores en el TreeView"""
        # Limpiar TreeView
        for item in tree.get_children():
            tree.delete(item)
        
        try:
            proveedores = self.db.fetch_all("""
                SELECT id, nombre, contacto, telefono, email, activo
                FROM proveedores
                ORDER BY nombre
            """)
            
            for prov in proveedores:
                estado = 'Activo' if prov['activo'] == 1 else 'Inactivo'
                tree.insert('', 'end', iid=prov['id'], values=(
                    prov['nombre'],
                    prov['contacto'] or '',
                    prov['telefono'] or '',
                    prov['email'] or '',
                    estado
                ))
        except Exception as e:
            logger.error(f"Error al cargar proveedores: {e}")
            messagebox.showerror("Error", f"No se pudieron cargar los proveedores: {str(e)}")
    
    def _new_proveedor_dialog(self):
        """Muestra el diálogo para crear un nuevo proveedor"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Nuevo Proveedor")
        dialog.geometry("500x550")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Frame principal
        main_frame = ttk.Frame(dialog, style='Card.TFrame', padding=20)
        main_frame.pack(fill='both', expand=True)
        
        # Campos
        row = 0
        ttk.Label(main_frame, text="Nombre *:", style='Info.TLabel').grid(row=row, column=0, sticky='w', pady=5)
        nombre_entry = ttk.Entry(main_frame, width=40, font=('Segoe UI', 11))
        nombre_entry.grid(row=row, column=1, pady=5, padx=10)
        nombre_entry.focus()
        
        row += 1
        ttk.Label(main_frame, text="Contacto:", style='Info.TLabel').grid(row=row, column=0, sticky='w', pady=5)
        contacto_entry = ttk.Entry(main_frame, width=40, font=('Segoe UI', 11))
        contacto_entry.grid(row=row, column=1, pady=5, padx=10)
        
        row += 1
        ttk.Label(main_frame, text="Teléfono:", style='Info.TLabel').grid(row=row, column=0, sticky='w', pady=5)
        telefono_entry = ttk.Entry(main_frame, width=40, font=('Segoe UI', 11))
        telefono_entry.grid(row=row, column=1, pady=5, padx=10)
        
        row += 1
        ttk.Label(main_frame, text="Email:", style='Info.TLabel').grid(row=row, column=0, sticky='w', pady=5)
        email_entry = ttk.Entry(main_frame, width=40, font=('Segoe UI', 11))
        email_entry.grid(row=row, column=1, pady=5, padx=10)
        
        row += 1
        ttk.Label(main_frame, text="Dirección:", style='Info.TLabel').grid(row=row, column=0, sticky='w', pady=5)
        direccion_text = tk.Text(main_frame, width=40, height=3, font=('Segoe UI', 10))
        direccion_text.grid(row=row, column=1, pady=5, padx=10)
        
        row += 1
        ttk.Label(main_frame, text="Notas:", style='Info.TLabel').grid(row=row, column=0, sticky='w', pady=5)
        notas_text = tk.Text(main_frame, width=40, height=3, font=('Segoe UI', 10))
        notas_text.grid(row=row, column=1, pady=5, padx=10)
        
        # Botones
        row += 1
        btn_frame = ttk.Frame(main_frame, style='Card.TFrame')
        btn_frame.grid(row=row, column=0, columnspan=2, pady=20)
        
        def save_proveedor():
            nombre = nombre_entry.get().strip()
            if not nombre:
                messagebox.showwarning("Advertencia", "El nombre es obligatorio", parent=dialog)
                return
            
            try:
                self.db.execute_query("""
                    INSERT INTO proveedores (nombre, contacto, telefono, email, direccion, notas, activo)
                    VALUES (?, ?, ?, ?, ?, ?, 1)
                """, (
                    nombre,
                    contacto_entry.get().strip() or None,
                    telefono_entry.get().strip() or None,
                    email_entry.get().strip() or None,
                    direccion_text.get('1.0', 'end').strip() or None,
                    notas_text.get('1.0', 'end').strip() or None
                ))
                
                messagebox.showinfo("Éxito", "Proveedor creado correctamente", parent=dialog)
                self._load_proveedores(self.proveedores_tree)
                dialog.destroy()
            except Exception as e:
                logger.error(f"Error al crear proveedor: {e}")
                messagebox.showerror("Error", f"No se pudo crear el proveedor: {str(e)}", parent=dialog)
        
        ttk.Button(btn_frame, text="Guardar", command=save_proveedor,
                  style='Success.TButton').pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=dialog.destroy,
                  style='Danger.TButton').pack(side='left', padx=5)
    
    def _new_pedido_dialog(self):
        """Muestra el diálogo para crear un nuevo pedido"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Nuevo Pedido a Proveedor")
        dialog.geometry("1000x700")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Centrar ventana
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (1000 // 2)
        y = (dialog.winfo_screenheight() // 2) - (700 // 2)
        dialog.geometry(f"1000x700+{x}+{y}")
        
        # Frame principal con canvas scrollable
        canvas = tk.Canvas(dialog, bg=self.colors.get('bg_light', '#ECF0F1'), highlightthickness=0)
        scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        main_frame = ttk.Frame(canvas, padding=20)
        
        main_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=main_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Mouse wheel support
        def _on_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except:
                pass
        
        # Usar bind en lugar de bind_all
        canvas.bind("<MouseWheel>", _on_mousewheel)
        main_frame.bind("<MouseWheel>", _on_mousewheel)
        
        # Limpiar binding cuando se destruye
        def _on_destroy(event):
            try:
                canvas.unbind("<MouseWheel>")
                main_frame.unbind("<MouseWheel>")
            except:
                pass
        canvas.bind("<Destroy>", _on_destroy)
        
        # Título
        ttk.Label(main_frame, text="Nuevo Pedido a Proveedor",
                 font=('Segoe UI', 16, 'bold')).pack(pady=(0, 20))
        
        # Frame de información del pedido
        info_frame = ttk.LabelFrame(main_frame, text="Información del Pedido", padding=15)
        info_frame.pack(fill='x', pady=(0, 10))
        
        # Proveedor
        ttk.Label(info_frame, text="Proveedor:").grid(row=0, column=0, sticky='w', pady=5, padx=5)
        proveedores = self.db.fetch_all("SELECT id, nombre FROM proveedores WHERE activo = 1")
        proveedor_var = tk.StringVar()
        proveedor_combo = ttk.Combobox(info_frame, textvariable=proveedor_var, state='readonly', width=40)
        proveedor_combo['values'] = [f"{p['id']} - {p['nombre']}" for p in proveedores]
        proveedor_combo.grid(row=0, column=1, sticky='ew', pady=5, padx=5)
        if proveedores:
            proveedor_combo.current(0)
        
        # Fecha de entrega estimada
        ttk.Label(info_frame, text="Fecha Entrega Estimada:").grid(row=0, column=2, sticky='w', pady=5, padx=5)
        from datetime import timedelta
        fecha_entrega = ttk.Entry(info_frame, width=15)
        fecha_entrega.insert(0, (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d'))
        fecha_entrega.grid(row=0, column=3, sticky='w', pady=5, padx=5)
        
        # Notas
        ttk.Label(info_frame, text="Notas:").grid(row=1, column=0, sticky='w', pady=5, padx=5)
        notas_entry = ttk.Entry(info_frame, width=80)
        notas_entry.grid(row=1, column=1, columnspan=3, sticky='ew', pady=5, padx=5)
        
        info_frame.grid_columnconfigure(1, weight=1)
        
        # Frame de productos
        productos_frame = ttk.LabelFrame(main_frame, text="Productos", padding=15)
        productos_frame.pack(fill='x', pady=(0, 10))
        
        # Lista de productos disponibles (izquierda)
        left_frame = ttk.Frame(productos_frame)
        left_frame.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        ttk.Label(left_frame, text="Productos Disponibles", font=('Segoe UI', 11, 'bold')).pack()
        
        productos_tree = ttk.Treeview(left_frame, columns=('Nombre', 'Stock', 'Costo'),
                                     show='headings', style='Modern.Treeview', height=8)
        productos_tree.heading('Nombre', text='Producto')
        productos_tree.heading('Stock', text='Stock')
        productos_tree.heading('Costo', text='Costo')
        productos_tree.column('Nombre', width=200)
        productos_tree.column('Stock', width=80)
        productos_tree.column('Costo', width=100)
        productos_tree.pack(fill='both', expand=True)
        
        # Cargar productos
        productos = self.producto_service.listar_productos(solo_activos=True)
        for p in productos:
            productos_tree.insert('', 'end', values=(p.nombre, p.stock, format_currency(p.costo)),
                                tags=(p.id,))
        
        # Botones centrales
        center_frame = ttk.Frame(productos_frame)
        center_frame.pack(side='left', padx=5)
        
        def add_to_pedido():
            selection = productos_tree.selection()
            if not selection:
                messagebox.showwarning("Advertencia", "Seleccione un producto")
                return
            
            item = productos_tree.item(selection[0])
            producto_id = int(item['tags'][0])
            nombre = item['values'][0]
            costo_str = item['values'][2].replace('$', '').replace(',', '')
            costo = float(costo_str)
            
            # Validar que el costo sea mayor a 0
            if costo <= 0:
                messagebox.showerror("Error", 
                    f"El producto '{nombre}' tiene costo $0.00.\n" +
                    "Debe actualizar el costo del producto antes de agregarlo al pedido.")
                return
            
            # Diálogo para cantidad
            cantidad = tk.simpledialog.askinteger("Cantidad", f"Cantidad de {nombre}:", minvalue=1, maxvalue=10000)
            if cantidad:
                # Verificar si ya está en el pedido
                for child in pedido_tree.get_children():
                    if pedido_tree.item(child)['tags'][0] == producto_id:
                        messagebox.showinfo("Info", "El producto ya está en el pedido")
                        return
                
                subtotal = cantidad * costo
                pedido_tree.insert('', 'end', values=(nombre, cantidad, format_currency(costo),
                                                     format_currency(subtotal)), tags=(producto_id,))
                actualizar_total()
        
        def remove_from_pedido():
            selection = pedido_tree.selection()
            if selection:
                pedido_tree.delete(selection[0])
                actualizar_total()
        
        ttk.Button(center_frame, text="➡️\nAgregar", command=add_to_pedido,
                  style='Success.TButton').pack(pady=5)
        ttk.Button(center_frame, text="⬅️\nQuitar", command=remove_from_pedido,
                  style='Danger.TButton').pack(pady=5)
        
        # Lista de productos del pedido (derecha)
        right_frame = ttk.Frame(productos_frame)
        right_frame.pack(side='left', fill='both', expand=True)
        
        ttk.Label(right_frame, text="Productos del Pedido", font=('Segoe UI', 11, 'bold')).pack()
        
        pedido_tree = ttk.Treeview(right_frame, columns=('Producto', 'Cantidad', 'Precio', 'Subtotal'),
                                  show='headings', style='Modern.Treeview', height=8)
        pedido_tree.heading('Producto', text='Producto')
        pedido_tree.heading('Cantidad', text='Cantidad')
        pedido_tree.heading('Precio', text='Precio Unit.')
        pedido_tree.heading('Subtotal', text='Subtotal')
        pedido_tree.column('Producto', width=180)
        pedido_tree.column('Cantidad', width=80)
        pedido_tree.column('Precio', width=90)
        pedido_tree.column('Subtotal', width=90)
        pedido_tree.pack(fill='both', expand=True)
        
        # Frame de totales
        totales_frame = ttk.Frame(main_frame)
        totales_frame.pack(fill='x', pady=10)
        
        total_label = ttk.Label(totales_frame, text="TOTAL: $0.00",
                               font=('Segoe UI', 14, 'bold'))
        total_label.pack(side='right', padx=20)
        
        def actualizar_total():
            total = 0
            for child in pedido_tree.get_children():
                subtotal_str = pedido_tree.item(child)['values'][3].replace('$', '').replace(',', '')
                total += float(subtotal_str)
            total_label.config(text=f"TOTAL: {format_currency(total)}")
        
        # Botones de acción
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=20)
        
        def guardar_pedido():
            if not proveedor_var.get():
                messagebox.showerror("Error", "Debe seleccionar un proveedor", parent=dialog)
                return
            
            if not pedido_tree.get_children():
                messagebox.showerror("Error", "Debe agregar al menos un producto al pedido", parent=dialog)
                return
            
            try:
                # Obtener proveedor (nombre del combo)
                proveedor_texto = proveedor_var.get()
                proveedor_nombre = proveedor_texto.split(' - ')[1] if ' - ' in proveedor_texto else proveedor_texto
                
                # Generar número de pedido
                from utils import generate_order_number
                numero_pedido = generate_order_number()
                
                # Calcular totales
                subtotal = 0
                for child in pedido_tree.get_children():
                    subtotal_str = pedido_tree.item(child)['values'][3].replace('$', '').replace(',', '')
                    subtotal += float(subtotal_str)
                
                impuestos = subtotal * 0.16  # 16% IVA
                total = subtotal + impuestos
                
                # Insertar pedido con estado 'pendiente' - USANDO CAMPOS CORRECTOS
                self.db.execute_query(
                    """INSERT INTO pedidos (numero_pedido, proveedor, fecha, 
                       fecha_entrega, estado, total, notas, usuario_id)
                       VALUES (?, ?, ?, ?, 'pendiente', ?, ?, ?)""",
                    (numero_pedido, proveedor_nombre, datetime.now(), fecha_entrega.get(),
                     total, notas_entry.get(), self.auth_service.current_user.id)
                )
                
                # Obtener ID del pedido
                pedido_id = self.db.fetch_one("SELECT last_insert_rowid() as id")['id']
                
                # Insertar detalles del pedido - SIN CAMPO 'recibido'
                for child in pedido_tree.get_children():
                    values = pedido_tree.item(child)['values']
                    producto_id = pedido_tree.item(child)['tags'][0]
                    cantidad = int(values[1])
                    precio_str = values[2].replace('$', '').replace(',', '')
                    precio = float(precio_str)
                    subtotal_item_str = values[3].replace('$', '').replace(',', '')
                    subtotal_item = float(subtotal_item_str)
                    
                    # Validar que precio sea mayor a 0
                    if precio <= 0:
                        messagebox.showerror("Error", 
                            f"El precio unitario debe ser mayor a 0.\nProducto: {values[0]}", 
                            parent=dialog)
                        return
                    
                    self.db.execute_query(
                        """INSERT INTO detalles_pedido (pedido_id, producto_id, cantidad, 
                           precio_unitario, subtotal)
                           VALUES (?, ?, ?, ?, ?)""",
                        (pedido_id, producto_id, cantidad, precio, subtotal_item)
                    )
                
                messagebox.showinfo("Éxito", 
                    f"Pedido creado correctamente\nNúmero: {numero_pedido}\n\n" +
                    f"El pedido quedó en estado PENDIENTE.\n" +
                    f"Cuando se reciba la mercancía, use la opción 'Recibir Pedido'\n" +
                    f"para actualizar el stock automáticamente.",
                    parent=dialog)
                self._load_pedidos(self.pedidos_tree)
                dialog.destroy()
                
            except Exception as e:
                logger.error(f"Error al crear pedido: {e}")
                messagebox.showerror("Error", f"No se pudo crear el pedido: {str(e)}", parent=dialog)
        
        # Botones grandes y centrados
        button_container = ttk.Frame(btn_frame)
        button_container.pack(expand=True)
        
        save_btn = ttk.Button(button_container, text="✅ GUARDAR PEDIDO", command=guardar_pedido,
                  style='Success.TButton')
        save_btn.pack(side='left', padx=10, ipadx=20, ipady=10)
        
        cancel_btn = ttk.Button(button_container, text="❌ CANCELAR", command=dialog.destroy,
                  style='Danger.TButton')
        cancel_btn.pack(side='left', padx=10, ipadx=20, ipady=10)
    
    def _view_pedido_details(self):
        """Muestra los detalles de un pedido"""
        selection = self.pedidos_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un pedido")
            return
        
        try:
            pedido_id = selection[0]
            
            # Obtener información del pedido - USANDO CAMPOS CORRECTOS
            pedido = self.db.fetch_one(
                """SELECT p.*, u.nombre as usuario_nombre
                   FROM pedidos p
                   LEFT JOIN usuarios u ON p.usuario_id = u.id
                   WHERE p.id = ?""",
                (pedido_id,)
            )
            
            if not pedido:
                messagebox.showerror("Error", "Pedido no encontrado")
                return
            
            # Obtener detalles del pedido
            detalles = self.db.fetch_all(
                """SELECT dp.*, prod.nombre as producto_nombre
                   FROM detalles_pedido dp
                   LEFT JOIN productos prod ON dp.producto_id = prod.id
                   WHERE dp.pedido_id = ?""",
                (pedido_id,)
            )
            
            # Crear ventana de detalles
            details_window = tk.Toplevel(self.root)
            details_window.title(f"Detalles del Pedido {pedido['numero_pedido']}")
            details_window.geometry("800x600")
            details_window.transient(self.root)
            
            main_frame = ttk.Frame(details_window, padding=20)
            main_frame.pack(fill='both', expand=True)
            
            # Título
            ttk.Label(main_frame, text=f"Detalles del Pedido {pedido['numero_pedido']}",
                     font=('Segoe UI', 16, 'bold')).pack(pady=(0, 20))
            
            # Información del pedido
            info_frame = ttk.LabelFrame(main_frame, text="Información General", padding=15)
            info_frame.pack(fill='x', pady=(0, 10))
            
            info_text = f"""
Proveedor: {pedido['proveedor'] if pedido.get('proveedor') else 'Sin proveedor'}
Fecha de Pedido: {datetime.fromisoformat(pedido['fecha']).strftime('%d/%m/%Y %H:%M')}
Fecha Entrega: {pedido['fecha_entrega'] if pedido.get('fecha_entrega') else 'No especificada'}
Estado: {pedido['estado'].upper()}
Usuario: {pedido['usuario_nombre']}
Notas: {pedido['notas'] if pedido['notas'] else 'Sin notas'}
            """
            
            ttk.Label(info_frame, text=info_text, justify='left', font=('Segoe UI', 10)).pack()
            
            # Detalles de productos
            productos_frame = ttk.LabelFrame(main_frame, text="Productos del Pedido", padding=15)
            productos_frame.pack(fill='both', expand=True, pady=(0, 10))
            
            detalles_tree = ttk.Treeview(productos_frame,
                                        columns=('Producto', 'Cantidad', 'Precio', 'Subtotal', 'Recibido'),
                                        show='headings', style='Modern.Treeview', height=10)
            
            detalles_tree.heading('Producto', text='Producto')
            detalles_tree.heading('Cantidad', text='Cantidad')
            detalles_tree.heading('Precio', text='Precio Unit.')
            detalles_tree.heading('Subtotal', text='Subtotal')
            detalles_tree.heading('Recibido', text='Recibido')
            
            detalles_tree.column('Producto', width=250)
            detalles_tree.column('Cantidad', width=100)
            detalles_tree.column('Precio', width=100)
            detalles_tree.column('Subtotal', width=100)
            detalles_tree.column('Recibido', width=80)
            
            for detalle in detalles:
                try:
                    recibido_text = "Sí" if detalle['recibido'] else "No"
                except (IndexError, KeyError):
                    recibido_text = "N/A"
                detalles_tree.insert('', 'end', values=(
                    detalle['producto_nombre'],
                    detalle['cantidad'],
                    format_currency(detalle['precio_unitario']),
                    format_currency(detalle['subtotal']),
                    recibido_text
                ))
            
            detalles_tree.pack(fill='both', expand=True)
            
            # Totales
            totales_frame = ttk.Frame(main_frame)
            totales_frame.pack(fill='x', pady=10)
            
            try:
                subtotal_val = pedido['subtotal']
            except (IndexError, KeyError):
                subtotal_val = pedido['total']
            try:
                impuestos_val = pedido['impuestos']
            except (IndexError, KeyError):
                impuestos_val = 0.0
            totales_text = f"""
Subtotal: {format_currency(subtotal_val)}
Impuestos: {format_currency(impuestos_val)}
TOTAL: {format_currency(pedido['total'])}
            """
            
            ttk.Label(totales_frame, text=totales_text, font=('Segoe UI', 12, 'bold'),
                     foreground=self.colors.get('success', '#27AE60')).pack(side='right', padx=20)
            
            # Botón cerrar
            ttk.Button(main_frame, text="Cerrar", command=details_window.destroy,
                      style='Primary.TButton').pack(pady=10)
            
        except Exception as e:
            logger.error(f"Error al mostrar detalles del pedido: {e}")
            messagebox.showerror("Error", f"No se pudieron cargar los detalles: {str(e)}")
    
    def _receive_pedido(self):
        """Recibe un pedido y actualiza el inventario"""
        selection = self.pedidos_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un pedido")
            return
        
        try:
            pedido_id = selection[0]
            
            # Verificar estado del pedido
            pedido = self.db.fetch_one("SELECT * FROM pedidos WHERE id = ?", (pedido_id,))
            
            if not pedido:
                messagebox.showerror("Error", "Pedido no encontrado")
                return
            
            if pedido['estado'] == 'recibido':
                messagebox.showinfo("Información", "Este pedido ya fue recibido")
                return
            
            if pedido['estado'] == 'cancelado':
                messagebox.showwarning("Advertencia", "No se puede recibir un pedido cancelado")
                return
            
            # Confirmar recepción
            if not messagebox.askyesno("Confirmar", 
                                       f"¿Confirma la recepción del pedido {pedido['numero_pedido']}?\n\n" +
                                       "Esto actualizará el inventario de todos los productos."):
                return
            
            # Obtener detalles del pedido
            detalles = self.db.fetch_all(
                "SELECT * FROM detalles_pedido WHERE pedido_id = ?",
                (pedido_id,)
            )
            
            # Actualizar inventario para cada producto
            for detalle in detalles:
                producto_id = detalle['producto_id']
                cantidad = detalle['cantidad']
                
                # Obtener stock actual
                producto = self.db.fetch_one(
                    "SELECT stock FROM productos WHERE id = ?",
                    (producto_id,)
                )
                
                if producto:
                    stock_anterior = producto['stock']
                    stock_nuevo = stock_anterior + cantidad
                    
                    # Actualizar stock
                    self.db.execute_query(
                        "UPDATE productos SET stock = ?, fecha_modificacion = ? WHERE id = ?",
                        (stock_nuevo, datetime.now(), producto_id)
                    )
                    
                    # Registrar movimiento de inventario
                    self.db.execute_query(
                        """INSERT INTO movimientos_inventario 
                           (producto_id, tipo, cantidad, stock_anterior, stock_nuevo, 
                            usuario_id, referencia, fecha, motivo)
                           VALUES (?, 'compra', ?, ?, ?, ?, ?, ?, ?)""",
                        (producto_id, cantidad, stock_anterior, stock_nuevo,
                         self.auth_service.current_user.id,
                         f"Pedido {pedido['numero_pedido']}", datetime.now(),
                         f"Recepción de pedido a proveedor")
                    )
            
            # Actualizar estado del pedido - SIN campos que no existen
            self.db.execute_query(
                """UPDATE pedidos SET estado = 'recibido' WHERE id = ?""",
                (pedido_id,)
            )
            
            messagebox.showinfo("Éxito", 
                               f"Pedido {pedido['numero_pedido']} recibido correctamente.\n" +
                               f"Se actualizó el inventario de {len(detalles)} productos.")
            
            # Recargar lista de pedidos
            self._load_pedidos(self.pedidos_tree)
            
            logger.info(f"Pedido {pedido['numero_pedido']} recibido por usuario {self.auth_service.current_user.username}")
            
        except Exception as e:
            logger.error(f"Error al recibir pedido: {e}")
            messagebox.showerror("Error", f"No se pudo recibir el pedido: {str(e)}")
    
    def _cancel_pedido(self):
        """Cancela un pedido"""
        selection = self.pedidos_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un pedido")
            return
        
        if messagebox.askyesno("Confirmar", "¿Está seguro de cancelar este pedido?"):
            try:
                pedido_id = selection[0]
                self.db.execute_query("UPDATE pedidos SET estado = 'cancelado' WHERE id = ?", (pedido_id,))
                messagebox.showinfo("Éxito", "Pedido cancelado correctamente")
                self._load_pedidos(self.pedidos_tree)
            except Exception as e:
                logger.error(f"Error al cancelar pedido: {e}")
                messagebox.showerror("Error", f"No se pudo cancelar el pedido: {str(e)}")
    
    def _edit_proveedor(self):
        """Edita un proveedor existente"""
        selection = self.proveedores_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un proveedor")
            return
        
        try:
            prov_id = selection[0]
            
            # Obtener datos actuales del proveedor
            proveedor = self.db.fetch_one("SELECT * FROM proveedores WHERE id = ?", (prov_id,))
            
            if not proveedor:
                messagebox.showerror("Error", "Proveedor no encontrado")
                return
            
            # Crear diálogo de edición
            dialog = tk.Toplevel(self.root)
            dialog.title("Editar Proveedor")
            dialog.geometry("500x600")
            dialog.transient(self.root)
            dialog.grab_set()
            
            form_frame = ttk.Frame(dialog, padding=20)
            form_frame.pack(fill='both', expand=True)
            
            ttk.Label(form_frame, text="Editar Proveedor",
                     font=('Segoe UI', 16, 'bold')).pack(pady=(0, 20))
            
            # Nombre
            ttk.Label(form_frame, text="Nombre: *").pack(anchor='w', pady=(5, 0))
            nombre_entry = ttk.Entry(form_frame, width=50, font=('Segoe UI', 10))
            nombre_entry.insert(0, proveedor['nombre'])
            nombre_entry.pack(fill='x', pady=(0, 10))
            
            # Contacto
            ttk.Label(form_frame, text="Persona de Contacto:").pack(anchor='w', pady=(5, 0))
            contacto_entry = ttk.Entry(form_frame, width=50, font=('Segoe UI', 10))
            contacto_entry.insert(0, proveedor['contacto'] if proveedor['contacto'] else '')
            contacto_entry.pack(fill='x', pady=(0, 10))
            
            # Teléfono
            ttk.Label(form_frame, text="Teléfono:").pack(anchor='w', pady=(5, 0))
            telefono_entry = ttk.Entry(form_frame, width=50, font=('Segoe UI', 10))
            telefono_entry.insert(0, proveedor['telefono'] if proveedor['telefono'] else '')
            telefono_entry.pack(fill='x', pady=(0, 10))
            
            # Email
            ttk.Label(form_frame, text="Email:").pack(anchor='w', pady=(5, 0))
            email_entry = ttk.Entry(form_frame, width=50, font=('Segoe UI', 10))
            email_entry.insert(0, proveedor['email'] if proveedor['email'] else '')
            email_entry.pack(fill='x', pady=(0, 10))
            
            # Dirección
            ttk.Label(form_frame, text="Dirección:").pack(anchor='w', pady=(5, 0))
            direccion_entry = ttk.Entry(form_frame, width=50, font=('Segoe UI', 10))
            direccion_entry.insert(0, proveedor['direccion'] if proveedor['direccion'] else '')
            direccion_entry.pack(fill='x', pady=(0, 10))
            
            # Notas
            ttk.Label(form_frame, text="Notas:").pack(anchor='w', pady=(5, 0))
            notas_text = tk.Text(form_frame, width=50, height=4, font=('Segoe UI', 10))
            notas_text.insert('1.0', proveedor['notas'] if proveedor['notas'] else '')
            notas_text.pack(fill='x', pady=(0, 10))
            
            def guardar():
                nombre = nombre_entry.get().strip()
                
                if not nombre:
                    messagebox.showerror("Error", "El nombre es obligatorio")
                    return
                
                # Validar email si se proporciona
                email = email_entry.get().strip()
                if email:
                    import re
                    if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
                        messagebox.showerror("Error", "El formato del email no es válido")
                        return
                
                # Validar teléfono si se proporciona
                telefono = telefono_entry.get().strip()
                if telefono:
                    if not re.match(r'^[\d\s\-\(\)\+]{8,20}$', telefono):
                        messagebox.showerror("Error", "El formato del teléfono no es válido")
                        return
                
                try:
                    self.db.execute_query(
                        """UPDATE proveedores SET nombre = ?, contacto = ?, telefono = ?,
                           email = ?, direccion = ?, notas = ?
                           WHERE id = ?""",
                        (nombre, contacto_entry.get().strip(), telefono,
                         email, direccion_entry.get().strip(),
                         notas_text.get('1.0', 'end-1c').strip(), prov_id)
                    )
                    
                    messagebox.showinfo("Éxito", "Proveedor actualizado correctamente")
                    self._load_proveedores(self.proveedores_tree)
                    dialog.destroy()
                    
                except Exception as e:
                    logger.error(f"Error al actualizar proveedor: {e}")
                    messagebox.showerror("Error", f"No se pudo actualizar: {str(e)}")
            
            btn_frame = ttk.Frame(form_frame)
            btn_frame.pack(pady=20)
            
            ttk.Button(btn_frame, text="💾 Guardar", command=guardar,
                      style='Success.TButton').pack(side='left', padx=5)
            ttk.Button(btn_frame, text="Cancelar", command=dialog.destroy,
                      style='Danger.TButton').pack(side='left', padx=5)
            
        except Exception as e:
            logger.error(f"Error al editar proveedor: {e}")
            messagebox.showerror("Error", f"No se pudo editar el proveedor: {str(e)}")
    
    def _delete_proveedor(self):
        """Elimina un proveedor"""
        selection = self.proveedores_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione un proveedor")
            return
        
        if messagebox.askyesno("Confirmar", "¿Está seguro de eliminar este proveedor?"):
            try:
                prov_id = selection[0]
                self.db.execute_query("UPDATE proveedores SET activo = 0 WHERE id = ?", (prov_id,))
                messagebox.showinfo("Éxito", "Proveedor desactivado correctamente")
                self._load_proveedores(self.proveedores_tree)
            except Exception as e:
                logger.error(f"Error al desactivar proveedor: {e}")
                messagebox.showerror("Error", f"No se pudo desactivar el proveedor: {str(e)}")
    
    def show_reports(self):
        """Muestra los reportes"""
        self.clear_content()
        self.current_screen = "reports"
        
        # Configurar grid
        self.content_frame.grid_columnconfigure(0, weight=1)
        
        # Título
        title = ttk.Label(self.content_frame,
                         text="Reportes y Estadísticas",
                         font=(self.config.get('fonts', {}).get('family', 'Segoe UI'), 20, 'bold'),
                         background=self.colors.get('bg_light', '#ECF0F1'),
                         foreground=self.colors.get('primary', '#2C3E50'))
        title.pack(pady=20)
        
        # Frame de reportes
        reports_frame = ttk.Frame(self.content_frame, style='Main.TFrame')
        reports_frame.pack(fill='both', expand=True, padx=50, pady=20)
        
        # Reporte de ventas
        sales_card = ttk.LabelFrame(reports_frame,
                                    text="Reporte de Ventas",
                                    style='Card.TLabelframe',
                                    padding=20)
        sales_card.pack(fill='x', pady=10)
        
        ttk.Label(sales_card,
                 text="Genere reportes de ventas por período",
                 style='Info.TLabel').pack(pady=5)
        
        sales_btn_frame = ttk.Frame(sales_card, style='Card.TFrame')
        sales_btn_frame.pack(pady=10)
        
        ttk.Button(sales_btn_frame,
                  text="📊 Ventas del Día",
                  command=lambda: self._generate_report('ventas_dia'),
                  style='Primary.TButton').pack(side='left', padx=5)
        
        ttk.Button(sales_btn_frame,
                  text="📊 Ventas del Mes",
                  command=lambda: self._generate_report('ventas_mes'),
                  style='Primary.TButton').pack(side='left', padx=5)
        
        ttk.Button(sales_btn_frame,
                  text="📊 Ventas Personalizadas",
                  command=lambda: self._custom_sales_report(),
                  style='Primary.TButton').pack(side='left', padx=5)
        
        # Reporte de inventario
        inventory_card = ttk.LabelFrame(reports_frame,
                                       text="Reporte de Inventario",
                                       style='Card.TLabelframe',
                                       padding=20)
        inventory_card.pack(fill='x', pady=10)
        
        ttk.Label(inventory_card,
                 text="Consulte el estado actual del inventario",
                 style='Info.TLabel').pack(pady=5)
        
        inventory_btn_frame = ttk.Frame(inventory_card, style='Card.TFrame')
        inventory_btn_frame.pack(pady=10)
        
        ttk.Button(inventory_btn_frame,
                  text="📦 Inventario Completo",
                  command=lambda: self._generate_report('inventario_completo'),
                  style='Success.TButton').pack(side='left', padx=5)
        
        ttk.Button(inventory_btn_frame,
                  text="⚠️ Productos Bajo Stock",
                  command=lambda: self._generate_report('bajo_stock'),
                  style='Danger.TButton').pack(side='left', padx=5)
        
        # Reporte financiero
        financial_card = ttk.LabelFrame(reports_frame,
                                       text="Reporte Financiero",
                                       style='Card.TLabelframe',
                                       padding=20)
        financial_card.pack(fill='x', pady=10)
        
        ttk.Label(financial_card,
                 text="Análisis financiero y rentabilidad",
                 style='Info.TLabel').pack(pady=5)
        
        financial_btn_frame = ttk.Frame(financial_card, style='Card.TFrame')
        financial_btn_frame.pack(pady=10)
        
        ttk.Button(financial_btn_frame,
                  text="💰 Resumen Financiero",
                  command=lambda: self._generate_report('financiero'),
                  style='Success.TButton').pack(side='left', padx=5)
        
        # 💳 Reporte de métodos de pago
        payment_card = ttk.LabelFrame(reports_frame,
                                     text="Reporte de Métodos de Pago",
                                     style='Card.TLabelframe',
                                     padding=20)
        payment_card.pack(fill='x', pady=10)
        
        ttk.Label(payment_card,
                 text="Análisis de ventas por método de pago",
                 style='Info.TLabel').pack(pady=5)
        
        payment_btn_frame = ttk.Frame(payment_card, style='Card.TFrame')
        payment_btn_frame.pack(pady=10)
        
        ttk.Button(payment_btn_frame,
                  text="💳 Ventas por Método de Pago",
                  command=lambda: self._generate_report('metodos_pago'),
                  style='Primary.TButton').pack(side='left', padx=5)
        
        ttk.Button(payment_btn_frame,
                  text="📊 Gráfico de Métodos de Pago",
                  command=lambda: self._show_payment_methods_chart(),
                  style='Success.TButton').pack(side='left', padx=5)
    
    def _generate_report(self, report_type):
        """Genera un reporte específico"""
        try:
            if report_type == 'ventas_dia':
                self._report_ventas_dia()
            elif report_type == 'ventas_mes':
                self._report_ventas_mes()
            elif report_type == 'inventario_completo':
                self._report_inventario()
            elif report_type == 'bajo_stock':
                self._report_bajo_stock()
            elif report_type == 'financiero':
                self._report_financiero()
            elif report_type == 'metodos_pago':
                self._report_metodos_pago()
        except Exception as e:
            logger.error(f"Error al generar reporte: {e}")
            messagebox.showerror("Error", f"Error al generar reporte: {str(e)}")
    
    def _report_ventas_dia(self):
        """Reporte de ventas del día"""
        hoy = datetime.now().date()
        rows = self.db.fetch_all(
            """SELECT v.*, u.nombre as usuario_nombre 
               FROM ventas v 
               LEFT JOIN usuarios u ON v.usuario_id = u.id
               WHERE DATE(v.fecha) = ? AND v.estado = 'completada'
               ORDER BY v.fecha DESC""",
            (hoy,)
        )
        
        if not rows:
            messagebox.showinfo("Información", "No hay ventas registradas hoy")
            return
        
        total = sum(row['total'] for row in rows)
        
        report = f"REPORTE DE VENTAS DEL DÍA\n"
        report += f"Fecha: {hoy.strftime('%d/%m/%Y')}\n"
        report += f"=" * 50 + "\n\n"
        report += f"Total de ventas: {len(rows)}\n"
        report += f"Monto total: {format_currency(total)}\n\n"
        report += "Detalle de ventas:\n"
        report += "-" * 50 + "\n"
        
        for row in rows:
            fecha = datetime.fromisoformat(row['fecha']).strftime('%H:%M:%S')
            report += f"{row['numero_venta']} - {fecha} - {format_currency(row['total'])} - {row['usuario_nombre']}\n"
        
        self._show_report_window("Reporte de Ventas del Día", report)
    
    def _report_ventas_mes(self):
        """Reporte de ventas del mes"""
        hoy = datetime.now()
        inicio_mes = datetime(hoy.year, hoy.month, 1)
        
        rows = self.db.fetch_all(
            """SELECT v.* 
               FROM ventas v 
               WHERE v.fecha >= ? AND v.estado = 'completada'
               ORDER BY v.fecha DESC""",
            (inicio_mes,)
        )
        
        if not rows:
            messagebox.showinfo("Información", "No hay ventas registradas este mes")
            return
        
        total = sum(row['total'] for row in rows)
        
        report = f"REPORTE DE VENTAS DEL MES\n"
        report += f"Período: {inicio_mes.strftime('%m/%Y')}\n"
        report += f"=" * 50 + "\n\n"
        report += f"Total de ventas: {len(rows)}\n"
        report += f"Monto total: {format_currency(total)}\n"
        report += f"Promedio por venta: {format_currency(total / len(rows))}\n"
        
        self._show_report_window("Reporte de Ventas del Mes", report)
    
    def _report_inventario(self):
        """Reporte de inventario completo"""
        productos = self.producto_service.listar_productos(solo_activos=True)
        
        if not productos:
            messagebox.showinfo("Información", "No hay productos registrados")
            return
        
        valor_total = sum(p.stock * p.costo for p in productos)
        
        report = f"REPORTE DE INVENTARIO\n"
        report += f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
        report += f"=" * 80 + "\n\n"
        report += f"Total de productos: {len(productos)}\n"
        report += f"Valor total del inventario: {format_currency(valor_total)}\n\n"
        report += f"{'Producto':<30} {'Stock':>10} {'Costo':>12} {'Valor':>12}\n"
        report += "-" * 80 + "\n"
        
        for p in productos:
            valor = p.stock * p.costo
            report += f"{p.nombre[:28]:<30} {p.stock:>10} {format_currency(p.costo):>12} {format_currency(valor):>12}\n"
        
        self._show_report_window("Reporte de Inventario", report)
    
    def _report_bajo_stock(self):
        """Reporte de productos con stock bajo"""
        productos = self.producto_service.productos_bajo_stock()
        
        if not productos:
            messagebox.showinfo("Información", "No hay productos bajo stock")
            return
        
        report = f"REPORTE DE PRODUCTOS BAJO STOCK\n"
        report += f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
        report += f"=" * 80 + "\n\n"
        report += f"⚠️  Productos que requieren reabastecimiento: {len(productos)}\n\n"
        report += f"{'Producto':<30} {'Stock':>10} {'Mínimo':>10} {'Faltante':>10}\n"
        report += "-" * 80 + "\n"
        
        for p in productos:
            faltante = p.stock_minimo - p.stock
            report += f"{p.nombre[:28]:<30} {p.stock:>10} {p.stock_minimo:>10} {faltante:>10}\n"
        
        self._show_report_window("Productos Bajo Stock", report)
    
    def _report_financiero(self):
        """Reporte financiero general"""
        hoy = datetime.now()
        inicio_mes = datetime(hoy.year, hoy.month, 1)
        
        # Ventas del mes
        ventas_rows = self.db.fetch_all(
            "SELECT SUM(total) as total, COUNT(*) as cantidad FROM ventas WHERE fecha >= ? AND estado = 'completada'",
            (inicio_mes,)
        )
        
        ventas_total = ventas_rows[0]['total'] if ventas_rows and ventas_rows[0]['total'] else 0
        ventas_cantidad = ventas_rows[0]['cantidad'] if ventas_rows else 0
        
        # Valor del inventario
        productos = self.producto_service.listar_productos(solo_activos=True)
        valor_inventario = sum(p.stock * p.costo for p in productos)
        
        report = f"REPORTE FINANCIERO\n"
        report += f"Período: {inicio_mes.strftime('%m/%Y')}\n"
        report += f"=" * 50 + "\n\n"
        report += f"VENTAS DEL MES:\n"
        report += f"  Total vendido: {format_currency(ventas_total)}\n"
        report += f"  Número de ventas: {ventas_cantidad}\n"
        report += f"  Ticket promedio: {format_currency(ventas_total / ventas_cantidad if ventas_cantidad > 0 else 0)}\n\n"
        report += f"INVENTARIO:\n"
        report += f"  Productos activos: {len(productos)}\n"
        report += f"  Valor del inventario: {format_currency(valor_inventario)}\n"
        
        self._show_report_window("Reporte Financiero", report)
    
    def _report_metodos_pago(self):
        """Reporte de ventas por método de pago"""
        # Obtener ventas agrupadas por método de pago
        rows = self.db.fetch_all("""
            SELECT 
                metodo_pago,
                COUNT(*) as cantidad,
                SUM(total) as total,
                AVG(total) as promedio
            FROM ventas 
            WHERE estado = 'completada'
            GROUP BY metodo_pago
            ORDER BY total DESC
        """)
        
        if not rows:
            messagebox.showinfo("Información", "No hay ventas registradas")
            return
        
        # Calcular total general
        total_general = sum(row['total'] for row in rows)
        total_ventas = sum(row['cantidad'] for row in rows)
        
        # Diccionario de nombres amigables
        metodo_nombres = {
            'efectivo': 'Efectivo',
            'tarjeta_credito': 'Tarjeta de Crédito',
            'tarjeta_debito': 'Tarjeta de Débito',
            'transferencia': 'Transferencia',
            'otro': 'Otro'
        }
        
        report = f"REPORTE DE VENTAS POR MÉTODO DE PAGO\n"
        report += f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
        report += f"=" * 70 + "\n\n"
        report += f"RESUMEN GENERAL:\n"
        report += f"  Total de ventas: {total_ventas}\n"
        report += f"  Monto total: {format_currency(total_general)}\n\n"
        report += f"DESGLOSE POR MÉTODO DE PAGO:\n"
        report += f"{'-' * 70}\n"
        report += f"{'Método':<25} {'Cantidad':>10} {'Total':>15} {'Promedio':>15}\n"
        report += f"{'-' * 70}\n"
        
        for row in rows:
            metodo = metodo_nombres.get(row['metodo_pago'], row['metodo_pago'])
            porcentaje = (row['total'] / total_general * 100) if total_general > 0 else 0
            report += f"{metodo:<25} {row['cantidad']:>10} {format_currency(row['total']):>15} {format_currency(row['promedio']):>15}\n"
            report += f"{'':>25} ({porcentaje:>6.2f}% del total)\n"
        
        report += f"{'-' * 70}\n"
        report += f"{'TOTAL':<25} {total_ventas:>10} {format_currency(total_general):>15}\n"
        
        self._show_report_window("Reporte de Métodos de Pago", report)
    
    def _show_payment_methods_chart(self):
        """Muestra gráfico de ventas por método de pago"""
        try:
            # Obtener datos
            rows = self.db.fetch_all("""
                SELECT 
                    metodo_pago,
                    COUNT(*) as cantidad,
                    SUM(total) as total
                FROM ventas 
                WHERE estado = 'completada'
                GROUP BY metodo_pago
                ORDER BY total DESC
            """)
            
            if not rows:
                messagebox.showinfo("Información", "No hay datos para mostrar")
                return
            
            # Diccionario de nombres amigables
            metodo_nombres = {
                'efectivo': 'Efectivo',
                'tarjeta_credito': 'Tarjeta de Crédito',
                'tarjeta_debito': 'Tarjeta de Débito',
                'transferencia': 'Transferencia',
                'otro': 'Otro'
            }
            
            # Preparar datos para el gráfico
            metodos = [metodo_nombres.get(r['metodo_pago'], r['metodo_pago']) for r in rows]
            totales = [r['total'] for r in rows]
            cantidades = [r['cantidad'] for r in rows]
            
            # Crear ventana para el gráfico
            chart_window = tk.Toplevel(self.root)
            chart_window.title("Gráfico de Métodos de Pago")
            chart_window.geometry("1000x600")
            chart_window.transient(self.root)
            
            main_frame = ttk.Frame(chart_window, padding=20)
            main_frame.pack(fill='both', expand=True)
            
            # Título
            ttk.Label(main_frame,
                     text="Ventas por Método de Pago",
                     font=('Segoe UI', 16, 'bold')).pack(pady=(0, 20))
            
            # Crear figura con 2 subgráficos
            fig = Figure(figsize=(10, 5), dpi=100)
            
            # Gráfico de barras - Montos
            ax1 = fig.add_subplot(121)
            colors = ['#3498db', '#2ecc71', '#e74c3c', '#f39c12', '#9b59b6']
            bars1 = ax1.bar(metodos, totales, color=colors[:len(metodos)])
            ax1.set_title('Ventas por Monto', fontsize=12, fontweight='bold')
            ax1.set_ylabel('Monto ($)', fontsize=10)
            ax1.tick_params(axis='x', rotation=45)
            
            # Agregar valores en las barras
            for bar in bars1:
                height = bar.get_height()
                ax1.text(bar.get_x() + bar.get_width()/2., height,
                        f'${height:,.0f}',
                        ha='center', va='bottom', fontsize=9)
            
            # Gráfico circular - Cantidad de transacciones
            ax2 = fig.add_subplot(122)
            ax2.pie(cantidades, labels=metodos, autopct='%1.1f%%',
                   colors=colors[:len(metodos)], startangle=90)
            ax2.set_title('Distribución de Transacciones', fontsize=12, fontweight='bold')
            
            fig.tight_layout()
            
            # Agregar canvas
            canvas = FigureCanvasTkAgg(fig, main_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)
            
            # Botón cerrar
            ttk.Button(main_frame,
                      text="Cerrar",
                      command=chart_window.destroy,
                      style='Primary.TButton').pack(pady=10)
            
        except Exception as e:
            logger.error(f"Error al mostrar gráfico: {e}")
            messagebox.showerror("Error", f"No se pudo generar el gráfico: {str(e)}")
    
    def _custom_sales_report(self):
        """Reporte de ventas personalizado con filtros avanzados"""
        # Crear ventana de diálogo
        dialog = tk.Toplevel(self.root)
        dialog.title("Reporte Personalizado de Ventas")
        dialog.geometry("600x650")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Frame principal
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        # Título
        ttk.Label(main_frame,
                 text="Configurar Reporte Personalizado",
                 font=('Segoe UI', 14, 'bold')).pack(pady=(0, 20))
        
        # Frame de filtros
        filters_frame = ttk.LabelFrame(main_frame, text="Filtros de Búsqueda", padding=15)
        filters_frame.pack(fill='x', pady=(0, 20))
        
        # Rango de fechas
        date_frame = ttk.Frame(filters_frame)
        date_frame.pack(fill='x', pady=5)
        
        ttk.Label(date_frame, text="Desde:", width=12).pack(side='left', padx=5)
        from_date_entry = ttk.Entry(date_frame, width=15)
        from_date_entry.insert(0, (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        from_date_entry.pack(side='left', padx=5)
        
        ttk.Label(date_frame, text="Hasta:", width=8).pack(side='left', padx=5)
        to_date_entry = ttk.Entry(date_frame, width=15)
        to_date_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        to_date_entry.pack(side='left', padx=5)
        
        # Filtro de categoría
        category_frame = ttk.Frame(filters_frame)
        category_frame.pack(fill='x', pady=5)
        
        ttk.Label(category_frame, text="Categoría:", width=12).pack(side='left', padx=5)
        category_var = tk.StringVar(value="Todas")
        
        # Obtener categorías disponibles
        try:
            categories = ['Todas']
            productos = self.producto_service.listar_productos(solo_activos=False)
            unique_cats = set(p.categoria for p in productos if p.categoria)
            categories.extend(sorted(unique_cats))
        except:
            categories = ['Todas']
        
        category_combo = ttk.Combobox(category_frame, textvariable=category_var,
                                     values=categories, width=25, state='readonly')
        category_combo.pack(side='left', padx=5)
        
        # Filtro de usuario
        user_frame = ttk.Frame(filters_frame)
        user_frame.pack(fill='x', pady=5)
        
        ttk.Label(user_frame, text="Usuario:", width=12).pack(side='left', padx=5)
        user_var = tk.StringVar(value="Todos")
        
        # Obtener usuarios disponibles
        try:
            users = ['Todos']
            users_data = self.db.fetch_all("SELECT DISTINCT nombre FROM usuarios ORDER BY nombre")
            users.extend([u['nombre'] for u in users_data])
        except:
            users = ['Todos']
        
        user_combo = ttk.Combobox(user_frame, textvariable=user_var,
                                 values=users, width=25, state='readonly')
        user_combo.pack(side='left', padx=5)
        
        # Filtro de método de pago
        payment_frame = ttk.Frame(filters_frame)
        payment_frame.pack(fill='x', pady=5)
        
        ttk.Label(payment_frame, text="Método Pago:", width=12).pack(side='left', padx=5)
        payment_var = tk.StringVar(value="Todos")
        payment_combo = ttk.Combobox(payment_frame, textvariable=payment_var,
                                    values=['Todos', 'Efectivo', 'Tarjeta', 'Transferencia'],
                                    width=25, state='readonly')
        payment_combo.pack(side='left', padx=5)
        
        # Filtro de estado
        status_frame = ttk.Frame(filters_frame)
        status_frame.pack(fill='x', pady=5)
        
        ttk.Label(status_frame, text="Estado:", width=12).pack(side='left', padx=5)
        status_var = tk.StringVar(value="Todos")
        status_combo = ttk.Combobox(status_frame, textvariable=status_var,
                                   values=['Todos', 'Completada', 'Cancelada'],
                                   width=25, state='readonly')
        status_combo.pack(side='left', padx=5)
        
        # Opciones de agrupación
        grouping_frame = ttk.LabelFrame(main_frame, text="Agrupación y Análisis", padding=15)
        grouping_frame.pack(fill='x', pady=(0, 20))
        
        group_by_var = tk.StringVar(value="dia")
        ttk.Radiobutton(grouping_frame, text="Por día", variable=group_by_var,
                       value="dia").pack(anchor='w', pady=2)
        ttk.Radiobutton(grouping_frame, text="Por semana", variable=group_by_var,
                       value="semana").pack(anchor='w', pady=2)
        ttk.Radiobutton(grouping_frame, text="Por mes", variable=group_by_var,
                       value="mes").pack(anchor='w', pady=2)
        ttk.Radiobutton(grouping_frame, text="Sin agrupación (detallado)", variable=group_by_var,
                       value="detalle").pack(anchor='w', pady=2)
        
        # Función para generar el reporte
        def generate_report():
            try:
                from_date = from_date_entry.get()
                to_date = to_date_entry.get()
                category = category_var.get()
                user = user_var.get()
                payment = payment_var.get()
                status = status_var.get()
                group_by = group_by_var.get()
                
                # Construir query base
                query = """
                    SELECT v.*, u.nombre as usuario_nombre, dv.producto_id, dv.cantidad,
                           dv.precio_unitario, dv.subtotal, p.nombre as producto_nombre,
                           p.categoria as producto_categoria
                    FROM ventas v
                    LEFT JOIN usuarios u ON v.usuario_id = u.id
                    LEFT JOIN detalles_venta dv ON v.id = dv.venta_id
                    LEFT JOIN productos p ON dv.producto_id = p.id
                    WHERE DATE(v.fecha) BETWEEN ? AND ?
                """
                params = [from_date, to_date]
                
                # Aplicar filtros
                if user != "Todos":
                    query += " AND u.nombre = ?"
                    params.append(user)
                
                if payment != "Todos":
                    query += " AND v.metodo_pago = ?"
                    params.append(payment.lower())
                
                if status != "Todos":
                    query += " AND v.estado = ?"
                    params.append(status.lower())
                
                if category != "Todas":
                    query += " AND p.categoria = ?"
                    params.append(category)
                
                query += " ORDER BY v.fecha DESC"
                
                # Ejecutar query
                results = self.db.fetch_all(query, tuple(params))
                
                # Generar contenido del reporte
                content = self._generate_custom_report_content(
                    results, from_date, to_date, category, user, payment, status, group_by
                )
                
                # Cerrar diálogo y mostrar reporte
                dialog.destroy()
                self._show_report_window(f"Reporte Personalizado ({from_date} a {to_date})", content)
                
            except Exception as e:
                logger.error(f"Error al generar reporte personalizado: {e}")
                messagebox.showerror("Error", f"Error al generar reporte: {str(e)}")
        
        # Botones
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)
        
        ttk.Button(btn_frame,
                  text="📊 Generar Reporte",
                  command=generate_report,
                  style='Success.TButton').pack(side='left', padx=5)
        
        ttk.Button(btn_frame,
                  text="Cancelar",
                  command=dialog.destroy,
                  style='Secondary.TButton').pack(side='left', padx=5)
    
    def _generate_custom_report_content(self, results, from_date, to_date, category, user, payment, status, group_by):
        """Genera el contenido del reporte personalizado"""
        content = "=" * 80 + "\n"
        content += "REPORTE PERSONALIZADO DE VENTAS\n"
        content += "=" * 80 + "\n\n"
        
        # Información de filtros aplicados
        content += f"Período: {from_date} a {to_date}\n"
        content += f"Categoría: {category}\n"
        content += f"Usuario: {user}\n"
        content += f"Método de Pago: {payment}\n"
        content += f"Estado: {status}\n"
        content += f"Agrupación: {group_by.title()}\n"
        content += f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        content += "=" * 80 + "\n\n"
        
        if not results:
            content += "No se encontraron ventas con los filtros especificados.\n"
            return content
        
        # Agrupar datos
        if group_by == "detalle":
            # Mostrar todas las ventas con detalles
            ventas_dict = {}
            for row in results:
                venta_id = row['id']
                if venta_id not in ventas_dict:
                    ventas_dict[venta_id] = {
                        'numero': row['numero_venta'],
                        'fecha': row['fecha'],
                        'usuario': row['usuario_nombre'] or 'N/A',
                        'metodo': row['metodo_pago'].title(),
                        'total': row['total'],
                        'productos': []
                    }
                if row['producto_nombre']:
                    ventas_dict[venta_id]['productos'].append({
                        'nombre': row['producto_nombre'],
                        'cantidad': row['cantidad'],
                        'precio': row['precio_unitario'],
                        'subtotal': row['subtotal']
                    })
            
            for venta in ventas_dict.values():
                content += f"\nVenta #{venta['numero']} - {venta['fecha'][:16]}\n"
                content += f"Usuario: {venta['usuario']} | Método: {venta['metodo']}\n"
                content += "-" * 80 + "\n"
                for prod in venta['productos']:
                    content += f"  {prod['nombre']:<40} x{prod['cantidad']:<3} @ {format_currency(prod['precio']):<10} = {format_currency(prod['subtotal'])}\n"
                content += "-" * 80 + "\n"
                content += f"TOTAL: {format_currency(venta['total'])}\n"
                content += "\n"
        
        else:
            # Agrupar por período
            grouped_data = {}
            for row in results:
                fecha = datetime.fromisoformat(row['fecha'])
                
                if group_by == "dia":
                    key = fecha.strftime('%Y-%m-%d')
                elif group_by == "semana":
                    key = f"{fecha.year}-S{fecha.isocalendar()[1]}"
                else:  # mes
                    key = fecha.strftime('%Y-%m')
                
                if key not in grouped_data:
                    grouped_data[key] = {'total': 0, 'cantidad_ventas': 0, 'productos_vendidos': 0}
                
                # Solo contar la venta una vez por venta_id
                if row['id'] not in [v.get('id') for v in grouped_data[key].get('ventas', [])]:
                    if 'ventas' not in grouped_data[key]:
                        grouped_data[key]['ventas'] = []
                    grouped_data[key]['ventas'].append({'id': row['id'], 'total': row['total']})
                    grouped_data[key]['total'] += row['total']
                    grouped_data[key]['cantidad_ventas'] += 1
                
                if row['cantidad']:
                    grouped_data[key]['productos_vendidos'] += row['cantidad']
            
            content += f"{'Período':<20} {'Ventas':<10} {'Productos':<12} {'Total':<15}\n"
            content += "=" * 80 + "\n"
            
            total_general = 0
            total_ventas = 0
            total_productos = 0
            
            for periodo in sorted(grouped_data.keys()):
                data = grouped_data[periodo]
                content += f"{periodo:<20} {data['cantidad_ventas']:<10} {data['productos_vendidos']:<12} {format_currency(data['total']):<15}\n"
                total_general += data['total']
                total_ventas += data['cantidad_ventas']
                total_productos += data['productos_vendidos']
            
            content += "=" * 80 + "\n"
            content += f"{'TOTALES':<20} {total_ventas:<10} {total_productos:<12} {format_currency(total_general):<15}\n"
            content += f"\nPromedio por venta: {format_currency(total_general / total_ventas if total_ventas > 0 else 0)}\n"
        
        return content
    
    def _show_report_window(self, title, content):
        """Muestra una ventana con el reporte"""
        window = tk.Toplevel(self.root)
        window.title(title)
        window.geometry("800x600")
        window.transient(self.root)
        
        # Frame principal
        main_frame = ttk.Frame(window, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        # Texto del reporte
        text_widget = tk.Text(main_frame,
                             font=('Courier New', 10),
                             wrap='none',
                             bg='white',
                             fg='#2C3E50')
        text_widget.pack(fill='both', expand=True)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(text_widget, orient='vertical', command=text_widget.yview)
        v_scrollbar.pack(side='right', fill='y')
        text_widget.configure(yscrollcommand=v_scrollbar.set)
        
        h_scrollbar = ttk.Scrollbar(main_frame, orient='horizontal', command=text_widget.xview)
        h_scrollbar.pack(fill='x')
        text_widget.configure(xscrollcommand=h_scrollbar.set)
        
        # Insertar contenido
        text_widget.insert('1.0', content)
        text_widget.config(state='disabled')
        
        # Botones
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)
        
        ttk.Button(btn_frame,
                  text="💾 Guardar",
                  command=lambda: self._save_report(content, title),
                  style='Success.TButton').pack(side='left', padx=5)
        
        ttk.Button(btn_frame,
                  text="Cerrar",
                  command=window.destroy,
                  style='Primary.TButton').pack(side='left', padx=5)
    
    def _save_report(self, content, title):
        """Guarda el reporte en archivo con múltiples formatos"""
        # Diálogo para seleccionar formato
        format_dialog = tk.Toplevel(self.root)
        format_dialog.title("Exportar Reporte")
        format_dialog.geometry("400x250")
        format_dialog.transient(self.root)
        format_dialog.grab_set()
        
        main_frame = ttk.Frame(format_dialog, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        ttk.Label(main_frame, text="Seleccione el formato de exportación",
                 font=('Segoe UI', 12, 'bold')).pack(pady=(0, 20))
        
        format_var = tk.StringVar(value="txt")
        
        ttk.Radiobutton(main_frame, text="📄 Archivo de Texto (.txt)",
                       variable=format_var, value="txt").pack(anchor='w', pady=5)
        ttk.Radiobutton(main_frame, text="📕 Documento PDF (.pdf)",
                       variable=format_var, value="pdf").pack(anchor='w', pady=5)
        ttk.Radiobutton(main_frame, text="📊 Hoja de Excel (.xlsx)",
                       variable=format_var, value="xlsx").pack(anchor='w', pady=5)
        
        def export_report():
            formato = format_var.get()
            
            if formato == "txt":
                self._export_report_txt(content, title)
            elif formato == "pdf":
                self._export_report_pdf(content, title)
            elif formato == "xlsx":
                self._export_report_xlsx(content, title)
            
            format_dialog.destroy()
        
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=20)
        
        ttk.Button(btn_frame, text="💾 Exportar", command=export_report,
                  style='Success.TButton').pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=format_dialog.destroy,
                  style='Danger.TButton').pack(side='left', padx=5)
    
    def _export_report_txt(self, content, title):
        """Exporta reporte como archivo de texto"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Archivos de texto", "*.txt"), ("Todos los archivos", "*.*")],
            initialfile=f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(content)
                messagebox.showinfo("Éxito", f"Reporte guardado en:\n{filename}")
                logger.info(f"Reporte TXT guardado: {filename}")
            except Exception as e:
                logger.error(f"Error al guardar reporte TXT: {e}")
                messagebox.showerror("Error", f"Error al guardar: {str(e)}")
    
    def _export_report_pdf(self, content, title):
        """Exporta reporte como PDF"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*")],
            initialfile=f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
        
        if filename:
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Preformatted
                from reportlab.lib.enums import TA_CENTER, TA_LEFT
                
                # Crear documento PDF
                doc = SimpleDocTemplate(filename, pagesize=letter,
                                       topMargin=0.5*inch, bottomMargin=0.5*inch,
                                       leftMargin=0.5*inch, rightMargin=0.5*inch)
                
                story = []
                styles = getSampleStyleSheet()
                
                # Estilo personalizado para título
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    textColor='#2C3E50',
                    spaceAfter=20,
                    alignment=TA_CENTER
                )
                
                # Estilo para contenido
                content_style = ParagraphStyle(
                    'ContentStyle',
                    parent=styles['Code'],
                    fontSize=9,
                    fontName='Courier',
                    alignment=TA_LEFT
                )
                
                # Agregar título
                story.append(Paragraph(title, title_style))
                story.append(Spacer(1, 0.2*inch))
                
                # Agregar contenido
                story.append(Preformatted(content, content_style))
                
                # Generar PDF
                doc.build(story)
                
                messagebox.showinfo("Éxito", f"Reporte PDF guardado en:\n{filename}")
                logger.info(f"Reporte PDF guardado: {filename}")
                
            except ImportError:
                messagebox.showerror("Error", 
                                   "La librería reportlab no está instalada.\n" +
                                   "Instálela con: pip install reportlab")
            except Exception as e:
                logger.error(f"Error al guardar reporte PDF: {e}")
                messagebox.showerror("Error", f"Error al guardar PDF: {str(e)}")
    
    def _export_report_xlsx(self, content, title):
        """Exporta reporte como Excel"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
            initialfile=f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if filename:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                
                # Crear workbook
                wb = Workbook()
                ws = wb.active
                ws.title = title[:31]  # Excel limita a 31 caracteres
                
                # Estilo para título
                title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
                title_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
                title_alignment = Alignment(horizontal='center', vertical='center')
                
                # Agregar título
                ws['A1'] = title
                ws['A1'].font = title_font
                ws['A1'].fill = title_fill
                ws['A1'].alignment = title_alignment
                ws.merge_cells('A1:E1')
                ws.row_dimensions[1].height = 30
                
                # Agregar contenido línea por línea
                lines = content.split('\n')
                for idx, line in enumerate(lines, start=3):
                    ws[f'A{idx}'] = line
                    ws[f'A{idx}'].font = Font(name='Courier New', size=9)
                
                # Ajustar ancho de columna
                ws.column_dimensions['A'].width = 100
                
                # Guardar
                wb.save(filename)
                
                messagebox.showinfo("Éxito", f"Reporte Excel guardado en:\n{filename}")
                logger.info(f"Reporte Excel guardado: {filename}")
                
            except ImportError:
                messagebox.showerror("Error",
                                   "La librería openpyxl no está instalada.\n" +
                                   "Instálela con: pip install openpyxl")
            except Exception as e:
                logger.error(f"Error al guardar reporte Excel: {e}")
                messagebox.showerror("Error", f"Error al guardar Excel: {str(e)}")
    
    def show_settings(self):
        """Muestra la configuración"""
        self.clear_content()
        self.current_screen = "settings"
        
        # Título
        title = ttk.Label(self.content_frame,
                         text="Configuración del Sistema",
                         font=(self.config.get('fonts', {}).get('family', 'Segoe UI'), 20, 'bold'),
                         background=self.colors.get('bg_light', '#ECF0F1'),
                         foreground=self.colors.get('primary', '#2C3E50'))
        title.pack(pady=20)
        
        # Frame de configuración
        config_frame = ttk.Frame(self.content_frame, style='Main.TFrame')
        config_frame.pack(fill='both', expand=True, padx=50, pady=20)
        
        # Información del usuario
        user_card = ttk.LabelFrame(config_frame,
                                   text="Información del Usuario",
                                   style='Card.TLabelframe',
                                   padding=20)
        user_card.pack(fill='x', pady=10)
        
        if self.auth_service.current_user:
            user = self.auth_service.current_user
            info_text = f"""
Usuario: {user.username}
Nombre: {user.nombre}
Email: {user.email or 'No especificado'}
Rol: {user.rol.upper()}
Estado: {'Activo' if user.activo else 'Inactivo'}
            """
            ttk.Label(user_card,
                     text=info_text,
                     style='Info.TLabel',
                     justify='left').pack(pady=5)
        
        # Configuración de la aplicación
        app_card = ttk.LabelFrame(config_frame,
                                  text="Configuración de la Aplicación",
                                  style='Card.TLabelframe',
                                  padding=20)
        app_card.pack(fill='x', pady=10)
        
        ttk.Label(app_card,
                 text="Configuración avanzada disponible en config.json",
                 style='Info.TLabel').pack(pady=5)
        
        btn_frame = ttk.Frame(app_card, style='Card.TFrame')
        btn_frame.pack(pady=10)
        
        ttk.Button(btn_frame,
                  text="🔄 Cambiar Contraseña",
                  command=self._change_password_dialog,
                  style='Primary.TButton').pack(side='left', padx=5)
        
        ttk.Button(btn_frame,
                  text="💾 Respaldar Base de Datos",
                  command=self._backup_database,
                  style='Success.TButton').pack(side='left', padx=5)
        
        # Información del sistema
        system_card = ttk.LabelFrame(config_frame,
                                     text="Información del Sistema",
                                     style='Card.TLabelframe',
                                     padding=20)
        system_card.pack(fill='x', pady=10)
        
        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'heladeria.db')
        system_info = f"""
Versión: 2.0.0
Base de datos: {db_path}
Directorio de datos: {self.app_data_dir}
Python: {sys.version.split()[0]}
        """
        ttk.Label(system_card,
                 text=system_info,
                 style='Info.TLabel',
                 justify='left').pack(pady=5)
    
    def _change_password_dialog(self):
        """Diálogo para cambiar contraseña"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Cambiar Contraseña")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        form_frame = ttk.Frame(dialog, padding=20)
        form_frame.pack(fill='both', expand=True)
        
        # Contraseña actual
        ttk.Label(form_frame, text="Contraseña actual:").pack(pady=5)
        current_pwd = ttk.Entry(form_frame, show='●', width=30)
        current_pwd.pack(pady=5)
        
        # Nueva contraseña
        ttk.Label(form_frame, text="Nueva contraseña:").pack(pady=5)
        new_pwd = ttk.Entry(form_frame, show='●', width=30)
        new_pwd.pack(pady=5)
        
        # Confirmar contraseña
        ttk.Label(form_frame, text="Confirmar contraseña:").pack(pady=5)
        confirm_pwd = ttk.Entry(form_frame, show='●', width=30)
        confirm_pwd.pack(pady=5)
        
        def change_password():
            current = current_pwd.get()
            new = new_pwd.get()
            confirm = confirm_pwd.get()
            
            if not all([current, new, confirm]):
                messagebox.showerror("Error", "Todos los campos son obligatorios")
                return
            
            if new != confirm:
                messagebox.showerror("Error", "Las contraseñas no coinciden")
                return
            
            if len(new) < 6:
                messagebox.showerror("Error", "La contraseña debe tener al menos 6 caracteres")
                return
            
            # Verificar contraseña actual
            user = self.auth_service.current_user
            if not user.verificar_password(current):
                messagebox.showerror("Error", "Contraseña actual incorrecta")
                return
            
            # Cambiar contraseña
            from models import Usuario
            new_hash = Usuario.hash_password(new)
            success = self.db.execute_query(
                "UPDATE usuarios SET password_hash = ? WHERE id = ?",
                (new_hash, user.id)
            )
            
            if success:
                messagebox.showinfo("Éxito", "Contraseña cambiada correctamente")
                dialog.destroy()
            else:
                messagebox.showerror("Error", "No se pudo cambiar la contraseña")
        
        # Botones
        btn_frame = ttk.Frame(form_frame)
        btn_frame.pack(pady=20)
        
        ttk.Button(btn_frame, text="Cambiar", command=change_password, style='Success.TButton').pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=dialog.destroy, style='Danger.TButton').pack(side='left', padx=5)
    
    def _backup_database(self):
        """Respalda la base de datos"""
        try:
            import shutil
            
            db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'heladeria.db')
            backup_dir = os.path.join(self.app_data_dir, 'backups')
            ensure_directory(backup_dir)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = os.path.join(backup_dir, f'heladeria_backup_{timestamp}.db')
            
            shutil.copy2(db_path, backup_path)
            
            messagebox.showinfo("Éxito", f"Respaldo creado correctamente:\n{backup_path}")
            logger.info(f"Base de datos respaldada: {backup_path}")
            
        except Exception as e:
            logger.error(f"Error al respaldar: {e}")
            messagebox.showerror("Error", f"Error al crear respaldo: {str(e)}")
    
    def show_users_manager(self):
        """Muestra la pantalla de gestión de usuarios con paginación"""
        self.clear_content()
        self.current_screen = "users"
        
        # Variables de paginación
        if not hasattr(self, 'users_current_page'):
            self.users_current_page = 1
        if not hasattr(self, 'users_items_per_page'):
            self.users_items_per_page = 25
        if not hasattr(self, 'users_search_term'):
            self.users_search_term = ''
        
        # Frame principal
        main_frame = ttk.Frame(self.content_frame)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill='x', pady=(0, 20))
        
        ttk.Label(
            header_frame,
            text="👥 Gestión de Usuarios",
            font=('Segoe UI', 18, 'bold')
        ).pack(side='left')
        
        # Botones de acción
        actions_frame = ttk.Frame(header_frame)
        actions_frame.pack(side='right')
        
        ttk.Button(
            actions_frame,
            text="➕ Nuevo Usuario",
            style='Primary.TButton',
            command=self._new_user_dialog
        ).pack(side='left', padx=5)
        
        ttk.Button(
            actions_frame,
            text="🔄 Actualizar",
            command=lambda: self._load_users_paginated()
        ).pack(side='left', padx=5)
        
        # Búsqueda
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(search_frame, text="Buscar:", font=('Segoe UI', 10)).pack(side='left', padx=(0, 10))
        
        self.users_search_var = tk.StringVar(value=self.users_search_term)
        search_entry = ttk.Entry(search_frame, textvariable=self.users_search_var, width=40)
        search_entry.pack(side='left', padx=(0, 10))
        search_entry.bind('<Return>', lambda e: self._load_users_paginated())
        
        ttk.Button(
            search_frame,
            text="🔍 Buscar",
            command=self._load_users_paginated
        ).pack(side='left')
        
        # Frame del TreeView
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill='both', expand=True, pady=(0, 10))
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        vsb.pack(side='right', fill='y')
        
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        hsb.pack(side='bottom', fill='x')
        
        # TreeView
        columns = ('ID', 'Usuario', 'Nombre', 'Email', 'Rol', 'Estado', 'Fecha Creación')
        self.users_tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show='headings',
            style='Modern.Treeview',
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set
        )
        
        vsb.config(command=self.users_tree.yview)
        hsb.config(command=self.users_tree.xview)
        
        # Configurar columnas
        self.users_tree.heading('ID', text='ID')
        self.users_tree.heading('Usuario', text='Usuario')
        self.users_tree.heading('Nombre', text='Nombre Completo')
        self.users_tree.heading('Email', text='Email')
        self.users_tree.heading('Rol', text='Rol')
        self.users_tree.heading('Estado', text='Estado')
        self.users_tree.heading('Fecha Creación', text='Fecha Creación')
        
        self.users_tree.column('ID', width=50, anchor='center')
        self.users_tree.column('Usuario', width=120)
        self.users_tree.column('Nombre', width=200)
        self.users_tree.column('Email', width=200)
        self.users_tree.column('Rol', width=100, anchor='center')
        self.users_tree.column('Estado', width=80, anchor='center')
        self.users_tree.column('Fecha Creación', width=150, anchor='center')
        
        self.users_tree.pack(fill='both', expand=True)
        
        # Frame de paginación
        pagination_frame = ttk.Frame(main_frame)
        pagination_frame.pack(fill='x', pady=(10, 0))
        
        # Tamaño de página
        size_frame = ttk.Frame(pagination_frame)
        size_frame.pack(side='left')
        
        ttk.Label(size_frame, text="Registros por página:").pack(side='left', padx=(0, 5))
        
        self.users_page_size_var = tk.StringVar(value=str(self.users_items_per_page))
        page_size_combo = ttk.Combobox(
            size_frame,
            textvariable=self.users_page_size_var,
            values=['25', '50', '100', '200'],
            width=8,
            state='readonly'
        )
        page_size_combo.pack(side='left')
        page_size_combo.bind('<<ComboboxSelected>>', self._users_change_page_size)
        
        # Controles de navegación
        nav_frame = ttk.Frame(pagination_frame)
        nav_frame.pack(side='right')
        
        self.users_prev_btn = ttk.Button(
            nav_frame,
            text="◀ Anterior",
            command=self._users_prev_page,
            state='disabled'
        )
        self.users_prev_btn.pack(side='left', padx=5)
        
        self.users_page_label = ttk.Label(
            nav_frame,
            text="Página 1 de 1",
            font=('Segoe UI', 10)
        )
        self.users_page_label.pack(side='left', padx=10)
        
        self.users_next_btn = ttk.Button(
            nav_frame,
            text="Siguiente ▶",
            command=self._users_next_page,
            state='disabled'
        )
        self.users_next_btn.pack(side='left', padx=5)
        
        # Botones de acciones
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill='x', pady=(10, 0))
        
        ttk.Button(
            buttons_frame,
            text="✏️ Editar Usuario",
            style='Primary.TButton',
            command=self._edit_user_dialog
        ).pack(side='left', padx=5)
        
        ttk.Button(
            buttons_frame,
            text="🔑 Resetear Contraseña",
            command=self._reset_user_password
        ).pack(side='left', padx=5)
        
        ttk.Button(
            buttons_frame,
            text="🗑️ Eliminar Usuario",
            style='Danger.TButton',
            command=self._delete_user
        ).pack(side='left', padx=5)
        
        # Cargar datos
        self._load_users_paginated()
    
    def _load_users_paginated(self):
        """Carga usuarios con paginación"""
        try:
            # Actualizar término de búsqueda
            if hasattr(self, 'users_search_var'):
                self.users_search_term = self.users_search_var.get().strip()
            
            # Limpiar TreeView
            for item in self.users_tree.get_children():
                self.users_tree.delete(item)
            
            # Obtener todos los usuarios
            query = """
                SELECT id, username, nombre, email, rol, activo, fecha_creacion
                FROM usuarios
                WHERE 1=1
            """
            params = []
            
            # Aplicar filtro de búsqueda
            if self.users_search_term:
                query += " AND (username LIKE ? OR nombre LIKE ? OR email LIKE ?)"
                search_param = f"%{self.users_search_term}%"
                params.extend([search_param, search_param, search_param])
            
            query += " ORDER BY id DESC"
            
            usuarios = self.db.fetch_all(query, tuple(params))
            
            # Calcular paginación
            total_usuarios = len(usuarios)
            total_pages = (total_usuarios + self.users_items_per_page - 1) // self.users_items_per_page
            
            if total_pages == 0:
                total_pages = 1
            
            # Ajustar página actual si es necesario
            if self.users_current_page > total_pages:
                self.users_current_page = total_pages
            
            # Calcular índices
            start_idx = (self.users_current_page - 1) * self.users_items_per_page
            end_idx = start_idx + self.users_items_per_page
            
            # Obtener usuarios de la página actual
            usuarios_pagina = usuarios[start_idx:end_idx]
            
            # Insertar en TreeView
            for user in usuarios_pagina:
                estado = "Activo" if user['activo'] else "Inactivo"
                
                # Formatear rol
                rol_display = {
                    'admin': 'Administrador',
                    'vendedor': 'Vendedor',
                    'supervisor': 'Supervisor'
                }.get(user['rol'], user['rol'])
                
                # Formatear fecha
                fecha = user['fecha_creacion'][:10] if user['fecha_creacion'] else ''
                
                self.users_tree.insert('', 'end', values=(
                    user['id'],
                    user['username'],
                    user['nombre'],
                    user['email'] or '',
                    rol_display,
                    estado,
                    fecha
                ))
            
            # Actualizar controles de paginación
            self.users_page_label.config(
                text=f"Página {self.users_current_page} de {total_pages} ({total_usuarios} registros)"
            )
            
            # Actualizar estado de botones
            self.users_prev_btn.config(state='normal' if self.users_current_page > 1 else 'disabled')
            self.users_next_btn.config(state='normal' if self.users_current_page < total_pages else 'disabled')
            
        except Exception as e:
            logger.error(f"Error al cargar usuarios: {e}")
            messagebox.showerror("Error", f"Error al cargar usuarios: {str(e)}")
    
    def _users_prev_page(self):
        """Ir a la página anterior"""
        if self.users_current_page > 1:
            self.users_current_page -= 1
            self._load_users_paginated()
    
    def _users_next_page(self):
        """Ir a la página siguiente"""
        self.users_current_page += 1
        self._load_users_paginated()
    
    def _users_change_page_size(self, event=None):
        """Cambiar tamaño de página"""
        try:
            self.users_items_per_page = int(self.users_page_size_var.get())
            self.users_current_page = 1  # Volver a la primera página
            self._load_users_paginated()
        except Exception as e:
            logger.error(f"Error al cambiar tamaño de página: {e}")
    
    def _new_user_dialog(self):
        """Diálogo para crear un nuevo usuario"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Nuevo Usuario")
        dialog.geometry("500x650")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Centrar diálogo
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (650 // 2)
        dialog.geometry(f"500x650+{x}+{y}")
        
        # Frame principal
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        # Título
        ttk.Label(
            main_frame,
            text="➕ Crear Nuevo Usuario",
            font=('Segoe UI', 14, 'bold')
        ).pack(pady=(0, 20))
        
        # Frame de campos
        fields_frame = ttk.Frame(main_frame)
        fields_frame.pack(fill='both', expand=True)
        
        # Usuario
        ttk.Label(fields_frame, text="Usuario (username):*", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        username_var = tk.StringVar()
        username_entry = ttk.Entry(fields_frame, textvariable=username_var, width=40)
        username_entry.pack(fill='x', pady=(0, 15))
        username_entry.focus()
        
        # Nombre
        ttk.Label(fields_frame, text="Nombre:*", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        nombre_var = tk.StringVar()
        ttk.Entry(fields_frame, textvariable=nombre_var, width=40).pack(fill='x', pady=(0, 15))
        
        # Apellido
        ttk.Label(fields_frame, text="Apellido:", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        apellido_var = tk.StringVar()
        ttk.Entry(fields_frame, textvariable=apellido_var, width=40).pack(fill='x', pady=(0, 15))
        
        # Email
        ttk.Label(fields_frame, text="Email:", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        email_var = tk.StringVar()
        ttk.Entry(fields_frame, textvariable=email_var, width=40).pack(fill='x', pady=(0, 15))
        
        # Contraseña
        ttk.Label(fields_frame, text="Contraseña:*", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        password_var = tk.StringVar()
        ttk.Entry(fields_frame, textvariable=password_var, show='•', width=40).pack(fill='x', pady=(0, 15))
        
        # Confirmar contraseña
        ttk.Label(fields_frame, text="Confirmar Contraseña:*", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        confirm_password_var = tk.StringVar()
        ttk.Entry(fields_frame, textvariable=confirm_password_var, show='•', width=40).pack(fill='x', pady=(0, 15))
        
        # Rol
        ttk.Label(fields_frame, text="Rol:*", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        rol_var = tk.StringVar(value='cajero')
        rol_combo = ttk.Combobox(
            fields_frame,
            textvariable=rol_var,
            values=['admin', 'supervisor', 'cajero'],
            state='readonly',
            width=38
        )
        rol_combo.pack(fill='x', pady=(0, 15))
        
        # Nota
        ttk.Label(
            fields_frame,
            text="* Campos obligatorios",
            font=('Segoe UI', 9, 'italic'),
            foreground='gray'
        ).pack(anchor='w', pady=(10, 0))
        
        # Función para guardar
        def save_user():
            # Validaciones
            username = username_var.get().strip()
            nombre = nombre_var.get().strip()
            apellido = apellido_var.get().strip()
            email = email_var.get().strip()
            password = password_var.get()
            confirm_password = confirm_password_var.get()
            rol = rol_var.get()
            
            if not username:
                messagebox.showerror("Error", "El nombre de usuario es obligatorio")
                username_entry.focus()
                return
            
            if len(username) < 3:
                messagebox.showerror("Error", "El usuario debe tener al menos 3 caracteres")
                username_entry.focus()
                return
            
            if not nombre:
                messagebox.showerror("Error", "El nombre es obligatorio")
                return
            
            if not password:
                messagebox.showerror("Error", "La contraseña es obligatoria")
                return
            
            if len(password) < 6:
                messagebox.showerror("Error", "La contraseña debe tener al menos 6 caracteres")
                return
            
            if password != confirm_password:
                messagebox.showerror("Error", "Las contraseñas no coinciden")
                return
            
            # Validar email si se proporciona
            if email:
                import re
                email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                if not re.match(email_pattern, email):
                    messagebox.showerror("Error", "El formato del email no es válido")
                    return
            
            try:
                # Verificar si el usuario ya existe
                existing = self.db.fetch_one(
                    "SELECT id FROM usuarios WHERE username = ?",
                    (username,)
                )
                
                if existing:
                    messagebox.showerror("Error", "El nombre de usuario ya existe")
                    username_entry.focus()
                    return
                
                # Hash de la contraseña
                import bcrypt
                password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                
                # Crear usuario
                success = self.db.execute_query("""
                    INSERT INTO usuarios (username, password_hash, nombre, apellido, email, rol, activo)
                    VALUES (?, ?, ?, ?, ?, ?, 1)
                """, (username, password_hash, nombre, apellido if apellido else None, email if email else None, rol))
                
                if not success:
                    messagebox.showerror("Error", "No se pudo crear el usuario. Verifique los datos.")
                    logger.error(f"Fallo al crear usuario: {username}")
                    return
                
                messagebox.showinfo("Éxito", f"Usuario '{username}' creado correctamente")
                logger.info(f"Nuevo usuario creado: {username}")
                
                dialog.destroy()
                self._load_users_paginated()
                
            except Exception as e:
                logger.error(f"Error al crear usuario: {e}")
                messagebox.showerror("Error", f"Error al crear usuario: {str(e)}")
        
        # Botones
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill='x', pady=(20, 0))
        
        ttk.Button(
            buttons_frame,
            text="💾 Guardar",
            style='Primary.TButton',
            command=save_user
        ).pack(side='left', padx=5)
        
        ttk.Button(
            buttons_frame,
            text="❌ Cancelar",
            command=dialog.destroy
        ).pack(side='left', padx=5)
    
    def _edit_user_dialog(self):
        """Diálogo para editar un usuario"""
        selection = self.users_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Por favor, selecciona un usuario para editar")
            return
        
        item = self.users_tree.item(selection[0])
        values = item['values']
        user_id = values[0]
        
        # Obtener datos del usuario
        user_data = self.db.fetch_one(
            "SELECT id, username, nombre, email, rol, activo FROM usuarios WHERE id = ?",
            (user_id,)
        )
        
        if not user_data:
            messagebox.showerror("Error", "No se encontró el usuario")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Editar Usuario")
        dialog.geometry("500x520")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Centrar diálogo
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (520 // 2)
        dialog.geometry(f"500x520+{x}+{y}")
        
        # Frame principal
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        # Título
        ttk.Label(
            main_frame,
            text=f"✏️ Editar Usuario: {user_data['username']}",
            font=('Segoe UI', 14, 'bold')
        ).pack(pady=(0, 20))
        
        # Frame de campos
        fields_frame = ttk.Frame(main_frame)
        fields_frame.pack(fill='both', expand=True)
        
        # Usuario (solo lectura)
        ttk.Label(fields_frame, text="Usuario (username):", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        username_label = ttk.Label(fields_frame, text=user_data['username'], font=('Segoe UI', 10, 'bold'))
        username_label.pack(anchor='w', pady=(0, 15))
        
        # Nombre completo
        ttk.Label(fields_frame, text="Nombre Completo:*", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        nombre_var = tk.StringVar(value=user_data['nombre'])
        nombre_entry = ttk.Entry(fields_frame, textvariable=nombre_var, width=40)
        nombre_entry.pack(fill='x', pady=(0, 15))
        nombre_entry.focus()
        
        # Email
        ttk.Label(fields_frame, text="Email:", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        email_var = tk.StringVar(value=user_data['email'] if user_data['email'] else '')
        ttk.Entry(fields_frame, textvariable=email_var, width=40).pack(fill='x', pady=(0, 15))
        
        # Rol
        ttk.Label(fields_frame, text="Rol:*", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        rol_var = tk.StringVar(value=user_data['rol'])
        rol_combo = ttk.Combobox(
            fields_frame,
            textvariable=rol_var,
            values=['admin', 'supervisor', 'cajero'],
            state='readonly',
            width=38
        )
        rol_combo.pack(fill='x', pady=(0, 15))
        
        # Estado
        ttk.Label(fields_frame, text="Estado:", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        activo_var = tk.BooleanVar(value=bool(user_data['activo']))
        ttk.Checkbutton(fields_frame, text="Usuario activo", variable=activo_var).pack(anchor='w', pady=(0, 15))
        
        # Función para guardar
        def save_user():
            nombre = nombre_var.get().strip()
            email = email_var.get().strip()
            rol = rol_var.get()
            activo = 1 if activo_var.get() else 0
            
            if not nombre:
                messagebox.showerror("Error", "El nombre completo es obligatorio")
                nombre_entry.focus()
                return
            
            # Validar email si se proporciona
            if email:
                import re
                email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                if not re.match(email_pattern, email):
                    messagebox.showerror("Error", "El formato del email no es válido")
                    return
            
            try:
                # Actualizar usuario
                self.db.execute_query("""
                    UPDATE usuarios
                    SET nombre = ?, email = ?, rol = ?, activo = ?
                    WHERE id = ?
                """, (nombre, email if email else None, rol, activo, user_id))
                
                messagebox.showinfo("Éxito", "Usuario actualizado correctamente")
                logger.info(f"Usuario actualizado: {user_data['username']}")
                
                dialog.destroy()
                self._load_users_paginated()
                
            except Exception as e:
                logger.error(f"Error al actualizar usuario: {e}")
                messagebox.showerror("Error", f"Error al actualizar usuario: {str(e)}")
        
        # Botones
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill='x', pady=(20, 0))
        
        ttk.Button(
            buttons_frame,
            text="💾 Guardar Cambios",
            style='Primary.TButton',
            command=save_user
        ).pack(side='left', padx=5)
        
        ttk.Button(
            buttons_frame,
            text="❌ Cancelar",
            command=dialog.destroy
        ).pack(side='left', padx=5)
    
    def _delete_user(self):
        """Eliminar (desactivar) un usuario"""
        selection = self.users_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Por favor, selecciona un usuario para eliminar")
            return
        
        item = self.users_tree.item(selection[0])
        values = item['values']
        user_id = values[0]
        username = values[1]
        
        # No permitir eliminar al usuario actual
        if self.auth_service.current_user and self.auth_service.current_user.id == user_id:
            messagebox.showerror("Error", "No puedes eliminar tu propia cuenta")
            return
        
        # Confirmación
        if not messagebox.askyesno(
            "Confirmar Eliminación",
            f"¿Estás seguro de que deseas desactivar el usuario '{username}'?\n\n"
            "El usuario no será eliminado permanentemente, solo se desactivará."
        ):
            return
        
        try:
            # Desactivar usuario
            self.db.execute_query(
                "UPDATE usuarios SET activo = 0 WHERE id = ?",
                (user_id,)
            )
            
            messagebox.showinfo("Éxito", f"Usuario '{username}' desactivado correctamente")
            logger.info(f"Usuario desactivado: {username}")
            
            self._load_users_paginated()
            
        except Exception as e:
            logger.error(f"Error al eliminar usuario: {e}")
            messagebox.showerror("Error", f"Error al eliminar usuario: {str(e)}")
    
    def _reset_user_password(self):
        """Resetear contraseña de un usuario"""
        selection = self.users_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Por favor, selecciona un usuario para resetear su contraseña")
            return
        
        item = self.users_tree.item(selection[0])
        values = item['values']
        user_id = values[0]
        username = values[1]
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Resetear Contraseña")
        dialog.geometry("450x300")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Centrar diálogo
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (dialog.winfo_screenheight() // 2) - (300 // 2)
        dialog.geometry(f"450x300+{x}+{y}")
        
        # Frame principal
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        # Título
        ttk.Label(
            main_frame,
            text=f"🔑 Resetear Contraseña: {username}",
            font=('Segoe UI', 14, 'bold')
        ).pack(pady=(0, 20))
        
        # Campos
        ttk.Label(main_frame, text="Nueva Contraseña:*", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        password_var = tk.StringVar()
        password_entry = ttk.Entry(main_frame, textvariable=password_var, show='•', width=40)
        password_entry.pack(fill='x', pady=(0, 15))
        password_entry.focus()
        
        ttk.Label(main_frame, text="Confirmar Contraseña:*", font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))
        confirm_password_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=confirm_password_var, show='•', width=40).pack(fill='x', pady=(0, 15))
        
        # Función para guardar
        def save_password():
            password = password_var.get()
            confirm_password = confirm_password_var.get()
            
            if not password:
                messagebox.showerror("Error", "La contraseña es obligatoria")
                password_entry.focus()
                return
            
            if len(password) < 6:
                messagebox.showerror("Error", "La contraseña debe tener al menos 6 caracteres")
                password_entry.focus()
                return
            
            if password != confirm_password:
                messagebox.showerror("Error", "Las contraseñas no coinciden")
                return
            
            try:
                # Hash de la contraseña
                import bcrypt
                password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                
                # Actualizar contraseña
                self.db.execute_query(
                    "UPDATE usuarios SET password_hash = ? WHERE id = ?",
                    (password_hash, user_id)
                )
                
                messagebox.showinfo("Éxito", f"Contraseña de '{username}' reseteada correctamente")
                logger.info(f"Contraseña reseteada para usuario: {username}")
                
                dialog.destroy()
                
            except Exception as e:
                logger.error(f"Error al resetear contraseña: {e}")
                messagebox.showerror("Error", f"Error al resetear contraseña: {str(e)}")
        
        # Botones
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill='x', pady=(20, 0))
        
        ttk.Button(
            buttons_frame,
            text="💾 Guardar",
            style='Primary.TButton',
            command=save_password
        ).pack(side='left', padx=5)
        
        ttk.Button(
            buttons_frame,
            text="❌ Cancelar",
            command=dialog.destroy
        ).pack(side='left', padx=5)
    
    def show_help(self):
        """Muestra la ayuda"""
        self.clear_content()
        self.current_screen = "help"
        
        # Título
        title = ttk.Label(self.content_frame,
                         text="Ayuda y Documentación",
                         font=(self.config.get('fonts', {}).get('family', 'Segoe UI'), 20, 'bold'),
                         background=self.colors.get('bg_light', '#ECF0F1'),
                         foreground=self.colors.get('primary', '#2C3E50'))
        title.pack(pady=20)
        
        # Frame de ayuda
        help_frame = ttk.Frame(self.content_frame, style='Main.TFrame')
        help_frame.pack(fill='both', expand=True, padx=50, pady=20)
        
        # Notebook para pestañas
        notebook = ttk.Notebook(help_frame)
        notebook.pack(fill='both', expand=True)
        
        # Pestaña: Inicio Rápido
        quick_start_frame = ttk.Frame(notebook, padding=20)
        notebook.add(quick_start_frame, text="Inicio Rápido")
        
        quick_start_text = """
BIENVENIDO AL SISTEMA TPV - SISTEMA VENTAS PROFESIONAL

Este sistema le permite gestionar las ventas, inventario y reportes de su heladería
de manera eficiente y profesional.

INICIO RÁPIDO:

1. LOGIN
   - Usuario por defecto: admin
   - Contraseña: admin123
   - Cambie la contraseña en Configuración

2. DASHBOARD
   - Visualice estadísticas de ventas
   - Vea productos bajo stock
   - Analice ventas de los últimos 7 días

3. NUEVA VENTA
   - Busque productos disponibles
   - Haga doble clic para agregar al carrito
   - Procese la venta con el botón "Procesar Venta"

4. GESTIÓN DE PRODUCTOS
   - Cree nuevos productos
   - Edite información de productos existentes
   - Ajuste stock según necesidad

5. REPORTES
   - Genere reportes de ventas diarias y mensuales
   - Consulte estado del inventario
   - Revise análisis financiero
        """
        
        text1 = tk.Text(quick_start_frame, wrap='word', font=('Segoe UI', 10))
        text1.pack(fill='both', expand=True)
        text1.insert('1.0', quick_start_text)
        text1.config(state='disabled', bg='white')
        
        # Pestaña: Funciones
        functions_frame = ttk.Frame(notebook, padding=20)
        notebook.add(functions_frame, text="Funciones")
        
        functions_text = """
FUNCIONES DEL SISTEMA

🏠 INICIO (Dashboard)
   - Estadísticas en tiempo real
   - Gráficos de ventas
   - Alertas de stock bajo

🛒 NUEVA VENTA
   - Búsqueda rápida de productos
   - Gestión de carrito
   - Procesamiento de ventas

📊 HISTORIAL DE VENTAS
   - Consulta de ventas por fecha
   - Detalles de cada venta
   - Impresión de tickets

📦 GESTIÓN DE PRODUCTOS
   - Alta, baja y modificación de productos
   - Control de stock
   - Categorización de productos

📈 REPORTES
   - Reportes de ventas (día/mes)
   - Inventario completo
   - Productos bajo stock
   - Análisis financiero

👥 GESTIÓN DE USUARIOS
   - Crear y editar usuarios
   - Resetear contraseñas
   - Gestionar roles y permisos
   - Activar/desactivar usuarios

⚙️ CONFIGURACIÓN
   - Cambio de contraseña
   - Respaldo de base de datos
   - Información del sistema

❓ AYUDA
   - Documentación completa
   - Guía de uso
   - Atajos de teclado
        """
        
        text2 = tk.Text(functions_frame, wrap='word', font=('Segoe UI', 10))
        text2.pack(fill='both', expand=True)
        text2.insert('1.0', functions_text)
        text2.config(state='disabled', bg='white')
        
        # Pestaña: Atajos de Teclado
        shortcuts_frame = ttk.Frame(notebook, padding=20)
        notebook.add(shortcuts_frame, text="Atajos")
        
        shortcuts_text = """
ATAJOS DE TECLADO

GENERALES:
   F1  - Ayuda
   F5  - Actualizar pantalla actual
   Esc - Cerrar diálogo actual

NAVEGACIÓN:
   Ctrl+1 - Ir a Dashboard
   Ctrl+2 - Nueva Venta
   Ctrl+3 - Historial de Ventas
   Ctrl+4 - Gestión de Productos
   Ctrl+5 - Reportes
   Ctrl+6 - Configuración
   Ctrl+7 - Gestión de Usuarios

PUNTO DE VENTA:
   F2  - Buscar producto
   F3  - Agregar al carrito
   F4  - Procesar venta
   Del - Quitar del carrito

FORMULARIOS:
   Enter     - Confirmar/Aceptar
   Esc       - Cancelar
   Tab       - Siguiente campo
   Shift+Tab - Campo anterior

REPORTES:
   Ctrl+S - Guardar reporte
   Ctrl+P - Imprimir
        """
        
        text3 = tk.Text(shortcuts_frame, wrap='word', font=('Courier New', 10))
        text3.pack(fill='both', expand=True)
        text3.insert('1.0', shortcuts_text)
        text3.config(state='disabled', bg='white')
        
        # Pestaña: Soporte
        support_frame = ttk.Frame(notebook, padding=20)
        notebook.add(support_frame, text="Soporte")
        
        support_text = """
SOPORTE TÉCNICO

Para obtener ayuda adicional o reportar problemas:

📧 Email: jesusjcopes@gmail.com
📞 Teléfono: +54 340815670623
💬 Chat: Disponible en horario de oficina

HORARIO DE ATENCIÓN:
   Lunes a Viernes: 9:00 AM - 6:00 PM
   Sábados: 9:00 AM - 1:00 PM
   Domingos y feriados: Cerrado

INFORMACIÓN DE LA VERSIÓN:
   Versión: 2.0.0
   Fecha: Octubre 2025
   Desarrollador: Jesús Copes
   Licencia: MIT

NOTAS DE LA VERSIÓN:
   - Interfaz modernizada
   - Mejoras en rendimiento
   - Nuevos reportes
   - Gestión avanzada de inventario
   - Respaldos automáticos

Para más información, consulte el archivo README.md
y MANUAL.md en el directorio de la aplicación.
        """
        
        text4 = tk.Text(support_frame, wrap='word', font=('Segoe UI', 10))
        text4.pack(fill='both', expand=True)
        text4.insert('1.0', support_text)
        text4.config(state='disabled', bg='white')
        
        # Botón para abrir manual completo
        ttk.Button(help_frame,
                  text="📖 Abrir Manual Completo (MANUAL.md)",
                  command=self._open_manual,
                  style='Primary.TButton').pack(pady=10)
    
    def _open_manual(self):
        """Abre el archivo MANUAL.md"""
        manual_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'MANUAL.md')
        if os.path.exists(manual_path):
            import webbrowser
            webbrowser.open(manual_path)
        else:
            messagebox.showinfo("Información", "Archivo MANUAL.md no encontrado")
    
    def run(self):
        """Inicia la aplicación"""
        self.root.mainloop()


def main():
    """Función principal"""
    try:
        # Crear ventana principal
        root = tk.Tk()
        
        # Crear aplicación
        app = ModernTPV(root)
        
        # Ejecutar
        app.run()
        
    except Exception as e:
        logger.critical(f"Error fatal en la aplicación: {e}", exc_info=True)
        messagebox.showerror("Error Fatal",
                           f"Se produjo un error fatal:\n{str(e)}\n\nPor favor contacte al soporte técnico.")
        sys.exit(1)


if __name__ == "__main__":
    main()
